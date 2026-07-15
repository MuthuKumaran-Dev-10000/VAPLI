import os
import re
import sys
import json
import smtplib
import threading
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Attempt to load reportlab and openpyxl
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Safely parses an ISO 8601 datetime string, supporting Z offset and variable subseconds
def parse_iso_datetime(iso_str):
    if not iso_str:
        return None
    try:
        clean_str = iso_str.rstrip("Z")
        if "." in clean_str:
            parts = clean_str.split(".")
            if len(parts[1]) > 6:
                clean_str = parts[0] + "." + parts[1][:6]
        return datetime.fromisoformat(clean_str)
    except Exception as e:
        return None

# Extracts email addresses from Firebase structures (lists, dicts, or strings)
def parse_email_list(receivers_val):
    if not receivers_val:
        return []
    if isinstance(receivers_val, list):
        return [str(r).strip() for r in receivers_val if r and str(r).strip()]
    if isinstance(receivers_val, dict):
        return [str(v).strip() for v in receivers_val.values() if v and str(v).strip()]
    if isinstance(receivers_val, str):
        return [r.strip() for r in receivers_val.split(",") if r.strip()]
    return []

# Safely converts Firebase database lists/arrays into dictionaries keyed by ID or index
def normalize_to_dict(data):
    if not data:
        return {}
    if isinstance(data, dict):
        return data
    if isinstance(data, list):
        result = {}
        for idx, item in enumerate(data):
            if item is None:
                continue
            if isinstance(item, dict) and "id" in item:
                result[str(item["id"])] = item
            else:
                result[str(idx)] = item
        return result
    return {}

# Global Context Variables
CONFIG = {
    "db_url": "",
    "web_api_key": "",
    "smtp_server": "smtp.gmail.com",
    "smtp_port": 587,
    "smtp_email": "",
    "smtp_password": "",
    "server_host": "0.0.0.0",
    "server_port": 8080,
    "client_id": "dummy_client_id",
    "env_mode": "production", # 'production' or 'development'
    "external_url": "",
    # Schedule configuration
    "ast_reminder_hour": 15,       # 3 PM Arabia Standard Time (UTC+3)
    "ast_reminder_minute": 0,
    "ast_reminder_enabled": True,
    "ist_reminder_hour": 15,       # 3 PM India Standard Time (UTC+5:30)
    "ist_reminder_minute": 0,
    "ist_reminder_enabled": True,
    "syd_reminder_enabled": True,   # Keep Sydney schedule active
    "server_start_time": None       # Set at startup
}

def load_env():
    """Loads VAPLI environment variables from parent .env/.env file"""
    env_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".env", ".env"))
    if not os.path.exists(env_path):
        print(f"[-] Env file not found at: {env_path}")
        return

    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            m = re.match(r"^([^=]+)=(.*)$", line)
            if m:
                key = m.group(1).strip()
                val = m.group(2).strip()
                if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
                    val = val[1:-1]
                os.environ[key] = val
                
    # Read from merged environment variables
    if "FIREBASE_DATABASE_URL" in os.environ:
        CONFIG["db_url"] = os.environ["FIREBASE_DATABASE_URL"].rstrip("/")
    if "FIREBASE_WEB_API_KEY" in os.environ:
        CONFIG["web_api_key"] = os.environ["FIREBASE_WEB_API_KEY"]
    if "SMTP_SERVER" in os.environ:
        CONFIG["smtp_server"] = os.environ["SMTP_SERVER"]
    if "SMTP_PORT" in os.environ:
        CONFIG["smtp_port"] = int(os.environ["SMTP_PORT"])
    if "SMTP_EMAIL" in os.environ:
        CONFIG["smtp_email"] = os.environ["SMTP_EMAIL"]
    if "SMTP_PASSWORD" in os.environ:
        CONFIG["smtp_password"] = os.environ["SMTP_PASSWORD"]
    if "SERVER_HOST" in os.environ:
        CONFIG["server_host"] = os.environ["SERVER_HOST"]
    if "SERVER_PORT" in os.environ:
        CONFIG["server_port"] = int(os.environ["SERVER_PORT"])
    elif "PORT" in os.environ:
        CONFIG["server_port"] = int(os.environ["PORT"])
        CONFIG["server_host"] = "0.0.0.0" # Bind to all interfaces when deployed to Render
        
    if "DEFAULT_CLIENT_ID" in os.environ:
        CONFIG["client_id"] = os.environ["DEFAULT_CLIENT_ID"]
    if "DEFAULT_ENV_MODE" in os.environ:
        CONFIG["env_mode"] = os.environ["DEFAULT_ENV_MODE"].lower()
    if "FEEDBACK_BASE_URL" in os.environ:
        CONFIG["external_url"] = os.environ["FEEDBACK_BASE_URL"].rstrip("/")
    elif "SERVER_EXTERNAL_URL" in os.environ:
        CONFIG["external_url"] = os.environ["SERVER_EXTERNAL_URL"].rstrip("/")

    print("[+] Environment loaded successfully.")
    print(f"    Database URL: {CONFIG['db_url']}")
    print(f"    SMTP Account: {CONFIG['smtp_email'] or 'Not Configured'}")
    print(f"    Default Client: {CONFIG['client_id']}")
    print(f"    Mode: {CONFIG['env_mode']}")

def db_ref_path(raw_path):
    """Calculates database path prefix matching DatabaseModeService.dart rules"""
    normalized = raw_path.lstrip("/")
    resolved = normalized
    
    global_paths = {"users", "clients"}
    scoped_paths = {
        "tanks", "tank_tree", "readings", "alerts", "completed_tasks",
        "violations", "settings", "reading_feedback"
    }

    if resolved:
        head = resolved.split("/")[0]
        if head not in global_paths and head in scoped_paths and CONFIG["client_id"]:
            resolved = f"{CONFIG['client_id']}/{resolved}"

    if CONFIG["env_mode"] == "development":
        resolved = f"testDB/{resolved}" if resolved else "testDB"
        
    return resolved

def fetch_db(path):
    """Fetches database node as JSON using REST API"""
    full_path = db_ref_path(path)
    encoded_path = urllib.parse.quote(full_path, safe='/')
    url = f"{CONFIG['db_url']}/{encoded_path}.json"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as res:
            return json.loads(res.read().decode("utf-8"))
    except Exception as e:
        print(f"[-] Database fetch error on {path}: {e}")
        return None

def write_db(path, data, method="PUT"):
    """Writes database node using REST API"""
    full_path = db_ref_path(path)
    encoded_path = urllib.parse.quote(full_path, safe='/')
    url = f"{CONFIG['db_url']}/{encoded_path}.json"
    try:
        payload = json.dumps(data).encode("utf-8")
        req = urllib.request.Request(url, data=payload, method=method)
        req.add_header("Content-Type", "application/json")
        with urllib.request.urlopen(req, timeout=10) as res:
            return json.loads(res.read().decode("utf-8"))
    except Exception as e:
        print(f"[-] Database write error on {path}: {e}")
        return None

# ==============================================================================
# REPORT GENERATORS (PDF, EXCEL)
# ==============================================================================

def format_reading_value(val):
    """Formats a database reading value (string, dict, or number) into a clean string"""
    if val is None:
        return "-"
    if isinstance(val, dict):
        if "left" in val or "right" in val:
            return f"{val.get('left', '')} / {val.get('right', '')}"
        elif "value" in val:
            return str(val.get("value", "-"))
        else:
            return ", ".join(f"{k}: {v}" for k, v in val.items())
    return str(val)

def has_take_action_keyword(val):
    """Checks if a reading value indicates an action is required"""
    if val is None:
        return False
    val_str = str(val).lower()
    return "take action" in val_str or "not ok" in val_str

def get_client_name(client_id):
    """Fetches the client list from database and resolves the user-friendly client name"""
    try:
        url = f"{CONFIG['db_url']}/clients.json"
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=5) as res:
            clients_data = json.loads(res.read().decode("utf-8"))
        clients_dict = normalize_to_dict(clients_data)
        for cid, cinfo in clients_dict.items():
            if isinstance(cinfo, dict) and cinfo.get("db_key") == client_id:
                return cinfo.get("name") or client_id
    except:
        pass
    return client_id.upper().replace("_", " ")

def get_selected_params_for_folder(folder_id, folder_tanks, report_format_configs):
    """Resolves and sorts selected parameters for a folder based on database report configurations"""
    folder_config = report_format_configs.get(folder_id, {}) if report_format_configs else {}
    selected_params_map = folder_config.get("selected_params", {}) if folder_config else {}
    
    discovered = {}
    for tank in folder_tanks:
        props = tank.get("inspection_properties", [])
        if not isinstance(props, list):
            continue
        for prop in props:
            p_type = prop.get("type", "text")
            if p_type == "group":
                continue
            label = prop.get("label") or prop.get("name") or ""
            label = str(label).strip()
            if not label:
                continue
            
            options = prop.get("options", [])
            if isinstance(options, list):
                options = sorted([str(o).strip() for o in options if o is not None])
            else:
                options = []
            options_str = ",".join(options)
            
            db_key = f"{label}_{p_type}_{options_str}"
            fallback_key = f"{label}_{p_type}"
            
            if db_key not in discovered:
                discovered[db_key] = {
                    "db_key": db_key,
                    "fallback_key": fallback_key,
                    "label": label,
                    "type": p_type,
                    "prop_id": prop.get("id", ""),
                    "options": options,
                    "selected": True,
                    "order": 0
                }
                
    items = []
    max_order = 0
    
    for db_key, item in discovered.items():
        selected = True
        order = 0
        
        cfg = None
        if selected_params_map:
            if db_key in selected_params_map:
                cfg = selected_params_map[db_key]
            else:
                # Try fallback matching
                for k, v in selected_params_map.items():
                    if k.startswith(item["fallback_key"]):
                        cfg = v
                        break
        
        if cfg and isinstance(cfg, dict):
            selected = cfg.get("selected", True)
            order = int(cfg.get("order", 0))
            if selected and order > max_order:
                max_order = order
                
        items.append({
            "label": item["label"],
            "prop_id": item["prop_id"],
            "selected": selected,
            "order": order
        })
        
    selected_items = [i for i in items if i["selected"]]
    
    configured = sorted([i for i in selected_items if i["order"] > 0], key=lambda x: x["order"])
    unconfigured = sorted([i for i in selected_items if i["order"] == 0], key=lambda x: x["label"])
    
    next_order = max_order + 1
    for i in unconfigured:
        i["order"] = next_order
        next_order += 1
        
    return configured + unconfigured

def build_pdf_subtable(tanks, params, readings, active_alerts, cell_style, cell_style_bold, header_style):
    """Helper to build a ReportLab Table for a subcategory of tanks"""
    num_params = len(params)
    param_col_width = 370.0 / num_params if num_params > 0 else 370.0
    col_widths = [75, 95] + [param_col_width] * num_params
    
    headers = [
        Paragraph("Date Time", header_style),
        Paragraph("Asset Name", header_style)
    ]
    for p in params:
        headers.append(Paragraph(p["label"], header_style))
        
    table_data = [headers]
    
    table_commands = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F497D')), # Dark Blue header
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]
    
    row_idx = 1
    for tank in tanks:
        tid = tank.get("id")
        tank_code = tank.get("tank_code")
        tank_name = tank.get("tank_name", "Asset")
        display_name = f"{tank_name} ({tank_code})" if tank_code else tank_name
        
        reading = readings.get(tid)
        
        # Determine highest alert severity for row background color
        tank_severity = None
        for a in active_alerts:
            if a.get("tank_id") == tid:
                sev = (a.get("constraint_severity") or a.get("severity") or "warning").lower()
                if sev == "critical":
                    tank_severity = "critical"
                    break
                elif sev == "warning":
                    tank_severity = "warning"
                elif not tank_severity:
                    tank_severity = "info"
                    
        default_row_color = '#FFFFFF' if row_idx % 2 == 1 else '#F8FAFC'
        
        if not reading:
            # Readings not taken
            row = [
                Paragraph("-", cell_style),
                Paragraph(display_name, cell_style_bold),
                Paragraph("------- Readings not taken ------", cell_style_bold)
            ]
            for _ in range(num_params - 1):
                row.append(Paragraph("", cell_style))
                
            table_data.append(row)
            
            # Set row background color: Pending light green
            table_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#C8E6C9')))
            table_commands.append(('SPAN', (2, row_idx), (-1, row_idx)))
            table_commands.append(('ALIGN', (2, row_idx), (-1, row_idx), 'CENTER'))
        else:
            # Readings exist
            captured_at = reading.get("captured_at", "")
            try:
                dt = parse_iso_datetime(captured_at)
                dt_str = dt.strftime("%d/%m/%y %H:%M") if dt else captured_at.split(".")[0].replace("T", " ")[:14]
            except:
                dt_str = captured_at.split(".")[0].replace("T", " ")[:14]
                
            row = [
                Paragraph(dt_str, cell_style),
                Paragraph(display_name, cell_style_bold)
            ]
            
            insp_vals = reading.get("inspection_values", {})
            col_idx = 2
            for p in params:
                p_label = p["label"]
                p_id = p["prop_id"]
                
                val = insp_vals.get(p_label)
                if val is None:
                    val = insp_vals.get(p_id)
                val_str = format_reading_value(val)
                
                # Check for parameter-specific image URL
                p_img = insp_vals.get(f"{p_id}__image_url")
                if not p_img:
                    p_img = insp_vals.get(f"{p_label}__image_url")
                    
                cell_content = val_str
                if p_img:
                    cell_content += f'<br/><a href="{p_img}" color="#1565C0"><u>[Photo]</u></a>'
                    
                # Evaluate violations
                is_violated = False
                cell_severity = None
                
                for a in active_alerts:
                    if a.get("tank_id") == tid:
                        label = a.get("constraint_label") or a.get("param_label") or ""
                        if label.strip().lower() == p_label.strip().lower():
                            is_violated = True
                            cell_severity = (a.get("constraint_severity") or a.get("severity") or "critical").lower()
                            break
                            
                if not is_violated and has_take_action_keyword(val):
                    is_violated = True
                    cell_severity = "critical"
                    
                if is_violated:
                    row.append(Paragraph(cell_content, cell_style_bold))
                    color_hex = '#FFCDD2' if cell_severity == 'critical' else ('#FFF9C4' if cell_severity == 'warning' else '#BBDEFB')
                    table_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor(color_hex)))
                else:
                    row.append(Paragraph(cell_content, cell_style))
                    table_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor(default_row_color)))
                    
                col_idx += 1
                
            table_data.append(row)
            
            if tank_severity:
                color_hex = '#FFCDD2' if tank_severity == 'critical' else ('#FFF9C4' if tank_severity == 'warning' else '#BBDEFB')
                table_commands.append(('BACKGROUND', (0, row_idx), (1, row_idx), colors.HexColor(color_hex)))
            else:
                table_commands.append(('BACKGROUND', (0, row_idx), (1, row_idx), colors.HexColor(default_row_color)))
                
        row_idx += 1
        
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle(table_commands))
    return t

def generate_inspection_pdf(filepath, date_str):
    """Generates today's inspection report PDF, matching the layout strategy of the Flutter app"""
    if not REPORTLAB_AVAILABLE:
        print("[-] reportlab library not available. Skipping PDF generation.")
        return False
        
    try:
        tanks_data = normalize_to_dict(fetch_db("tanks"))
        readings_data = normalize_to_dict(fetch_db("readings"))
        tree_data = normalize_to_dict(fetch_db("tank_tree"))
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        report_format_configs = normalize_to_dict(fetch_db("settings/report_format"))
        
        today_readings = {}
        for r_id, r in readings_data.items():
            try:
                r_date = r.get("captured_at", "").split("T")[0]
                if r_date == date_str:
                    today_readings[r.get("tank_id")] = r
            except:
                continue

        active_alerts = []
        for a_id, a in alerts_data.items():
            if not a.get("resolved", False) and a.get("status", "").lower() != "completed":
                active_alerts.append(a)

        nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
        folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
        folders.sort(key=lambda x: x.get("order", 0))

        def get_folder_tanks(folder_id):
            tank_ids = []
            queue = [folder_id]
            while queue:
                curr = queue.pop(0)
                children = [n for n in nodes if n.get("parent_id") == curr]
                for c in children:
                    if c.get("type") == "leaf" and c.get("tank_id"):
                        tank_ids.append(c.get("tank_id"))
                    elif c.get("type") == "folder":
                        queue.append(c.get("id"))
            return tank_ids

        client_name = get_client_name(CONFIG['client_id'])
        generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")

        total_configured = len(tanks_data)
        inspected_tanks = [tid for tid in tanks_data.keys() if tid in today_readings]
        pending_tanks = [tank for tid, tank in tanks_data.items() if tid not in today_readings]
        pending_tanks.sort(key=lambda x: x.get("tank_name", ""))
        compliance_rate = (len(inspected_tanks) / total_configured * 100) if total_configured > 0 else 0.0

        story = []
        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'DocTitle', parent=styles['Heading1'],
            fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#CB8C3E'),
            alignment=1, spaceAfter=8
        )
        meta_style = ParagraphStyle(
            'DocMeta', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#555555'),
            alignment=1, spaceAfter=15
        )
        section_style = ParagraphStyle(
            'SecTitle', parent=styles['Heading2'],
            fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#1A365D'),
            spaceBefore=10, spaceAfter=6
        )
        subsection_style = ParagraphStyle(
            'SubSecTitle', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=9.5, textColor=colors.HexColor('#2B6CB0'),
            spaceBefore=6, spaceAfter=4
        )
        cell_style = ParagraphStyle(
            'CellText', parent=styles['Normal'],
            fontName='Helvetica', fontSize=7, leading=9, textColor=colors.HexColor('#2D3748')
        )
        cell_style_bold = ParagraphStyle(
            'CellTextBold', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=7, leading=9, textColor=colors.HexColor('#2D3748')
        )
        header_style = ParagraphStyle(
            'HeaderStyle', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#FFFFFF'),
            alignment=1
        )

        # SUMMARY PAGE (Page 1)
        story.append(Paragraph("VAPLI Inspection Telemetry Report", title_style))
        story.append(Paragraph(f"Client: {client_name} | Date: {date_str} | Generated: {generated_at}", meta_style))
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("Professional Inspection Summary", section_style))
        story.append(Spacer(1, 5))
        
        stats_table_data = [
            [Paragraph("<b>Metric</b>", cell_style_bold), Paragraph("<b>Value / Count</b>", cell_style_bold)],
            [Paragraph("Total Configured Assets", cell_style), Paragraph(str(total_configured), cell_style_bold)],
            [Paragraph("Assets with Readings", cell_style), Paragraph(str(len(inspected_tanks)), cell_style_bold)],
            [Paragraph("Assets Pending Readings", cell_style), Paragraph(str(len(pending_tanks)), cell_style_bold)],
            [Paragraph("Compliance Rate", cell_style), Paragraph(f"{compliance_rate:.1f}%", cell_style_bold)],
            [Paragraph("Active Unresolved Alerts", cell_style), Paragraph(str(len(active_alerts)), cell_style_bold)]
        ]
        stats_table = Table(stats_table_data, colWidths=[200, 100])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 15))

        story.append(Paragraph("Assets Pending Inspections (No Readings Today)", ParagraphStyle('PendingTitle', parent=section_style, textColor=colors.HexColor('#9B2C2C'))))
        story.append(Spacer(1, 5))
        
        if not pending_tanks:
            story.append(Paragraph("<b>No assets pending readings today. All recorded.</b>", ParagraphStyle('NoPending', parent=cell_style, textColor=colors.HexColor('#2F855A'))))
        else:
            pending_rows = []
            temp_row = []
            for t in pending_tanks:
                t_display = f"&bull; {t.get('tank_name')} ({t.get('tank_code', '')})"
                temp_row.append(Paragraph(t_display, cell_style))
                if len(temp_row) == 3:
                    pending_rows.append(temp_row)
                    temp_row = []
            if temp_row:
                while len(temp_row) < 3:
                    temp_row.append(Paragraph("", cell_style))
                pending_rows.append(temp_row)
                
            pending_table = Table(pending_rows, colWidths=[180, 180, 180])
            pending_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(pending_table)
            
        story.append(PageBreak())

        # DETAILED TABLES FOR EACH GROUP
        has_detailed_data = False
        for folder in folders:
            tank_ids = get_folder_tanks(folder.get("id"))
            if not tank_ids:
                continue
            
            folder_tanks = [tanks_data[tid] for tid in tank_ids if tid in tanks_data]
            if not folder_tanks:
                continue
                
            selected_params = get_selected_params_for_folder(folder.get("id"), folder_tanks, report_format_configs)
            if not selected_params:
                continue
                
            has_detailed_data = True
            normal_tanks_list = []
            violated_tanks_list = []
            
            for tank in folder_tanks:
                tid = tank.get("id")
                has_active_alert = any(a.get("tank_id") == tid for a in active_alerts)
                if has_active_alert:
                    violated_tanks_list.append(tank)
                else:
                    normal_tanks_list.append(tank)
                    
            folder_name = folder.get("name")
            
            # Normal Tanks Subtable
            if normal_tanks_list:
                story.append(Paragraph(f"{folder_name} — Normal Assets", subsection_style))
                story.append(Spacer(1, 3))
                story.append(build_pdf_subtable(normal_tanks_list, selected_params, today_readings, active_alerts, cell_style, cell_style_bold, header_style))
                story.append(Spacer(1, 15))
                
            # Violated Tanks Subtable
            if violated_tanks_list:
                story.append(Paragraph(f"{folder_name} — Violated Assets", ParagraphStyle('ViolatedSubSec', parent=subsection_style, textColor=colors.HexColor('#9B2C2C'))))
                story.append(Spacer(1, 3))
                story.append(build_pdf_subtable(violated_tanks_list, selected_params, today_readings, active_alerts, cell_style, cell_style_bold, header_style))
                story.append(Spacer(1, 15))
                
        if not has_detailed_data:
            story.append(Paragraph("No detailed asset records found matching report config filters.", cell_style))
            
        doc.build(story)
        return True
    except Exception as e:
        print(f"[-] Error generating detailed inspection PDF: {e}")
        import traceback
        traceback.print_exc()
        return False

def generate_inspection_excel(filepath, date_str):
    """Generates today's inspection report Excel sheet"""
    if not OPENPYXL_AVAILABLE:
        print("[-] openpyxl library not available. Skipping Excel generation.")
        return False

    tanks_data = normalize_to_dict(fetch_db("tanks"))
    readings_data = normalize_to_dict(fetch_db("readings"))
    tree_data = normalize_to_dict(fetch_db("tank_tree"))

    today_readings = {}
    for r_id, r in readings_data.items():
        try:
            r_date = r.get("captured_at", "").split("T")[0]
            if r_date == date_str:
                today_readings[r.get("tank_id")] = r
        except:
            continue

    nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
    folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
    folders.sort(key=lambda x: x.get("order", 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection_Report"

    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="CB8C3E")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="252830")
    header_fill = PatternFill(start_color="CB8C3E", end_color="CB8C3E", fill_type="solid")
    section_fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    border_side = Side(border_style="thin", color="E0E0E0")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Title Block
    ws["A1"] = "VAPLI Telemetry Summary"
    ws["A1"].font = title_font
    ws["A2"] = f"Client: {CONFIG['client_id'].upper()} | Date: {date_str} | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)

    curr_row = 4

    def get_folder_tanks(folder_id):
        tank_ids = []
        queue = [folder_id]
        while queue:
            curr = queue.pop(0)
            children = [n for n in nodes if n.get("parent_id") == curr]
            for c in children:
                if c.get("type") == "leaf" and c.get("tank_id"):
                    tank_ids.append(c.get("tank_id"))
                elif c.get("type") == "folder":
                    queue.append(c.get("id"))
        return tank_ids

    for folder in folders:
        tank_ids = get_folder_tanks(folder.get("id"))
        if not tank_ids:
            continue
        
        folder_tanks = [tanks_data[tid] for tid in tank_ids if tid in tanks_data]
        if not folder_tanks:
            continue

        # Section Header
        ws.cell(row=curr_row, column=1, value=f"Group: {folder.get('name')}").font = section_font
        ws.cell(row=curr_row, column=1).fill = section_fill
        curr_row += 1

        # Unique parameters
        group_params = []
        for tank in folder_tanks:
            props = tank.get("inspection_properties", [])
            for p in props:
                label = p.get("label", "")
                if label and label not in group_params:
                    group_params.append(label)

        # Table Header
        headers = ["Asset Name"] + group_params + ["Capture Time"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        curr_row += 1

        # Table Data
        for tank in folder_tanks:
            tid = tank.get("id")
            reading = today_readings.get(tid)
            
            ws.cell(row=curr_row, column=1, value=tank.get("tank_name", "Asset")).border = thin_border
            
            if not reading:
                # Merge row
                ws.cell(row=curr_row, column=2, value="Readings not taken").border = thin_border
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=len(group_params)+2)
                # Fill missing color
                for c in range(1, len(group_params) + 3):
                    ws.cell(row=curr_row, column=c).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    ws.cell(row=curr_row, column=c).border = thin_border
            else:
                insp_vals = reading.get("inspection_values", {})
                for col_idx, gp in enumerate(group_params, start=2):
                    val = insp_vals.get(gp, "-")
                    if isinstance(val, dict):
                        if "left" in val or "right" in val:
                            val = f"{val.get('left', '')} / {val.get('right', '')}"
                        elif "value" in val:
                            val = val.get("value", "-")
                        else:
                            val = ", ".join(f"{k}: {v}" for k, v in val.items())
                    ws.cell(row=curr_row, column=col_idx, value=str(val)).border = thin_border
                
                # Time
                dt = parse_iso_datetime(reading.get("captured_at"))
                time_str = dt.strftime("%I:%M %p") if dt else (reading.get("captured_at", "") or "").split("T")[-1][:5]
                ws.cell(row=curr_row, column=len(group_params)+2, value=time_str).border = thin_border
            curr_row += 1
        
        curr_row += 2 # gap

    # Auto fit
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filepath)
    return True

def generate_alerts_pdf(filepath, date_str):
    """Generates alerts report PDF, renaming alarm references to alerts"""
    if not REPORTLAB_AVAILABLE:
        return False

    try:
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        
        today_alerts = []
        for a_id, a in alerts_data.items():
            try:
                raw_time = a.get("captured_at") or a.get("timestamp") or ""
                a_date = raw_time.split("T")[0]
                if a_date == date_str:
                    today_alerts.append(a)
            except:
                continue

        client_name = get_client_name(CONFIG['client_id'])
        generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")

        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'DocTitle', parent=styles['Heading1'],
            fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#D32F2F'),
            alignment=1, spaceAfter=8
        )
        meta_style = ParagraphStyle(
            'DocMeta', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#555555'),
            alignment=1, spaceAfter=15
        )
        cell_style = ParagraphStyle(
            'CellText', parent=styles['Normal'],
            fontName='Helvetica', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#2D3748')
        )
        cell_style_bold = ParagraphStyle(
            'CellTextBold', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#2D3748')
        )
        header_cell_style = ParagraphStyle(
            'HeaderCellText', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor('#FFFFFF'),
            alignment=1
        )

        story = []
        story.append(Paragraph(f"VAPLI Daily Alert Report", title_style))
        story.append(Paragraph(f"Client: {client_name} | Date: {date_str} | Generated: {generated_at}", meta_style))

        headers = [
            Paragraph("Time", header_cell_style),
            Paragraph("Asset Name", header_cell_style),
            Paragraph("Severity", header_cell_style),
            Paragraph("Parameter", header_cell_style),
            Paragraph("Value", header_cell_style),
            Paragraph("Alert Message", header_cell_style),
            Paragraph("IF-THEN Detail", header_cell_style)
        ]
        table_data = [headers]
        table_commands = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#9B2C2C')), # deep alert red
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]

        row_idx = 1
        for a in today_alerts:
            raw_time = a.get("captured_at") or a.get("timestamp") or ""
            dt = parse_iso_datetime(raw_time)
            time_str = dt.strftime("%I:%M %p") if dt else raw_time.split("T")[-1][:5]
            
            sev = (a.get("constraint_severity") or a.get("severity") or "WARNING").upper()
            label = a.get("constraint_label") or a.get("param_label") or "-"
            val = a.get("violated_value") or a.get("param_value") or "-"
            msg = a.get("message") or a.get("alert_title") or "-"
            if_then_val = a.get("if_then", "-") or "-"
            
            row = [
                Paragraph(time_str, cell_style),
                Paragraph(f"{a.get('tank_name')} ({a.get('tank_code', '')})", cell_style_bold),
                Paragraph(sev, cell_style_bold),
                Paragraph(label, cell_style),
                Paragraph(str(val), cell_style_bold),
                Paragraph(msg, cell_style),
                Paragraph(if_then_val, cell_style)
            ]
            table_data.append(row)
            
            # Highlight row backgrounds
            sev_lower = sev.lower()
            color_hex = '#FFCDD2' if sev_lower == 'critical' else ('#FFF9C4' if sev_lower == 'warning' else '#EDF2F7')
            table_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(color_hex)))
            row_idx += 1

        col_widths = [55, 95, 55, 75, 45, 125, 90]
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle(table_commands))

        story.append(t)
        doc.build(story)
        return True
    except Exception as e:
        print(f"[-] Error generating alerts PDF: {e}")
        return False

def generate_alerts_excel(filepath, date_str):
    """Generates alerts report Excel sheet"""
    if not OPENPYXL_AVAILABLE:
        return False

    try:
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        
        today_alerts = []
        for a_id, a in alerts_data.items():
            try:
                raw_time = a.get("captured_at") or a.get("timestamp") or ""
                a_date = raw_time.split("T")[0]
                if a_date == date_str:
                    today_alerts.append(a)
            except:
                continue

        wb = Workbook()
        ws = wb.active
        ws.title = "Alerts_Summary"

        title_font = Font(name="Calibri", size=16, bold=True, color="D32F2F")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        
        red_header_fill = PatternFill(start_color="9B2C2C", end_color="9B2C2C", fill_type="solid")
        red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        gray_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        left_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')

        ws.append(["VAPLI Daily Alert Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.row_dimensions[1].height = 30
        
        client_name = get_client_name(CONFIG['client_id'])
        ws.append([f"Client: {client_name} | Date: {date_str} | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.append([]) # spacer

        headers = ["Time", "Asset Name", "Asset Code", "Severity", "Parameter Checked", "Violated Value", "Alert Title", "Alert Message", "IF-THEN Detail"]
        ws.append(headers)
        ws.row_dimensions[4].height = 24
        
        for col_idx in range(1, 10):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = red_header_fill
            cell.alignment = center_align
            cell.border = thin_border

        curr_row = 5
        for a in today_alerts:
            raw_time = a.get("captured_at") or a.get("timestamp") or ""
            dt = parse_iso_datetime(raw_time)
            time_str = dt.strftime("%I:%M %p") if dt else raw_time.split("T")[-1][:5]

            sev = (a.get("constraint_severity") or a.get("severity") or "WARNING").upper()
            label = a.get("constraint_label") or a.get("param_label") or "-"
            val = a.get("violated_value") or a.get("param_value") or "-"
            title = a.get("alert_title") or "-"
            msg = a.get("message") or "-"
            if_then_val = a.get("if_then", "-") or "-"

            ws.append([time_str, a.get("tank_name"), a.get("tank_code"), sev, label, val, title, msg, if_then_val])
            ws.row_dimensions[curr_row].height = 20

            sev_lower = sev.lower()
            fill_to_use = red_fill if sev_lower == "critical" else (yellow_fill if sev_lower == "warning" else gray_fill)
            
            for col_idx in range(1, 10):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = bold_font if col_idx in [2, 3, 4, 6] else regular_font
                cell.fill = fill_to_use
                cell.border = thin_border
                cell.alignment = center_align if col_idx in [1, 3, 4, 6] else left_align

            curr_row += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"[-] Error generating alerts Excel: {e}")
        return False

# ==============================================================================
# SMTP MAILER AND EMAIL BUILDERS
# ==============================================================================

def send_email(recipients, subject, html_body, attachments=None):
    """Logs into SMTP server and sends MIME email"""
    if not CONFIG["smtp_email"] or not CONFIG["smtp_password"]:
        print("[-] SMTP email or password not configured in .env. Cannot send email.")
        return False
    
    if not recipients:
        print("[-] No recipient email addresses supplied.")
        return False

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = CONFIG["smtp_email"]
        msg["To"] = ", ".join(recipients)

        msg.attach(MIMEText(html_body, "html"))

        if attachments:
            for filepath in attachments:
                if os.path.exists(filepath):
                    with open(filepath, "rb") as f:
                        part = MIMEApplication(f.read(), Name=os.path.basename(filepath))
                    part['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
                    msg.attach(part)

        # Connect
        server = smtplib.SMTP(CONFIG["smtp_server"], CONFIG["smtp_port"])
        server.starttls()
        server.login(CONFIG["smtp_email"], CONFIG["smtp_password"])
        server.sendmail(CONFIG["smtp_email"], recipients, msg.as_string())
        server.quit()
        print(f"[+] Email sent successfully to: {recipients}")
        return True
    except Exception as e:
        print(f"[-] Failed to send email via SMTP: {e}")
        return False

def build_modern_template(title, header_title, body_content):
    """Wraps body content in a clean, modern, and highly professional HTML template"""
    client_name = get_client_name(CONFIG['client_id'])
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
                background-color: #FAFAFA;
                color: #2D3748;
                margin: 0;
                padding: 0;
                -webkit-font-smoothing: antialiased;
            }}
            .container {{
                max-width: 600px;
                margin: 30px auto;
                background: #FFFFFF;
                border-radius: 8px;
                overflow: hidden;
                box-shadow: 0 1px 3px rgba(0,0,0,0.05);
                border: 1px solid #E2E8F0;
                padding: 30px;
            }}
            .header {{
                border-bottom: 2px solid #EDF2F7;
                padding-bottom: 15px;
                margin-bottom: 25px;
            }}
            .header h1 {{
                color: #1A365D;
                margin: 0;
                font-size: 20px;
                font-weight: 600;
            }}
            .header p {{
                color: #718096;
                margin: 5px 0 0 0;
                font-size: 13px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            .content {{
                line-height: 1.5;
                font-size: 14px;
            }}
            .content p {{
                margin: 0 0 15px 0;
            }}
            .footer {{
                border-top: 1px solid #EDF2F7;
                padding-top: 15px;
                margin-top: 30px;
                font-size: 11px;
                color: #A0AEC0;
                text-align: center;
                line-height: 1.5;
            }}
            .footer a {{
                color: #3182CE;
                text-decoration: none;
            }}
            .badge {{
                display: inline-block;
                padding: 2px 8px;
                border-radius: 4px;
                font-size: 10px;
                font-weight: 600;
                text-transform: uppercase;
            }}
            .badge-critical {{ background-color: #FED7D7; color: #9B2C2C; }}
            .badge-warning {{ background-color: #FEFCBF; color: #975A16; }}
            .badge-info {{ background-color: #EBF8FF; color: #2B6CB0; }}
            
            /* Tables */
            table.report-table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }}
            table.report-table th {{
                background-color: #F7FAFC;
                color: #4A5568;
                font-size: 11px;
                font-weight: 600;
                text-transform: uppercase;
                padding: 10px 12px;
                border-bottom: 2px solid #EDF2F7;
                text-align: left;
            }}
            table.report-table td {{
                padding: 10px 12px;
                border-bottom: 1px solid #EDF2F7;
                font-size: 13px;
            }}
            
            /* Form / CTA Link Button */
            .btn-cta {{
                display: inline-block;
                background-color: #3182CE;
                color: #FFFFFF !important;
                padding: 10px 20px;
                font-size: 13px;
                font-weight: 600;
                border-radius: 5px;
                text-decoration: none;
                margin: 20px 0 10px 0;
            }}
            .btn-cta:hover {{
                background-color: #2B6CB0;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>{header_title}</h1>
                <p>{title} — {client_name}</p>
            </div>
            <div class="content">
                {body_content}
            </div>
            <div class="footer">
                This is an automated report from your VAPLI lubrication monitoring system.<br>
                For support, contact <a href="mailto:muthukumarandeveloper@gmail.com">muthukumarandeveloper@gmail.com</a>
            </div>
        </div>
    </body>
    </html>
    """

# ==============================================================================
# INDIVIDUAL EMAIL ACTIONS
# ==============================================================================

def execute_daily_reports_email(target_date=None):
    """Compiles PDFs and Excel sheets and sends to report_receivers"""
    print("[*] Starting Daily Reports compilation...")
    if not target_date:
        target_date = get_ast_time().strftime("%Y-%m-%d")
    today_str = target_date
    
    # Create temp directory
    temp_dir = os.path.join(os.path.dirname(__file__), "temp_reports")
    os.makedirs(temp_dir, exist_ok=True)
    
    insp_pdf_path = os.path.join(temp_dir, f"InspectionReport_{today_str}.pdf")
    insp_xls_path = os.path.join(temp_dir, f"InspectionReport_{today_str}.xlsx")
    alerts_pdf_path = os.path.join(temp_dir, f"AlertsReport_{today_str}.pdf")
    alerts_xls_path = os.path.join(temp_dir, f"AlertsReport_{today_str}.xlsx")

    # Generate files
    generate_inspection_pdf(insp_pdf_path, today_str)
    generate_inspection_excel(insp_xls_path, today_str)
    generate_alerts_pdf(alerts_pdf_path, today_str)
    generate_alerts_excel(alerts_xls_path, today_str)

    # Fetch receivers
    settings = fetch_db("settings/Report_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))
        
    if not receivers:
        print("[-] No recipients stored under settings/Report_Recievers/Emailids")
        return

    client_name = get_client_name(CONFIG['client_id'])
    subject = f"VAPLI Telemetry & Alert Reports - {client_name}"
    
    body = f"""
    <p>Dear Team,</p>
    <p>Please find attached the official VAPLI telemetry and alert inspection reports compiled for today.</p>
    <p><b>Summary Details:</b></p>
    <ul>
        <li><b>Client:</b> {client_name}</li>
        <li><b>Date:</b> {datetime.now().strftime('%A, %d %B %Y')}</li>
        <li><b>Generated At:</b> {datetime.now().strftime('%I:%M %p %Z')}</li>
    </ul>
    <p>The attachments contain:
        <ol>
            <li><b>Today's Inspection Summary (PDF & Excel)</b>: Detailed parameter values and image links for all scanned lubrication assets.</li>
            <li><b>Alert Violation Log (PDF & Excel)</b>: Summary of active alerts, severity details, and triggered IF-THEN workflows.</li>
        </ol>
    </p>
    <p>Please inspect critical alerts immediately to protect physical asset lifecycles.</p>
    <br>
    <p>Warm regards,<br><b>VAPLI Automation Engine</b></p>
    """
    
    html_email = build_modern_template("Telemetry & Alerts Summary", "Inspection Summary Reports", body)

    attachments = [insp_pdf_path, insp_xls_path, alerts_pdf_path, alerts_xls_path]
    ok = send_email(receivers, subject, html_email, attachments)
    if ok:
        threading.Thread(target=send_email_sent_fcm, args=("Daily Inspection Report", len(receivers))).start()


    # Clean up files
    for path in attachments:
        try:
            if os.path.exists(path):
                os.remove(path)
        except:
            pass
    try:
        os.rmdir(temp_dir)
    except:
        pass

def execute_missing_tanks_email(target_date=None):
    """Generates compliance summary and sends personalized, signed form links to missing_tanks_receivers"""
    print("[*] Starting Missing Tanks compliance scan...")
    if not target_date:
        target_date = get_ast_time().strftime("%Y-%m-%d")
    today_str = target_date

    tanks_data = normalize_to_dict(fetch_db("tanks"))
    readings_data = normalize_to_dict(fetch_db("readings"))
    tree_data = normalize_to_dict(fetch_db("tank_tree"))
    settings = fetch_db("settings/Missing Tanks_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))

    if not receivers:
        print("[-] No recipients stored under settings/Missing Tanks_Recievers/Emailids")
        return

    # Gather today's readings
    today_readings = set()
    for r in readings_data.values():
        try:
            if r.get("captured_at", "").split("T")[0] == today_str:
                today_readings.add(r.get("tank_id"))
        except:
            continue

    nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
    top_folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
    top_folders.sort(key=lambda x: x.get("order", 0))

    def get_folder_tanks(folder_id):
        tank_ids = []
        queue = [folder_id]
        while queue:
            curr = queue.pop(0)
            children = [n for n in nodes if n.get("parent_id") == curr]
            for c in children:
                if c.get("type") == "leaf" and c.get("tank_id"):
                    tank_ids.append(c.get("tank_id"))
                elif c.get("type") == "folder":
                    queue.append(c.get("id"))
        return tank_ids

    # Group Tanks by Groups
    compliance_content = ""
    missing_tank_ids = []

    for folder in top_folders:
        t_ids = get_folder_tanks(folder.get("id"))
        if not t_ids:
            continue
        
        folder_tanks = [tanks_data[tid] for tid in t_ids if tid in tanks_data]
        if not folder_tanks:
            continue

        total_count = len(folder_tanks)
        inspected = [t for t in folder_tanks if t.get("id") in today_readings]
        uninspected = [t for t in folder_tanks if t.get("id") not in today_readings]
        inspected_count = len(inspected)
        
        missing_tank_ids.extend([t.get("id") for t in uninspected])
        
        percentage = (inspected_count / total_count) * 100 if total_count > 0 else 0

        compliance_content += f"<h3 style='margin-top:20px; color:#252830;'>Group: {folder.get('name')}</h3>"

        # Check conditions
        if inspected_count == 0:
            compliance_content += f"<p style='color:#D97706; font-weight:600;'>Group <b>{folder.get('name')}</b> has no tank recording Checkout for Readings</p>"
        elif percentage == 100.0:
            compliance_content += f"<p style='color:#2ECC71; font-weight:600;'>Group <b>{folder.get('name')}</b> is fully recorded.</p>"
        elif percentage < 50.0:
            compliance_content += f"<p>Group <b>{folder.get('name')}</b> has not been Recorded except these tanks</p>"
            compliance_content += render_3_column_list(inspected)
        else:
            compliance_content += f"<p>Group <b>{folder.get('name')}</b> has been Recorded except these tanks</p>"
            compliance_content += render_3_column_list(uninspected)

    client_name = get_client_name(CONFIG['client_id'])
    encoded_tanks = ",".join(missing_tank_ids)
    feedback_base = CONFIG.get("external_url") or "https://dummy-firebase-project-id.web.app"
    
    # Send personalized emails to each receiver
    import hashlib
    secret_salt = "vapli_compliance_secure_salt"
    
    for receiver in receivers:
        sig = hashlib.sha256((receiver + secret_salt).encode('utf-8')).hexdigest()
        feedback_url = f"{feedback_base}/feedback.html?client_id={CONFIG['client_id']}&date={today_str}&tanks={encoded_tanks}&env={CONFIG['env_mode']}&submitted_by={urllib.parse.quote(receiver)}&sig={sig}"
        
        feedback_btn = f"""
        <div style='text-align: center;'>
            <a href='{feedback_url}' class='btn-cta' target='_blank'>Submit Reason for Missed Inspections</a>
        </div>
        """

        body = f"""
        <p>Dear Supervisor,</p>
        <p>Here is the compliance report highlighting lubrication inspection gaps for today.</p>
        <hr style='border:0; border-top:1px solid #E2E8F0;'>
        {compliance_content}
        <hr style='border:0; border-top:1px solid #E2E8F0; margin-top:20px;'>
        <p style='margin-top:20px;'><b>Inspection Action Required:</b></p>
        <p>To keep the automation metrics clean, please submit the reasons for any missed inspections today by clicking the button below:</p>
        {feedback_btn}
        <br>
        <p>Thank you,<br><b>VAPLI Compliance Monitor</b></p>
        """

        subject = f"VAPLI Inspection Compliance Gaps - {datetime.now().strftime('%d/%m/%Y')} ({client_name})"
        html_email = build_modern_template("Inspection Gaps & Gaps Form", "Daily Compliance Alert", body)

        ok = send_email([receiver], subject, html_email)
        if ok:
            threading.Thread(target=send_email_sent_fcm, args=("Missing Tanks Compliance Alert", 1)).start()


def render_3_column_list(tanks):
    """Renders tank list as a responsive 3-column table/grid inside HTML emails"""
    if not tanks:
        return "<p style='color:#777;'>None</p>"
    
    html = "<table style='width:100%; border:none; margin: 10px 0;'><tr>"
    col = 0
    for t in tanks:
        if col > 0 and col % 3 == 0:
            html += "</tr><tr>"
        html += f"<td style='width:33.3%; padding: 4px 8px; border:none; font-size:12px; color:#555;'>&bull; {t.get('tank_name')} ({t.get('tank_code', '')})</td>"
        col += 1
    
    # Fill remaining cells in the row
    while col % 3 != 0:
        html += "<td style='width:33.3%; border:none;'></td>"
        col += 1
        
    html += "</tr></table>"
    return html

def execute_alerts_notification_email():
    """Sends list of current unresolved alerts to alerts_receivers, using Alerts instead of Alarms"""
    print("[*] Starting Alert violator notification mail...")
    
    # Fetch active dashboard alerts
    alerts_data = normalize_to_dict(fetch_db("alerts"))
    active_alerts = []
    for a_id, a in alerts_data.items():
        if not a.get("resolved", False) and a.get("show_dashboard_alert", True):
            active_alerts.append(a)

    settings = fetch_db("settings/Alerts_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))

    if not receivers:
        print("[-] No recipients stored under settings/Alerts_Recievers/Emailids")
        return

    if not active_alerts:
        print("[*] No active alerts to notify. Skipping alerts email.")
        return

    # Build active alert table rows
    table_rows = ""
    for a in active_alerts:
        sev = (a.get("constraint_severity") or a.get("severity") or "warning").lower()
        badge_class = f"badge-{sev}"
        
        raw_time = a.get("captured_at") or a.get("timestamp") or ""
        dt = parse_iso_datetime(raw_time)
        time_str = dt.strftime("%d/%m/%Y %I:%M %p") if dt else raw_time.replace("T", " ")[:16]

        label = a.get("constraint_label") or a.get("param_label") or "-"
        val = a.get("violated_value") or a.get("param_value") or "-"
        msg = a.get("message") or a.get("alert_title") or "-"

        table_rows += f"""
        <tr>
            <td style='white-space:nowrap;'>{time_str}</td>
            <td><b>{a.get('tank_name')}</b><br><span style='font-size:11px; color:#64748B;'>{a.get('tank_code', '')}</span></td>
            <td><span class='badge {badge_class}'>{sev.upper()}</span></td>
            <td>{label}</td>
            <td style='color:#EF4444; font-weight:600;'>{val}</td>
            <td>{msg}</td>
        </tr>
        """

    alert_table = f"""
    <table class='report-table'>
        <thead>
            <tr>
                <th>Time</th>
                <th>Asset</th>
                <th>Severity</th>
                <th>Parameter</th>
                <th>Violated Value</th>
                <th>Message</th>
            </tr>
        </thead>
        <tbody>
            {table_rows}
        </tbody>
    </table>
    """

    client_name = get_client_name(CONFIG['client_id'])
    body = f"""
    <p>Dear Team,</p>
    <p>Our constraint engine has surfaced the following active alerts that require your immediate attention.</p>
    {alert_table}
    <p style='margin-top:20px; font-weight:600; color:#EF4444;'>Action Required:</p>
    <p>Log in to the VAPLI Admin Dashboard to inspect these violations, resolve the issues on-site, and mark these alerts as resolved.</p>
    <br>
    <p>Regards,<br><b>VAPLI Alert Center</b></p>
    """

    subject = f"URGENT: Active Alert Notification - {len(active_alerts)} Alerts Unresolved ({client_name})"
    html_email = build_modern_template("Active Alert Center", "Active Alert Notification", body)

    ok = send_email(receivers, subject, html_email)
    if ok:
        threading.Thread(target=send_email_sent_fcm, args=("Active Alerts Notification", len(receivers))).start()


def execute_weekly_reports_email():
    """Compiles weekly inspection PDF and Excel reports and emails to report receivers"""
    print("[*] Starting Weekly Inspection Reports compilation...")
    today_str = datetime.now().strftime("%Y-%m-%d")
    # Date window: last 7 days
    start_date = (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d")
    
    temp_dir = os.path.join(os.path.dirname(__file__), "temp_weekly")
    os.makedirs(temp_dir, exist_ok=True)
    
    pdf_filename = os.path.join(temp_dir, f"WeeklyInspectionReport_{start_date}_to_{today_str}.pdf")
    excel_filename = os.path.join(temp_dir, f"WeeklyInspectionReport_{start_date}_to_{today_str}.xlsx")
    
    print("[*] Generating Weekly Inspection PDF...")
    pdf_ok = generate_weekly_inspection_pdf(pdf_filename, start_date, today_str)
    print("[*] Generating Weekly Inspection Excel...")
    excel_ok = generate_weekly_inspection_excel(excel_filename, start_date, today_str)
    
    if not pdf_ok and not excel_ok:
        print("[-] Weekly report generation failed. Email aborted.")
        return
        
    settings = fetch_db("settings/Report_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))
    if not receivers:
        print("[-] No report receivers defined under settings/Report_Recievers/Emailids")
        return

    client_name = get_client_name(CONFIG['client_id'])
    
    body = f"""
    <p>Dear Team,</p>
    <p>Please find attached the official VAPLI Weekly Inspection Telemetry Reports compiled for the period <b>{start_date}</b> to <b>{today_str}</b>.</p>
    <p><b>Summary Details:</b></p>
    <ul>
        <li><b>Client:</b> {client_name}</li>
        <li><b>Report Period:</b> {start_date} to {today_str}</li>
        <li><b>Generated At:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</li>
    </ul>
    <p>The attachments contain:
        <ol>
            <li><b>Weekly Inspection PDF Report</b>: Professional summary and dynamic multi-column day-by-day telemetry matching your folder configurations.</li>
            <li><b>Weekly Telemetry Spreadsheet (Excel)</b>: Full raw reading columns color-coded for quick verification.</li>
        </ol>
    </p>
    <br>
    <p>Warm regards,<br><b>VAPLI Automation Engine</b></p>
    """
    
    subject = f"VAPLI Weekly Telemetry Reports - {client_name}"
    html_email = build_modern_template("Weekly Telemetry Summary", "Weekly Telemetry Report", body)

    attachments = []
    if pdf_ok and os.path.exists(pdf_filename):
        attachments.append(pdf_filename)
    if excel_ok and os.path.exists(excel_filename):
        attachments.append(excel_filename)

    ok = send_email(receivers, subject, html_email, attachments)
    if ok:
        threading.Thread(target=send_email_sent_fcm, args=("Weekly Inspection Report", len(receivers))).start()


    # Cleanup
    for path in attachments:
        try:
            if os.path.exists(path):
                os.remove(path)
        except:
            pass
    try:
        os.rmdir(temp_dir)
    except:
        pass

def execute_weekly_alerts_email():
    """Compiles weekly alerts PDF and Excel reports and emails to report receivers"""
    print("[*] Starting Weekly Alerts Reports compilation...")
    today_str = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d")
    
    temp_dir = os.path.join(os.path.dirname(__file__), "temp_weekly")
    os.makedirs(temp_dir, exist_ok=True)
    
    pdf_filename = os.path.join(temp_dir, f"WeeklyAlertsReport_{start_date}_to_{today_str}.pdf")
    excel_filename = os.path.join(temp_dir, f"WeeklyAlertsReport_{start_date}_to_{today_str}.xlsx")
    
    print("[*] Generating Weekly Alerts PDF...")
    pdf_ok = generate_weekly_alerts_pdf(pdf_filename, start_date, today_str)
    print("[*] Generating Weekly Alerts Excel...")
    excel_ok = generate_weekly_alerts_excel(excel_filename, start_date, today_str)
    
    if not pdf_ok and not excel_ok:
        print("[-] Weekly Alerts report generation failed. Email aborted.")
        return
        
    settings = fetch_db("settings/Report_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))
    if not receivers:
        print("[-] No report receivers defined under settings/Report_Recievers/Emailids")
        return

    client_name = get_client_name(CONFIG['client_id'])
    
    body = f"""
    <p>Dear Team,</p>
    <p>Please find attached the official VAPLI Weekly Alerts Summary Reports compiled for the period <b>{start_date}</b> to <b>{today_str}</b>.</p>
    <p><b>Summary Details:</b></p>
    <ul>
        <li><b>Client:</b> {client_name}</li>
        <li><b>Report Period:</b> {start_date} to {today_str}</li>
        <li><b>Generated At:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</li>
    </ul>
    <p>The attachments contain:
        <ol>
            <li><b>Weekly Alert PDF Log</b>: Full list of unresolved and resolved alerts logged during this week, formatted with priority severity highlights.</li>
            <li><b>Weekly Alerts Log (Excel)</b>: Detailed spreadsheet mapping times, parameter labels, values, and alert messages.</li>
        </ol>
    </p>
    <br>
    <p>Warm regards,<br><b>VAPLI Alert Center</b></p>
    """
    
    subject = f"VAPLI Weekly Alerts Summary - {client_name}"
    html_email = build_modern_template("Weekly Alerts Summary", "Weekly Alert Report", body)

    attachments = []
    if pdf_ok and os.path.exists(pdf_filename):
        attachments.append(pdf_filename)
    if excel_ok and os.path.exists(excel_filename):
        attachments.append(excel_filename)

    ok = send_email(receivers, subject, html_email, attachments)
    if ok:
        threading.Thread(target=send_email_sent_fcm, args=("Weekly Alerts Report", len(receivers))).start()


    # Cleanup
    for path in attachments:
        try:
            if os.path.exists(path):
                os.remove(path)
        except:
            pass
    try:
        os.rmdir(temp_dir)
    except:
        pass

def generate_weekly_inspection_pdf(filepath, start_date_str, end_date_str):
    """Generates a weekly inspection telemetry report matching the multi-column format of the Flutter app"""
    if not REPORTLAB_AVAILABLE:
        print("[-] reportlab library not available. Skipping PDF generation.")
        return False
        
    try:
        tanks_data = normalize_to_dict(fetch_db("tanks"))
        readings_data = normalize_to_dict(fetch_db("readings"))
        tree_data = normalize_to_dict(fetch_db("tank_tree"))
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        report_format_configs = normalize_to_dict(fetch_db("settings/report_format"))
        
        # Calculate date range list (list of 7 days)
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        delta = end_date - start_date
        days_list = []
        for i in range(delta.days + 1):
            days_list.append(start_date + timedelta(days=i))
        
        # Gather readings within this week's date range
        weekly_readings = {} # tank_id -> date_str -> reading
        for r_id, r in readings_data.items():
            try:
                r_date = r.get("captured_at", "").split("T")[0]
                if start_date_str <= r_date <= end_date_str:
                    tid = r.get("tank_id")
                    if tid not in weekly_readings:
                        weekly_readings[tid] = {}
                    existing = weekly_readings[tid].get(r_date)
                    if not existing or r.get("captured_at", "") > existing.get("captured_at", ""):
                        weekly_readings[tid][r_date] = r
            except:
                continue

        active_alerts = []
        for a_id, a in alerts_data.items():
            if not a.get("resolved", False) and a.get("status", "").lower() != "completed":
                active_alerts.append(a)

        nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
        folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
        folders.sort(key=lambda x: x.get("order", 0))

        def get_folder_tanks(folder_id):
            tank_ids = []
            queue = [folder_id]
            while queue:
                curr = queue.pop(0)
                children = [n for n in nodes if n.get("parent_id") == curr]
                for c in children:
                    if c.get("type") == "leaf" and c.get("tank_id"):
                        tank_ids.append(c.get("tank_id"))
                    elif c.get("type") == "folder":
                        queue.append(c.get("id"))
            return tank_ids

        client_name = get_client_name(CONFIG['client_id'])
        generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")

        # Summary calculations
        total_configured = len(tanks_data)
        total_possible = total_configured * len(days_list)
        total_scanned = sum(len(weekly_readings.get(tid, {})) for tid in tanks_data.keys())
        compliance_rate = (total_scanned / total_possible * 100) if total_possible > 0 else 0.0

        story = []
        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'DocTitle', parent=styles['Heading1'],
            fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#CB8C3E'),
            alignment=1, spaceAfter=8
        )
        meta_style = ParagraphStyle(
            'DocMeta', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#555555'),
            alignment=1, spaceAfter=15
        )
        section_style = ParagraphStyle(
            'SecTitle', parent=styles['Heading2'],
            fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#1A365D'),
            spaceBefore=10, spaceAfter=6
        )
        subsection_style = ParagraphStyle(
            'SubSecTitle', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=9.5, textColor=colors.HexColor('#2B6CB0'),
            spaceBefore=6, spaceAfter=4
        )
        cell_style = ParagraphStyle(
            'CellText', parent=styles['Normal'],
            fontName='Helvetica', fontSize=6.5, leading=8.5, textColor=colors.HexColor('#2D3748')
        )
        cell_style_bold = ParagraphStyle(
            'CellTextBold', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=6.5, leading=8.5, textColor=colors.HexColor('#2D3748')
        )
        header_style = ParagraphStyle(
            'HeaderStyle', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=7.0, leading=9.0, textColor=colors.HexColor('#FFFFFF'),
            alignment=1
        )
        header_sub_style = ParagraphStyle(
            'HeaderSubStyle', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=5.5, leading=7.5, textColor=colors.HexColor('#2D3748'),
            alignment=1
        )

        # SUMMARY PAGE (Page 1)
        story.append(Paragraph("VAPLI Weekly Inspection Telemetry Report", title_style))
        story.append(Paragraph(f"Client: {client_name} | Period: {start_date_str} to {end_date_str} | Generated: {generated_at}", meta_style))
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("Weekly Performance Summary", section_style))
        story.append(Spacer(1, 5))
        
        stats_table_data = [
            [Paragraph("<b>Metric</b>", cell_style_bold), Paragraph("<b>Value / Count</b>", cell_style_bold)],
            [Paragraph("Total Configured Assets", cell_style), Paragraph(str(total_configured), cell_style_bold)],
            [Paragraph("Total Inspections Expected", cell_style), Paragraph(str(total_possible), cell_style_bold)],
            [Paragraph("Total Inspections Completed", cell_style), Paragraph(str(total_scanned), cell_style_bold)],
            [Paragraph("Weekly Compliance Rate", cell_style), Paragraph(f"{compliance_rate:.1f}%", cell_style_bold)],
            [Paragraph("Active Unresolved Alerts", cell_style), Paragraph(str(len(active_alerts)), cell_style_bold)]
        ]
        stats_table = Table(stats_table_data, colWidths=[200, 100])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 15))

        story.append(PageBreak())

        # DETAILED TABLES FOR EACH GROUP
        has_detailed_data = False
        for folder in folders:
            tank_ids = get_folder_tanks(folder.get("id"))
            if not tank_ids:
                continue
            
            folder_tanks = [tanks_data[tid] for tid in tank_ids if tid in tanks_data]
            if not folder_tanks:
                continue
                
            selected_params = get_selected_params_for_folder(folder.get("id"), folder_tanks, report_format_configs)
            if not selected_params:
                continue
                
            has_detailed_data = True
            normal_tanks_list = []
            violated_tanks_list = []
            
            for tank in folder_tanks:
                tid = tank.get("id")
                has_active_alert = any(a.get("tank_id") == tid for a in active_alerts)
                if has_active_alert:
                    violated_tanks_list.append(tank)
                else:
                    normal_tanks_list.append(tank)
                    
            folder_name = folder.get("name")
            
            if normal_tanks_list:
                story.append(Paragraph(f"{folder_name} — Normal Assets", subsection_style))
                story.append(Spacer(1, 3))
                story.append(build_pdf_weekly_subtable(normal_tanks_list, selected_params, weekly_readings, days_list, active_alerts, cell_style, cell_style_bold, header_style, header_sub_style))
                story.append(Spacer(1, 15))
                
            if violated_tanks_list:
                story.append(Paragraph(f"{folder_name} — Violated Assets", ParagraphStyle('ViolatedSubSec', parent=subsection_style, textColor=colors.HexColor('#9B2C2C'))))
                story.append(Spacer(1, 3))
                story.append(build_pdf_weekly_subtable(violated_tanks_list, selected_params, weekly_readings, days_list, active_alerts, cell_style, cell_style_bold, header_style, header_sub_style))
                story.append(Spacer(1, 15))
                
        if not has_detailed_data:
            story.append(Paragraph("No detailed asset records found matching weekly report filters.", cell_style))
            
        doc.build(story)
        return True
    except Exception as e:
        print(f"[-] Error generating weekly inspection PDF: {e}")
        import traceback
        traceback.print_exc()
        return False

def build_pdf_weekly_subtable(tanks, params, readings, days, active_alerts, cell_style, cell_style_bold, header_style, header_sub_style):
    """Helper to build a multi-column weekly table in ReportLab"""
    num_params = len(params)
    num_days = len(days)
    
    # Calculate column widths: total width 540
    param_col_width = 430.0 / (num_params * num_days) if num_params > 0 else 430.0
    col_widths = [110] + [param_col_width] * (num_params * num_days)
    
    # Header Row 0: Asset Name, then Parameter Labels repeated/spanned
    row0 = [Paragraph("Asset Name", header_style)]
    for p in params:
        p_label = p["label"]
        row0.append(Paragraph(p_label, header_style))
        for _ in range(num_days - 1):
            row0.append(Paragraph("", header_style))
            
    # Header Row 1: empty for Asset Name, then Day Labels (e.g. 21/06)
    row1 = [Paragraph("", header_style)]
    day_colors_list = ['#FFFFFF', '#EDF2F7']
    
    for p in params:
        for d_idx, day in enumerate(days):
            day_str = day.strftime("%d/%m")
            row1.append(Paragraph(day_str, header_sub_style))
            
    table_data = [row0, row1]
    
    # Grid and default styles
    table_commands = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F497D')), # Dark Blue header
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#CBD5E1')),
        ('SPAN', (0,0), (0,1)), # Span Asset Name vertically
        ('VALIGN', (0,0), (0,1), 'MIDDLE'),
        ('ALIGN', (0,0), (0,1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
    ]
    
    # Add parameter spans in Header Row 0 and header backgrounds
    for k in range(num_params):
        col_start = 1 + k * num_days
        col_end = col_start + num_days - 1
        table_commands.append(('SPAN', (col_start, 0), (col_end, 0)))
        table_commands.append(('ALIGN', (col_start, 0), (col_end, 0), 'CENTER'))
        
        # Color header row 1 with alternating day shading
        for d_idx in range(num_days):
            col_bg = day_colors_list[d_idx % 2]
            table_commands.append(('BACKGROUND', (col_start + d_idx, 1), (col_start + d_idx, 1), colors.HexColor(col_bg)))

    row_idx = 2
    for tank in tanks:
        tid = tank.get("id")
        tank_code = tank.get("tank_code")
        tank_name = tank.get("tank_name", "Asset")
        display_name = f"{tank_name} ({tank_code})" if tank_code else tank_name
        
        tank_readings = readings.get(tid, {}) # date_str -> reading
        
        # Determine overall alert severity
        tank_severity = None
        for a in active_alerts:
            if a.get("tank_id") == tid:
                sev = (a.get("constraint_severity") or a.get("severity") or "warning").lower()
                if sev == "critical":
                    tank_severity = "critical"
                    break
                elif sev == "warning":
                    tank_severity = "warning"
                    
        has_any_readings = len(tank_readings) > 0
        default_row_color = '#FFFFFF' if row_idx % 2 == 1 else '#F8FAFC'
        
        row = [Paragraph(display_name, cell_style_bold)]
        
        if not has_any_readings:
            row.append(Paragraph("------- Readings not taken ------", cell_style_bold))
            for _ in range(num_params * num_days - 1):
                row.append(Paragraph("", cell_style))
                
            table_data.append(row)
            table_commands.append(('BACKGROUND', (1, row_idx), (-1, row_idx), colors.HexColor('#C8E6C9')))
            table_commands.append(('SPAN', (1, row_idx), (-1, row_idx)))
            table_commands.append(('ALIGN', (1, row_idx), (-1, row_idx), 'CENTER'))
        else:
            for p in params:
                p_label = p["label"]
                p_id = p["prop_id"]
                
                for d_idx, day in enumerate(days):
                    day_str = day.strftime("%Y-%m-%d")
                    reading = tank_readings.get(day_str)
                    
                    cell_bg = default_row_color if d_idx % 2 == 0 else '#F1F5F9'
                    
                    if not reading:
                        row.append(Paragraph("-", cell_style))
                        table_commands.append(('BACKGROUND', (len(row)-1, row_idx), (len(row)-1, row_idx), colors.HexColor(cell_bg)))
                    else:
                        insp_vals = reading.get("inspection_values", {})
                        val = insp_vals.get(p_label)
                        if val is None:
                            val = insp_vals.get(p_id)
                        val_str = format_reading_value(val)
                        
                        p_img = insp_vals.get(f"{p_id}__image_url")
                        if not p_img:
                            p_img = insp_vals.get(f"{p_label}__image_url")
                            
                        cell_content = val_str
                        if p_img:
                            cell_content += f'<br/><a href="{p_img}" color="#1565C0"><u>[Photo]</u></a>'
                            
                        is_violated = False
                        cell_severity = None
                        
                        for a in active_alerts:
                            if a.get("tank_id") == tid:
                                label = a.get("constraint_label") or a.get("param_label") or ""
                                if label.strip().lower() == p_label.strip().lower():
                                    alert_time = a.get("captured_at") or a.get("timestamp") or ""
                                    if alert_time.startswith(day_str):
                                        is_violated = True
                                        cell_severity = (a.get("constraint_severity") or a.get("severity") or "critical").lower()
                                        break
                                        
                        if not is_violated and has_take_action_keyword(val):
                            is_violated = True
                            cell_severity = "critical"
                            
                        col_pos = len(row)
                        if is_violated:
                            row.append(Paragraph(cell_content, cell_style_bold))
                            color_hex = '#FFCDD2' if cell_severity == 'critical' else ('#FFF9C4' if cell_severity == 'warning' else '#BBDEFB')
                            table_commands.append(('BACKGROUND', (col_pos, row_idx), (col_pos, row_idx), colors.HexColor(color_hex)))
                        else:
                            row.append(Paragraph(cell_content, cell_style))
                            table_commands.append(('BACKGROUND', (col_pos, row_idx), (col_pos, row_idx), colors.HexColor(cell_bg)))
                            
            table_data.append(row)
            
            if tank_severity:
                color_hex = '#FFCDD2' if tank_severity == 'critical' else ('#FFF9C4' if tank_severity == 'warning' else '#BBDEFB')
                table_commands.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor(color_hex)))
            else:
                table_commands.append(('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor(default_row_color)))
                
        row_idx += 1
        
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle(table_commands))
    return t

def generate_weekly_inspection_excel(filepath, start_date_str, end_date_str):
    """Generates today's weekly inspection Excel sheet"""
    if not OPENPYXL_AVAILABLE:
        print("[-] openpyxl library not available. Skipping Excel generation.")
        return False
        
    try:
        tanks_data = normalize_to_dict(fetch_db("tanks"))
        readings_data = normalize_to_dict(fetch_db("readings"))
        tree_data = normalize_to_dict(fetch_db("tank_tree"))
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        report_format_configs = normalize_to_dict(fetch_db("settings/report_format"))
        
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        delta = end_date - start_date
        days_list = []
        for i in range(delta.days + 1):
            days_list.append(start_date + timedelta(days=i))
        
        weekly_readings = {}
        for r_id, r in readings_data.items():
            try:
                r_date = r.get("captured_at", "").split("T")[0]
                if start_date_str <= r_date <= end_date_str:
                    tid = r.get("tank_id")
                    if tid not in weekly_readings:
                        weekly_readings[tid] = {}
                    existing = weekly_readings[tid].get(r_date)
                    if not existing or r.get("captured_at", "") > existing.get("captured_at", ""):
                        weekly_readings[tid][r_date] = r
            except:
                continue

        active_alerts = []
        for a_id, a in alerts_data.items():
            if not a.get("resolved", False) and a.get("status", "").lower() != "completed":
                active_alerts.append(a)

        nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
        folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
        folders.sort(key=lambda x: x.get("order", 0))

        def get_folder_tanks(folder_id):
            tank_ids = []
            queue = [folder_id]
            while queue:
                curr = queue.pop(0)
                children = [n for n in nodes if n.get("parent_id") == curr]
                for c in children:
                    if c.get("type") == "leaf" and c.get("tank_id"):
                        tank_ids.append(c.get("tank_id"))
                    elif c.get("type") == "folder":
                        queue.append(c.get("id"))
            return tank_ids

        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Telemetry"
        
        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="CB8C3E")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        sub_header_font = Font(name="Calibri", size=9, bold=True, color="000000")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        
        blue_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Title
        ws.append(["VAPLI Weekly Inspection Telemetry Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.row_dimensions[1].height = 30
        
        client_name = get_client_name(CONFIG['client_id'])
        ws.append([f"Client: {client_name} | Period: {start_date_str} to {end_date_str}"])
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.append([]) # spacer
        
        curr_row = 4
        
        for folder in folders:
            tank_ids = get_folder_tanks(folder.get("id"))
            if not tank_ids:
                continue
            
            folder_tanks = [tanks_data[tid] for tid in tank_ids if tid in tanks_data]
            if not folder_tanks:
                continue
                
            selected_params = get_selected_params_for_folder(folder.get("id"), folder_tanks, report_format_configs)
            if not selected_params:
                continue
                
            num_params = len(selected_params)
            num_days = len(days_list)
            
            # Header Row 1: Folder name
            ws.cell(row=curr_row, column=1, value=folder.get("name").upper()).font = bold_font
            curr_row += 1
            
            # Header Row 2: Asset Name, then parameter names spanned
            ws.cell(row=curr_row, column=1, value="Asset Name").font = header_font
            ws.cell(row=curr_row, column=1).fill = blue_fill
            ws.cell(row=curr_row, column=1).alignment = center_align
            
            # Format the vertically spanned bottom cell of Asset Name before merging
            ws.cell(row=curr_row + 1, column=1).fill = blue_fill
            
            for k, p in enumerate(selected_params):
                col_start = 2 + k * num_days
                col_end = col_start + num_days - 1
                
                # Format all cells in the span BEFORE merging
                for col_idx in range(col_start, col_end + 1):
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.fill = blue_fill
                    cell.border = thin_border
                
                top_left = ws.cell(row=curr_row, column=col_start, value=p["label"])
                top_left.font = header_font
                top_left.alignment = center_align
                
                ws.merge_cells(start_row=curr_row, start_column=col_start, end_row=curr_row, end_column=col_end)
            
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row+1, end_column=1)
            
            curr_row += 1
            
            # Header Row 3: Dates for each day under parameters (skip col 1 which is merged)
            for k in range(num_params):
                for d_idx, day in enumerate(days_list):
                    col_idx = 2 + k * num_days + d_idx
                    day_lbl = day.strftime("%d/%m")
                    cell = ws.cell(row=curr_row, column=col_idx, value=day_lbl)
                    cell.font = sub_header_font
                    cell.alignment = center_align
                    cell.fill = gray_fill if d_idx % 2 == 1 else PatternFill(fill_type=None)
                    cell.border = thin_border
            
            curr_row += 1
            
            # Populate Tank rows
            for tank in folder_tanks:
                tid = tank.get("id")
                tank_code = tank.get("tank_code")
                tank_name = tank.get("tank_name", "Asset")
                display_name = f"{tank_name} ({tank_code})" if tank_code else tank_name
                
                tank_readings = weekly_readings.get(tid, {})
                has_any_readings = len(tank_readings) > 0
                
                # Determine overall severity
                tank_severity = None
                for a in active_alerts:
                    if a.get("tank_id") == tid:
                        sev = (a.get("constraint_severity") or a.get("severity") or "warning").lower()
                        if sev == "critical":
                            tank_severity = "critical"
                            break
                        elif sev == "warning":
                            tank_severity = "warning"
                            
                ws.row_dimensions[curr_row].height = 20
                asset_cell = ws.cell(row=curr_row, column=1, value=display_name)
                asset_cell.font = bold_font
                asset_cell.border = thin_border
                asset_cell.alignment = left_align
                
                if tank_severity:
                    asset_cell.fill = red_fill if tank_severity == "critical" else yellow_fill
                
                if not has_any_readings:
                    col_start = 2
                    col_end = col_start + num_params * num_days - 1
                    ws.cell(row=curr_row, column=col_start, value="------- Readings not taken ------").font = bold_font
                    ws.cell(row=curr_row, column=col_start).alignment = center_align
                    ws.merge_cells(start_row=curr_row, start_column=col_start, end_row=curr_row, end_column=col_end)
                    for col_idx in range(col_start, col_end + 1):
                        cell = ws.cell(row=curr_row, column=col_idx)
                        cell.fill = green_fill
                        cell.border = thin_border
                else:
                    for k, p in enumerate(selected_params):
                        p_label = p["label"]
                        p_id = p["prop_id"]
                        
                        for d_idx, day in enumerate(days_list):
                            day_str = day.strftime("%Y-%m-%d")
                            col_idx = 2 + k * num_days + d_idx
                            
                            reading = tank_readings.get(day_str)
                            cell = ws.cell(row=curr_row, column=col_idx)
                            cell.border = thin_border
                            cell.alignment = center_align
                            
                            default_cell_fill = gray_fill if d_idx % 2 == 1 else PatternFill(fill_type=None)
                            if default_cell_fill.fill_type:
                                cell.fill = default_cell_fill
                            
                            if not reading:
                                cell.value = "-"
                                cell.font = regular_font
                            else:
                                insp_vals = reading.get("inspection_values", {})
                                val = insp_vals.get(p_label)
                                if val is None:
                                    val = insp_vals.get(p_id)
                                val_str = format_reading_value(val)
                                cell.value = val_str
                                cell.font = regular_font
                                
                                is_violated = False
                                cell_severity = None
                                
                                for a in active_alerts:
                                    if a.get("tank_id") == tid:
                                        label = a.get("constraint_label") or a.get("param_label") or ""
                                        if label.strip().lower() == p_label.strip().lower():
                                            alert_time = a.get("captured_at") or a.get("timestamp") or ""
                                            if alert_time.startswith(day_str):
                                                is_violated = True
                                                cell_severity = (a.get("constraint_severity") or a.get("severity") or "critical").lower()
                                                break
                                                
                                if not is_violated and has_take_action_keyword(val):
                                    is_violated = True
                                    cell_severity = "critical"
                                    
                                if is_violated:
                                    cell.font = bold_font
                                    cell.fill = red_fill if cell_severity == "critical" else yellow_fill
                                    
                curr_row += 1
            
            curr_row += 2
            
        ws.column_dimensions['A'].width = 25
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"[-] Error generating weekly inspection Excel: {e}")
        return False

def generate_weekly_alerts_pdf(filepath, start_date_str, end_date_str):
    """Generates a weekly alerts violation report PDF"""
    if not REPORTLAB_AVAILABLE:
        return False
        
    try:
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        
        weekly_alerts = []
        for a_id, a in alerts_data.items():
            try:
                alert_time = a.get("captured_at") or a.get("timestamp") or ""
                alert_date = alert_time.split("T")[0]
                if start_date_str <= alert_date <= end_date_str:
                    weekly_alerts.append(a)
            except:
                continue
                
        client_name = get_client_name(CONFIG['client_id'])
        generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        story = []
        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'DocTitle', parent=styles['Heading1'],
            fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#D32F2F'),
            alignment=1, spaceAfter=8
        )
        meta_style = ParagraphStyle(
            'DocMeta', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#555555'),
            alignment=1, spaceAfter=15
        )
        cell_style = ParagraphStyle(
            'CellText', parent=styles['Normal'],
            fontName='Helvetica', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#2D3748')
        )
        cell_style_bold = ParagraphStyle(
            'CellTextBold', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#2D3748')
        )
        header_style = ParagraphStyle(
            'HeaderStyle', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor('#FFFFFF'),
            alignment=1
        )
        
        story.append(Paragraph("VAPLI Weekly Alert Summary Report", title_style))
        story.append(Paragraph(f"Client: {client_name} | Period: {start_date_str} to {end_date_str} | Generated: {generated_at}", meta_style))
        story.append(Spacer(1, 10))
        
        if not weekly_alerts:
            story.append(Paragraph("<b>No alerts logged during this week period.</b>", ParagraphStyle('NoAlerts', parent=cell_style, textColor=colors.HexColor('#2F855A'))))
        else:
            headers = [
                Paragraph("Date Time", header_style),
                Paragraph("Asset Name", header_style),
                Paragraph("Severity", header_style),
                Paragraph("Parameter", header_style),
                Paragraph("Value", header_style),
                Paragraph("Alert Message", header_style)
            ]
            
            table_data = [headers]
            table_commands = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#9B2C2C')),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]
            
            row_idx = 1
            for a in sorted(weekly_alerts, key=lambda x: x.get("timestamp") or x.get("captured_at") or "", reverse=True):
                raw_time = a.get("captured_at") or a.get("timestamp") or ""
                try:
                    dt = parse_iso_datetime(raw_time)
                    time_str = dt.strftime("%d/%m/%y %H:%M") if dt else raw_time.replace("T", " ")[:14]
                except:
                    time_str = raw_time.replace("T", " ")[:14]
                    
                tank_name = a.get("tank_name", "Asset")
                tank_code = a.get("tank_code", "")
                display_name = f"{tank_name} ({tank_code})" if tank_code else tank_name
                
                sev = (a.get("constraint_severity") or a.get("severity") or "warning").lower()
                label = a.get("constraint_label") or a.get("param_label") or "-"
                val = a.get("violated_value") or a.get("param_value") or "-"
                msg = a.get("message") or a.get("alert_title") or "-"
                
                row = [
                    Paragraph(time_str, cell_style),
                    Paragraph(display_name, cell_style_bold),
                    Paragraph(sev.upper(), cell_style_bold),
                    Paragraph(label, cell_style),
                    Paragraph(str(val), cell_style_bold),
                    Paragraph(msg, cell_style)
                ]
                table_data.append(row)
                
                color_hex = '#FFCDD2' if sev == 'critical' else ('#FFF9C4' if sev == 'warning' else '#BBDEFB')
                table_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(color_hex)))
                row_idx += 1
                
            t = Table(table_data, colWidths=[75, 100, 50, 75, 50, 190])
            t.setStyle(TableStyle(table_commands))
            story.append(t)
            
        doc.build(story)
        return True
    except Exception as e:
        print(f"[-] Error generating weekly alerts PDF: {e}")
        return False

def generate_weekly_alerts_excel(filepath, start_date_str, end_date_str):
    """Generates a weekly alerts violation report Excel sheet"""
    if not OPENPYXL_AVAILABLE:
        return False
        
    try:
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        
        weekly_alerts = []
        for a_id, a in alerts_data.items():
            try:
                alert_time = a.get("captured_at") or a.get("timestamp") or ""
                alert_date = alert_time.split("T")[0]
                if start_date_str <= alert_date <= end_date_str:
                    weekly_alerts.append(a)
            except:
                continue
                
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Alerts"
        
        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="D32F2F")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        
        red_header_fill = PatternFill(start_color="9B2C2C", end_color="9B2C2C", fill_type="solid")
        red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        blue_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        left_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')

        # Title
        ws.append(["VAPLI Weekly Alert Summary Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.row_dimensions[1].height = 30
        
        client_name = get_client_name(CONFIG['client_id'])
        ws.append([f"Client: {client_name} | Period: {start_date_str} to {end_date_str}"])
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.append([]) # spacer
        
        headers = ["Date Time", "Asset Name", "Severity", "Parameter Checked", "Violated Value", "Alert Message"]
        ws.append(headers)
        ws.row_dimensions[4].height = 24
        
        for col_idx in range(1, 7):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = red_header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        curr_row = 5
        for a in sorted(weekly_alerts, key=lambda x: x.get("timestamp") or x.get("captured_at") or "", reverse=True):
            raw_time = a.get("captured_at") or a.get("timestamp") or ""
            try:
                dt = parse_iso_datetime(raw_time)
                time_str = dt.strftime("%d/%m/%y %H:%M") if dt else raw_time.replace("T", " ")[:14]
            except:
                time_str = raw_time.replace("T", " ")[:14]
                
            tank_name = a.get("tank_name", "Asset")
            tank_code = a.get("tank_code", "")
            display_name = f"{tank_name} ({tank_code})" if tank_code else tank_name
            
            sev = (a.get("constraint_severity") or a.get("severity") or "warning").lower()
            label = a.get("constraint_label") or a.get("param_label") or "-"
            val = a.get("violated_value") or a.get("param_value") or "-"
            msg = a.get("message") or a.get("alert_title") or "-"
            
            ws.append([time_str, display_name, sev.upper(), label, val, msg])
            ws.row_dimensions[curr_row].height = 20
            
            fill_to_use = red_fill if sev == "critical" else (yellow_fill if sev == "warning" else blue_fill)
            
            for col_idx in range(1, 7):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = bold_font if col_idx in [2, 3, 5] else regular_font
                cell.fill = fill_to_use
                cell.border = thin_border
                cell.alignment = center_align if col_idx in [1, 3, 5] else left_align
                
            curr_row += 1
            
        # Adjust Column Widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 35
        
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"[-] Error generating weekly alerts Excel: {e}")
        return False

# ==============================================================================
# WEB SERVER FOR FEEDBACK SUBMISSION
# ==============================================================================

def _load_admin_html():
    """Reads admin.html from the Server directory at request time (single source of truth)"""
    try:
        html_path = os.path.join(os.path.dirname(__file__), "admin.html")
        with open(html_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        # Fallback minimal page if file not found
        return "<html><body><h1>VAPLI Control Center</h1><p>admin.html not found in Server directory.</p></body></html>"


class FeedbackHTTPRequestHandler(BaseHTTPRequestHandler):
    """Zero-dependency HTTP Request Handler to serve feedback forms and provide a server control panel"""

    def log_message(self, format, *args):
        # Suppress logging in console to avoid cluttering CLI menus
        pass

    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)


        if parsed_url.path == "/health":
            # Keep-alive health check endpoint for Cloud Run / UptimeRobot
            import time as pytime
            uptime_secs = int(pytime.time() - (CONFIG.get("server_start_time") or pytime.time()))
            health_data = {
                "status": "ok",
                "uptime_seconds": uptime_secs,
                "client_id": CONFIG["client_id"],
                "env_mode": CONFIG["env_mode"],
                "timestamp": datetime.utcnow().isoformat() + "Z"
            }
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(json.dumps(health_data).encode("utf-8"))

        elif parsed_url.path in ["/feedback", "/feedback.html"]:
            try:
                filepath = os.path.join(os.path.dirname(__file__), "feedback.html")
                with open(filepath, "r", encoding="utf-8") as f:
                    html_content = f.read()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_content.encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(f"Error loading form: {e}".encode("utf-8"))

        elif parsed_url.path in ["/admin", "/admin.html", "/"]:
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(_load_admin_html().encode("utf-8"))

        elif parsed_url.path == "/api/status":
            client_list = []
            try:
                url = f"{CONFIG['db_url']}/clients.json"
                req = urllib.request.Request(url)
                with urllib.request.urlopen(req, timeout=5) as res:
                    clients_data = json.loads(res.read().decode("utf-8"))
                clients_dict = normalize_to_dict(clients_data)
                for cid, cinfo in clients_dict.items():
                    if isinstance(cinfo, dict):
                        db_key = cinfo.get("db_key")
                        name = cinfo.get("name") or db_key or cid
                        if db_key:
                            client_list.append({"db_key": db_key, "name": name})
            except Exception as e:
                pass

            groups_list = []
            try:
                tree_data = normalize_to_dict(fetch_db("tank_tree"))
                nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
                folders = [n for n in nodes if n.get("type") == "folder"]
                folders.sort(key=lambda x: x.get("order", 0))
                for f in folders:
                    groups_list.append(f.get("name"))
            except Exception as e:
                pass

            syd_time = get_sydney_time().strftime("%Y-%m-%d %I:%M:%S %p")
            ast_time = get_ast_time().strftime("%Y-%m-%d %I:%M:%S %p")
            ist_time = get_ist_time().strftime("%Y-%m-%d %I:%M:%S %p")
            status_data = {
                "client_id": CONFIG["client_id"],
                "env_mode": CONFIG["env_mode"],
                "sydney_time": syd_time,
                "ast_time": ast_time,
                "ist_time": ist_time,
                "clients": client_list,
                "groups": groups_list,
                "smtp_configured": bool(CONFIG["smtp_email"] and CONFIG["smtp_password"]),
                "schedule": {
                    "ast_hour": CONFIG.get("ast_reminder_hour", 15),
                    "ast_minute": CONFIG.get("ast_reminder_minute", 0),
                    "ast_enabled": CONFIG.get("ast_reminder_enabled", True),
                    "ist_hour": CONFIG.get("ist_reminder_hour", 15),
                    "ist_minute": CONFIG.get("ist_reminder_minute", 0),
                    "ist_enabled": CONFIG.get("ist_reminder_enabled", True),
                    "syd_enabled": CONFIG.get("syd_reminder_enabled", True),
                }
            }
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(status_data).encode("utf-8"))

        elif parsed_url.path == "/api/schedule":
            schedule_data = {
                "ast_hour": CONFIG.get("ast_reminder_hour", 15),
                "ast_minute": CONFIG.get("ast_reminder_minute", 0),
                "ast_enabled": CONFIG.get("ast_reminder_enabled", True),
                "ist_hour": CONFIG.get("ist_reminder_hour", 15),
                "ist_minute": CONFIG.get("ist_reminder_minute", 0),
                "ist_enabled": CONFIG.get("ist_reminder_enabled", True),
                "syd_enabled": CONFIG.get("syd_reminder_enabled", True),
                "ast_current": get_ast_time().strftime("%Y-%m-%d %H:%M:%S"),
                "ist_current": get_ist_time().strftime("%Y-%m-%d %H:%M:%S"),
                "syd_current": get_sydney_time().strftime("%Y-%m-%d %H:%M:%S"),
            }
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(schedule_data).encode("utf-8"))

        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")

    def do_POST(self):

        parsed_url = urllib.parse.urlparse(self.path)
        
        if parsed_url.path == "/api/trigger":
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length).decode('utf-8')
            try:
                data = json.loads(post_data)
                action = data.get("action")
                params = data.get("params", {})
                
                success = False
                message = ""
                
                if action == "set_client":
                    client_id = params.get("client_id")
                    if client_id:
                        CONFIG["client_id"] = client_id
                        success = True
                        message = f"Client ID updated to {client_id}"
                    else:
                        message = "Missing client_id parameter"
                        
                elif action == "set_mode":
                    mode = params.get("mode")
                    if mode in ["production", "development"]:
                        CONFIG["env_mode"] = mode
                        success = True
                        message = f"Database Mode updated to {mode.upper()}"
                    else:
                        message = "Invalid database mode"
                        
                elif action == "test_smtp":
                    if not CONFIG["smtp_email"] or not CONFIG["smtp_password"]:
                        message = "SMTP credentials not configured"
                    else:
                        try:
                            server = smtplib.SMTP(CONFIG["smtp_server"], CONFIG["smtp_port"])
                            server.starttls()
                            server.login(CONFIG["smtp_email"], CONFIG["smtp_password"])
                            server.quit()
                            success = True
                            message = "SMTP connection test successful!"
                        except Exception as e:
                            message = f"SMTP test failed: {e}"
                            
                elif action == "send_daily":
                    threading.Thread(target=execute_daily_reports_email).start()
                    success = True
                    message = "Daily report email trigger queued"
                    
                elif action == "send_compliance":
                    threading.Thread(target=execute_missing_tanks_email).start()
                    success = True
                    message = "Compliance alert email trigger queued"
                    
                elif action == "send_active_alerts":
                    threading.Thread(target=execute_alerts_notification_email).start()
                    success = True
                    message = "Active alerts notification email trigger queued"
                    
                elif action == "send_weekly_inspection":
                    threading.Thread(target=execute_weekly_reports_email).start()
                    success = True
                    message = "Weekly inspection report email trigger queued"
                    
                elif action == "send_weekly_alerts":
                    threading.Thread(target=execute_weekly_alerts_email).start()
                    success = True
                    message = "Weekly alerts report email trigger queued"
                    
                elif action == "trigger_fcm":
                    fcm_type = int(params.get("fcm_type", 1))
                    if fcm_type == 4:
                        group_name = params.get("group_name")
                        if group_name:
                            threading.Thread(target=trigger_fcm_group_reminder, args=(group_name,)).start()
                            success = True
                            message = f"FCM Group Reminder queued for group '{group_name}'"
                        else:
                            message = "Missing group_name parameter"
                    elif fcm_type in (1, 5):
                        # type 1 = Daily Inspection Reminder, type 5 = Smart Tank Reminder
                        # Both use the rule-based tank reminder (0% / <50% / >=50% per group)
                        threading.Thread(target=trigger_fcm_tank_reminder, args=("Manual",)).start()
                        success = True
                        message = "Smart Tank Reminder FCM queued (rule-based per-group status)"
                    else:
                        # Types 2 (alert+image) and 3 (alert no image) use the test notifications
                        threading.Thread(target=trigger_fcm_notification, args=(fcm_type,)).start()
                        success = True
                        message = f"FCM test notification type {fcm_type} queued"
                else:
                    message = f"Unknown action: {action}"

                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps({"success": success, "message": message}).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "message": f"Error: {e}"}).encode("utf-8"))

        elif parsed_url.path == "/api/schedule":
            # Update schedule configuration
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length).decode('utf-8')
            try:
                data = json.loads(post_data)
                updated = []
                if "ast_hour" in data:
                    CONFIG["ast_reminder_hour"] = int(data["ast_hour"])
                    updated.append(f"AST hour={data['ast_hour']}")
                if "ast_minute" in data:
                    CONFIG["ast_reminder_minute"] = int(data["ast_minute"])
                    updated.append(f"AST minute={data['ast_minute']}")
                if "ast_enabled" in data:
                    CONFIG["ast_reminder_enabled"] = bool(data["ast_enabled"])
                    updated.append(f"AST enabled={data['ast_enabled']}")
                if "ist_hour" in data:
                    CONFIG["ist_reminder_hour"] = int(data["ist_hour"])
                    updated.append(f"IST hour={data['ist_hour']}")
                if "ist_minute" in data:
                    CONFIG["ist_reminder_minute"] = int(data["ist_minute"])
                    updated.append(f"IST minute={data['ist_minute']}")
                if "ist_enabled" in data:
                    CONFIG["ist_reminder_enabled"] = bool(data["ist_enabled"])
                    updated.append(f"IST enabled={data['ist_enabled']}")
                if "syd_enabled" in data:
                    CONFIG["syd_reminder_enabled"] = bool(data["syd_enabled"])
                    updated.append(f"SYD enabled={data['syd_enabled']}")
                msg = "Schedule updated: " + ", ".join(updated) if updated else "No changes"
                log_to_db(msg, "info")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps({"success": True, "message": msg}).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(json.dumps({"success": False, "message": f"Error: {e}"}).encode("utf-8"))

        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")


def fetch_db_global(path):
    """Fetches global database node without client-scoping prefixes"""
    normalized = path.lstrip("/")
    url = f"{CONFIG['db_url']}/{normalized}.json"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as res:
            return json.loads(res.read().decode("utf-8"))
    except Exception as e:
        return None

def write_db_global(path, data, method="PUT"):
    """Writes global database node without client-scoping prefixes"""
    normalized = path.lstrip("/")
    url = f"{CONFIG['db_url']}/{normalized}.json"
    try:
        payload = json.dumps(data).encode("utf-8")
        req = urllib.request.Request(url, data=payload, method=method)
        req.add_header("Content-Type", "application/json")
        with urllib.request.urlopen(req, timeout=10) as res:
            return json.loads(res.read().decode("utf-8"))
    except Exception as e:
        return None

def log_to_db(message, log_type="info"):
    """Logs server activities to the database for the hosted dashboard to display"""
    print(f"[*] {message}")
    log_entry = {
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "message": message,
        "type": log_type
    }
    write_db_global("server_logs", log_entry, method="POST")
    
    # Trim logs to prevent Firebase RTDB bloating (keep last 50)
    try:
        logs = fetch_db_global("server_logs")
        if logs and isinstance(logs, dict) and len(logs) > 50:
            sorted_keys = sorted(logs.keys())
            for k in sorted_keys[:-50]:
                write_db_global(f"server_logs/{k}", None)
    except Exception:
        pass

def execute_backend_action(action, params):
    """Executes backend python routines based on DB triggers"""
    try:
        if action == "set_client":
            client_id = params.get("client_id")
            if client_id:
                CONFIG["client_id"] = client_id
                return True, f"Client ID updated to {client_id}"
            return False, "Missing client_id parameter"
            
        elif action == "set_mode":
            mode = params.get("mode")
            if mode in ["production", "development"]:
                CONFIG["env_mode"] = mode
                return True, f"Database Mode updated to {mode.upper()}"
            return False, "Invalid database mode"
            
        elif action == "test_smtp":
            if not CONFIG["smtp_email"] or not CONFIG["smtp_password"]:
                return False, "SMTP credentials not configured"
            server = smtplib.SMTP(CONFIG["smtp_server"], CONFIG["smtp_port"])
            server.starttls()
            server.login(CONFIG["smtp_email"], CONFIG["smtp_password"])
            server.quit()
            return True, "SMTP connection test successful!"
            
        elif action == "send_daily":
            threading.Thread(target=execute_daily_reports_email).start()
            return True, "Daily report email trigger queued"
            
        elif action == "send_compliance":
            threading.Thread(target=execute_missing_tanks_email).start()
            return True, "Compliance alert email trigger queued"
            
        elif action == "send_active_alerts":
            threading.Thread(target=execute_alerts_notification_email).start()
            return True, "Active alerts notification email trigger queued"
            
        elif action == "send_weekly_inspection":
            threading.Thread(target=execute_weekly_reports_email).start()
            return True, "Weekly inspection report email trigger queued"
            
        elif action == "send_weekly_alerts":
            threading.Thread(target=execute_weekly_alerts_email).start()
            return True, "Weekly alerts report email trigger queued"
            
        elif action == "trigger_fcm":
            fcm_type = int(params.get("fcm_type", 1))
            if fcm_type == 4:
                group_name = params.get("group_name")
                if group_name:
                    threading.Thread(target=trigger_fcm_group_reminder, args=(group_name,)).start()
                    return True, f"FCM Group Reminder queued for group '{group_name}'"
                return False, "Missing group_name parameter"
            elif fcm_type in (1, 5):
                # type 1 and 5 both use the rule-based smart tank reminder
                threading.Thread(target=trigger_fcm_tank_reminder, args=("Manual",)).start()
                return True, "Smart Tank Reminder FCM queued (rule-based per-group status)"
            else:
                # Types 2 and 3 are test alert notifications
                threading.Thread(target=trigger_fcm_notification, args=(fcm_type,)).start()
                return True, f"FCM test notification type {fcm_type} queued"
                
        return False, f"Unknown action: {action}"
    except Exception as e:
        return False, f"Error: {e}"

def db_trigger_monitor_thread():
    """Polls Firebase RTDB for overrides (client_id, mode), pending triggers, and writes server heartbeats/logs"""
    import time as pytime
    pytime.sleep(3)
    log_to_db("Database Trigger Monitor worker started.", "info")
    
    while True:
        try:
            # 1. Heartbeat & Live Sydney Time update
            syd_time = get_sydney_time().strftime("%Y-%m-%d %I:%M:%S %p")
            state_update = {
                "client_id": CONFIG["client_id"],
                "env_mode": CONFIG["env_mode"],
                "sydney_time": syd_time,
                "smtp_configured": bool(CONFIG["smtp_email"] and CONFIG["smtp_password"]),
                "last_active": datetime.utcnow().isoformat() + "Z"
            }
            write_db_global("server_state", state_update)
            
            # 2. Check for state overrides from DB
            state_db = fetch_db_global("server_state_override")
            if state_db:
                client_id = state_db.get("client_id")
                env_mode = state_db.get("env_mode")
                if client_id and client_id != CONFIG["client_id"]:
                    CONFIG["client_id"] = client_id
                    log_to_db(f"Client ID updated via DB override to {client_id}", "success")
                if env_mode and env_mode != CONFIG["env_mode"]:
                    CONFIG["env_mode"] = env_mode
                    log_to_db(f"Database Mode updated via DB override to {env_mode.upper()}", "success")
                # Clear override after applying
                write_db_global("server_state_override", None)
                
            # 3. Check for pending triggers in /triggers
            triggers = fetch_db_global("triggers")
            if triggers and isinstance(triggers, dict):
                for tid, tdata in triggers.items():
                    if tdata and isinstance(tdata, dict) and tdata.get("status") == "pending":
                        action = tdata.get("action")
                        params = tdata.get("params", {})
                        
                        log_to_db(f"Executing trigger '{action}' from database...", "info")
                        write_db_global(f"triggers/{tid}/status", "running")
                        
                        success, msg = execute_backend_action(action, params)
                        
                        write_db_global(f"triggers/{tid}/status", "success" if success else "failed")
                        write_db_global(f"triggers/{tid}/message", msg)
                        log_to_db(f"Trigger '{action}' result: {'SUCCESS' if success else 'FAILED'} - {msg}", "success" if success else "error")
        except Exception as e:
            pass
        pytime.sleep(3) # poll every 3 seconds

def trigger_fcm_group_reminder(group_name):
    """Sends group inspection reminder FCM to active client topic directly"""
    if not initialize_firebase_admin():
        return
    import firebase_admin
    from firebase_admin import messaging
    
    topic_name = CONFIG["client_id"]
    title = f"🔍 Group Reminder: {group_name}"
    body = f"Attention: Lubrication readings for the '{group_name}' group are due. Please complete them today."
    
    try:
        notification = messaging.Notification(title=title, body=body)
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": "4"
        }
        message = messaging.Message(notification=notification, data=data_payload, topic=topic_name)
        response = messaging.send(message)
        print(f"\n[+] FCM Group Reminder sent successfully! Topic: {topic_name}, Group: {group_name}")
    except Exception as e:
        print(f"[-] FCM Group Reminder failed: {e}")

def get_sydney_time():
    """Calculates Sydney time (AEST/AEDT) from UTC without external pytz dependency"""
    utc_now = datetime.utcnow()
    year = utc_now.year
    first_sun_apr = datetime(year, 4, 1)
    while first_sun_apr.weekday() != 6:
        first_sun_apr += timedelta(days=1)
    first_sun_oct = datetime(year, 10, 1)
    while first_sun_oct.weekday() != 6:
        first_sun_oct += timedelta(days=1)
    if first_sun_apr <= utc_now < first_sun_oct:
        offset = 10
    else:
        offset = 11
    return utc_now + timedelta(hours=offset)

def get_ast_time():
    """Returns current time in Arabia Standard Time (AST, UTC+3) — used for Saudi Arabia"""
    return datetime.utcnow() + timedelta(hours=3)

def get_ist_time():
    """Returns current time in India Standard Time (IST, UTC+5:30)"""
    return datetime.utcnow() + timedelta(hours=5, minutes=30)

def trigger_fcm_tank_reminder(timezone_label="", target_date=None):
    """Sends smart FCM notification with per-group reading completion status and bullet-point pending list"""
    if not initialize_firebase_admin():
        print("[-] Firebase Admin not initialized. Cannot send tank reminder FCM.")
        return

    import firebase_admin
    from firebase_admin import messaging

    topic_name = CONFIG["client_id"]
    
    # Standardize date calculations to prevent timezone mismatches
    if not target_date:
        if timezone_label == "AST":
            target_date = get_ast_time().strftime("%Y-%m-%d")
        elif timezone_label == "IST":
            target_date = get_ist_time().strftime("%Y-%m-%d")
        elif timezone_label == "SYD":
            target_date = get_sydney_time().strftime("%Y-%m-%d")
        else:
            # Manual or generic: Default to the AST workday date
            target_date = get_ast_time().strftime("%Y-%m-%d")
            
    today_str = target_date

    try:
        tanks_data = normalize_to_dict(fetch_db("tanks"))
        readings_data = normalize_to_dict(fetch_db("readings"))
        tree_data = normalize_to_dict(fetch_db("tank_tree"))
    except Exception as e:
        print(f"[-] Tank Reminder FCM: Failed to fetch data - {e}")
        return

    today_readings = set()
    for r in readings_data.values():
        try:
            if r and r.get("captured_at", "").split("T")[0] == today_str:
                today_readings.add(r.get("tank_id"))
        except:
            continue

    nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
    top_folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
    top_folders.sort(key=lambda x: x.get("order", 0))

    def get_folder_tanks(folder_id):
        tank_ids = []
        queue = [folder_id]
        while queue:
            curr = queue.pop(0)
            children = [n for n in nodes if n.get("parent_id") == curr]
            for c in children:
                if c.get("type") == "leaf" and c.get("tank_id"):
                    tank_ids.append(c.get("tank_id"))
                elif c.get("type") == "folder":
                    queue.append(c.get("id"))
        return tank_ids

    body_lines = []
    for folder in top_folders:
        t_ids = get_folder_tanks(folder.get("id"))
        if not t_ids:
            continue

        folder_tanks = [tanks_data[tid] for tid in t_ids if tid in tanks_data]
        if not folder_tanks:
            continue

        total_count = len(folder_tanks)
        inspected = [t for t in folder_tanks if t.get("id") in today_readings]
        uninspected = [t for t in folder_tanks if t.get("id") not in today_readings]
        inspected_count = len(inspected)
        percentage = (inspected_count / total_count * 100) if total_count > 0 else 0
        folder_name = folder.get("name", "Group")

        if inspected_count == 0:
            # 0% — no readings at all
            body_lines.append(f"{folder_name} has not taken Readings at all")
        elif percentage < 50:
            # < 50% — "Group C has not been Taken reading except: \nBulluet points pending list"
            # Shows the pending list (uninspected)
            body_lines.append(f"{folder_name} has not been Taken reading except:")
            for t in uninspected:
                name = t.get("tank_name", "Asset")
                code = t.get("tank_code", "")
                body_lines.append(f"  \u2022 {name}" + (f" ({code})" if code else ""))
        elif percentage < 100:
            # >= 50% — "Group B has been Taken reading except: \nBulluet points pending list"
            # Also shows the pending list (uninspected)
            body_lines.append(f"{folder_name} has been Taken reading except:")
            for t in uninspected:
                name = t.get("tank_name", "Asset")
                code = t.get("tank_code", "")
                body_lines.append(f"  \u2022 {name}" + (f" ({code})" if code else ""))
        else:
            # 100% complete — positive completion note
            body_lines.append(f"{folder_name}: All readings complete \u2713")

    if not body_lines:
        print(f"[*] Tank Reminder FCM: No group data found for client {topic_name}")
        return

    tz_tag = f" [{timezone_label}]" if timezone_label else ""
    title = f"\U0001f4cb 3 PM Inspection Reminder{tz_tag}"
    body = "\n".join(body_lines)

    try:
        notification = messaging.Notification(title=title, body=body)
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": "1"
        }
        message = messaging.Message(notification=notification, data=data_payload, topic=topic_name)
        messaging.send(message)
        log_to_db(f"Tank Reminder FCM sent{tz_tag} for topic: {topic_name}", "success")
        print(f"[+] Tank Reminder FCM sent successfully{tz_tag} for {topic_name}")
    except Exception as e:
        print(f"[-] Tank Reminder FCM failed{tz_tag}: {e}")

def send_email_sent_fcm(email_type, recipient_count=0):
    """Sends FCM confirmation notification to client topic after an email has been dispatched"""
    if not initialize_firebase_admin():
        return
    try:
        import firebase_admin
        from firebase_admin import messaging

        topic_name = CONFIG["client_id"]
        title = f"\U0001f4e7 Email Sent: {email_type}"
        body = f"Report dispatched to {recipient_count} recipient(s) successfully."

        notification = messaging.Notification(title=title, body=body)
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": "email_sent",
            "email_type": email_type
        }
        message = messaging.Message(notification=notification, data=data_payload, topic=topic_name)
        messaging.send(message)
        print(f"[+] Email Sent FCM notification dispatched: {email_type} ({recipient_count} recipients)")
    except Exception as e:
        print(f"[-] Email Sent FCM failed: {e}")


def auto_scheduler_thread():
    """Background thread that monitors Sydney Time and automatically triggers daily/weekly reports and reminders"""
    import time as pytime
    pytime.sleep(10)
    print("\n[+] Auto Scheduler started. Monitoring AST, IST, and Sydney Time...")

    # Dict to track what has been triggered per-day to prevent double-firing
    # Key format: "YYYY-MM-DD_TZ_HOUR_MINUTE"
    last_triggered = {}

    def already_triggered(date_key, tz, hour, minute):
        return f"{date_key}_{tz}_{hour}_{minute}" in last_triggered

    def mark_triggered(date_key, tz, hour, minute):
        last_triggered[f"{date_key}_{tz}_{hour}_{minute}"] = True

    while True:
        try:
            utc_now = datetime.utcnow()
            today_utc = utc_now.strftime("%Y-%m-%d")

            # --- AST (Arabia Standard Time, UTC+3) ---
            if CONFIG.get("ast_reminder_enabled", True):
                ast_now = get_ast_time()
                ast_h = CONFIG.get("ast_reminder_hour", 15)
                ast_m = CONFIG.get("ast_reminder_minute", 0)
                ast_date = ast_now.strftime("%Y-%m-%d")
                if ast_now.hour == ast_h and ast_now.minute >= ast_m and not already_triggered(ast_date, "AST", ast_h, ast_m):
                    print(f"\n[*] Scheduler: Triggering {ast_h}:{ast_m:02d} AST Reminder...")
                    mark_triggered(ast_date, "AST", ast_h, ast_m)
                    threading.Thread(target=trigger_fcm_tank_reminder, args=("AST", ast_date)).start()
                    # Emails: weekday = daily, weekend = weekly
                    if ast_now.weekday() < 5:
                        print("    AST: Weekday - Triggering Daily Reports + Missing Tanks...")
                        threading.Thread(target=execute_daily_reports_email, args=(ast_date,)).start()
                        threading.Thread(target=execute_missing_tanks_email, args=(ast_date,)).start()
                    else:
                        print("    AST: Weekend - Triggering Weekly Reports + Missing Tanks...")
                        threading.Thread(target=execute_weekly_reports_email).start()
                        threading.Thread(target=execute_weekly_alerts_email).start()
                        threading.Thread(target=execute_missing_tanks_email, args=(ast_date,)).start()

            # --- IST (India Standard Time, UTC+5:30) ---
            if CONFIG.get("ist_reminder_enabled", True):
                ist_now = get_ist_time()
                ist_h = CONFIG.get("ist_reminder_hour", 15)
                ist_m = CONFIG.get("ist_reminder_minute", 0)
                ist_date = ist_now.strftime("%Y-%m-%d")
                if ist_now.hour == ist_h and ist_now.minute >= ist_m and not already_triggered(ist_date, "IST", ist_h, ist_m):
                    print(f"\n[*] Scheduler: Triggering {ist_h}:{ist_m:02d} IST Reminder...")
                    mark_triggered(ist_date, "IST", ist_h, ist_m)
                    # IST: FCM reminder only (emails already sent via AST)
                    threading.Thread(target=trigger_fcm_tank_reminder, args=("IST", ist_date)).start()

            # --- Sydney Time (AEST/AEDT) ---
            if CONFIG.get("syd_reminder_enabled", True):
                syd_now = get_sydney_time()
                syd_date = syd_now.strftime("%Y-%m-%d")
                syd_h = syd_now.hour
                syd_m = syd_now.minute

                if syd_h == 13 and syd_m >= 0 and not already_triggered(syd_date, "SYD", 13, 0):
                    print(f"\n[*] Scheduler: Triggering 1:00 PM Sydney FCM Reminder...")
                    mark_triggered(syd_date, "SYD", 13, 0)
                    threading.Thread(target=trigger_fcm_tank_reminder, args=("SYD", syd_date)).start()

                elif syd_h == 15 and syd_m >= 0 and not already_triggered(syd_date, "SYD", 15, 0):
                    print(f"\n[*] Scheduler: Triggering 3:00 PM Sydney Reports & Reminder...")
                    mark_triggered(syd_date, "SYD", 15, 0)
                    threading.Thread(target=trigger_fcm_tank_reminder, args=("SYD", syd_date)).start()
                    if syd_now.weekday() < 5:
                        print("    SYD: Weekday - Daily Reports...")
                        threading.Thread(target=execute_daily_reports_email, args=(syd_date,)).start()
                        threading.Thread(target=execute_missing_tanks_email, args=(syd_date,)).start()
                    else:
                        print("    SYD: Weekend - Weekly Reports...")
                        threading.Thread(target=execute_weekly_reports_email).start()
                        threading.Thread(target=execute_weekly_alerts_email).start()
                        threading.Thread(target=execute_missing_tanks_email, args=(syd_date,)).start()

            # Prune old trigger keys (keep only today's entries)
            current_dates = {utc_now.strftime("%Y-%m-%d")}
            for tz_key in list(last_triggered.keys()):
                key_date = tz_key.split("_")[0]
                if key_date not in current_dates:
                    del last_triggered[tz_key]

        except Exception as e:
            print(f"[-] Scheduler error: {e}")
        pytime.sleep(30)


def realtime_alerts_monitor_thread():
    """Polls the DB alerts node every 15 seconds. If a new alert is created, triggers FCM notification with details, and updates notified=True"""
    import time as pytime
    pytime.sleep(5)
    print("\n[+] Real-time Alert Notification Monitor started in background.")
    
    while True:
        try:
            client_id = CONFIG.get("client_id")
            if client_id:
                alerts_data = normalize_to_dict(fetch_db("alerts"))
                if alerts_data:
                    for alert_id, alert in alerts_data.items():
                        if not alert or not isinstance(alert, dict):
                            continue
                        if not alert.get("resolved", False) and not alert.get("notified", False):
                            success = send_realtime_alert_fcm(alert_id, alert)
                            if success:
                                write_db(f"alerts/{alert_id}/notified", True)
                                print(f"[+] Alert {alert_id} notified and marked in DB.")
        except Exception as e:
            pass
        pytime.sleep(15)

def send_realtime_alert_fcm(alert_id, alert):
    """Sends FCM notification for a real alert payload, resolving image URLs.
    Image priority: violation_image (Cloudinary from reading page) -> image_url -> img_url"""
    if not initialize_firebase_admin():
        return False

    import firebase_admin
    from firebase_admin import messaging

    topic_name = CONFIG["client_id"]
    tank_name = alert.get("tank_name", "Asset")
    tank_code = alert.get("tank_code", "")
    param = alert.get("constraint_label") or alert.get("param_label") or "Parameter"
    val = alert.get("violated_value") or alert.get("param_value") or "-"
    msg = alert.get("message") or alert.get("alert_title") or "Threshold Violated"
    severity = (alert.get("constraint_severity") or alert.get("severity") or "critical").upper()

    emoji = "\U0001f6a8" if severity == "CRITICAL" else "\u26a0\ufe0f"
    title = f"{emoji} {severity} Alert: {tank_name} ({tank_code})"
    body = f"{param} is out of limits: {val} - {msg}"

    # Priority: violation_image (from reading_entry_screen Cloudinary upload) > image_url > img_url
    image_url = alert.get("violation_image") or alert.get("image_url") or alert.get("img_url")

    try:
        notification = messaging.Notification(
            title=title,
            body=body,
            image=image_url
        )
        
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": "2",
            "alert_id": alert_id
        }
        
        message = messaging.Message(
            notification=notification,
            data=data_payload,
            topic=topic_name
        )
        
        messaging.send(message)
        print(f"\n[+] Real-time Alert FCM Sent: {title}")
        if image_url:
            print(f"    Image: {image_url}")
        return True
    except Exception as e:
        print(f"[-] Failed to send real-time alert FCM: {e}")
        return False

def start_web_server():
    """Starts the feedback web server in a separate background thread"""
    def run():
        host = CONFIG["server_host"]
        port = CONFIG["server_port"]
        server = HTTPServer((host, port), FeedbackHTTPRequestHandler)
        print(f"[+] Feedback Web Server started on http://{host if host != '0.0.0.0' else 'localhost'}:{port}")
        server.serve_forever()
    
    t = threading.Thread(target=run, daemon=True)
    t.start()

# ==============================================================================
# CLI MENU / CONTROLLER
# ==============================================================================

def send_compliance_feedback_email(unique_id, entry):
    """Sends the submitted compliance feedback report to the Report Receivers"""
    settings = fetch_db("settings/Report_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))
    if not receivers:
        print("[-] No report receivers defined. Cannot send compliance feedback email.")
        return False
        
    date_str = entry.get("date")
    submitted_by = entry.get("submitted_by")
    reason = entry.get("reason")
    comments = entry.get("comments")
    tanks_by_group = entry.get("tanks_by_group", {})
    
    tanks_html = ""
    for group_name, tank_list in tanks_by_group.items():
        tanks_html += f"<div class='group-header' style='font-weight:600; color:#2D3748; margin-top:10px; font-size:13px;'>{group_name}</div><ul class='tanks-list' style='margin:0; padding-left:20px; font-size:13px; line-height:1.5;'>"
        for t in tank_list:
            tanks_html += f"<li>{t.get('name')} ({t.get('code')})</li>"
        tanks_html += "</ul>"
        
    body = f"""
    <p>Dear Administrator,</p>
    <p>A missed inspection compliance report has been submitted by <b>{submitted_by}</b> for the date <b>{date_str}</b>.</p>
    
    <div class="section-title" style="font-size:14px; font-weight:600; color:#2B6CB0; margin:20px 0 10px 0; text-transform:uppercase; letter-spacing:0.5px;">Compliance Details</div>
    <table class="meta-table" style="width:100%; border-collapse:collapse; margin-bottom:20px;">
        <tr>
            <td class="label" style="padding:6px 0; font-size:13px; border-bottom:1px solid #EDF2F7; width:30%; font-weight:600; color:#4A5568;">Submitted By:</td>
            <td style="padding:6px 0; font-size:13px; border-bottom:1px solid #EDF2F7;">{submitted_by}</td>
        </tr>
        <tr>
            <td class="label" style="padding:6px 0; font-size:13px; border-bottom:1px solid #EDF2F7; width:30%; font-weight:600; color:#4A5568;">Date:</td>
            <td style="padding:6px 0; font-size:13px; border-bottom:1px solid #EDF2F7;">{date_str}</td>
        </tr>
        <tr>
            <td class="label" style="padding:6px 0; font-size:13px; border-bottom:1px solid #EDF2F7; width:30%; font-weight:600; color:#4A5568;">Reason:</td>
            <td style="padding:6px 0; font-size:13px; border-bottom:1px solid #EDF2F7;"><b>{reason}</b></td>
        </tr>
    </table>
    
    <div class="section-title" style="font-size:14px; font-weight:600; color:#2B6CB0; margin:20px 0 10px 0; text-transform:uppercase; letter-spacing:0.5px;">Comments / Explanation</div>
    <div class="content-box" style="background-color:#F7FAFC; border:1px solid #EDF2F7; border-radius:6px; padding:15px; font-size:13px; line-height:1.5; margin-bottom:20px;">
        {comments.replace('\n', '<br/>')}
    </div>
    
    <div class="section-title" style="font-size:14px; font-weight:600; color:#2B6CB0; margin:20px 0 10px 0; text-transform:uppercase; letter-spacing:0.5px;">Missed Assets Covered by This Report</div>
    <div class="content-box" style="background-color:#FFFFFF; border:1px solid #E2E8F0; border-radius:6px; padding:15px; font-size:13px; line-height:1.5; margin-bottom:20px;">
        {tanks_html}
    </div>
    """
    
    client_name = get_client_name(CONFIG['client_id'])
    subject = f"Missed Inspection Reason Submitted - {client_name}"
    html_email = build_modern_template("Supervisor Compliance Reason", "Missed Inspection Report", body)
    
    return send_email(receivers, subject, html_email)

def feedback_monitor_thread():
    """Background thread that monitors the database for new compliance feedback submissions and emails them immediately"""
    # Wait a few seconds to let main setup complete
    import time as pytime
    pytime.sleep(3)
    print("\n[+] Compliance Feedback Monitor started in background.")
    while True:
        try:
            if CONFIG.get("smtp_email") and CONFIG.get("smtp_password") and CONFIG.get("db_url") and CONFIG.get("client_id"):
                today_str = datetime.now().strftime("%Y-%m-%d")
                reasons_data = fetch_db(f"MissingReasons/{today_str}")
                reasons_dict = normalize_to_dict(reasons_data)
                if reasons_dict:
                    for unique_id, entry in reasons_dict.items():
                        if not entry or not isinstance(entry, dict):
                            continue
                        if not entry.get("email_sent", False):
                            submitted_by = entry.get("submitted_by", "")
                            signature = entry.get("signature", "")
                            
                            import hashlib
                            secret_salt = "vapli_compliance_secure_salt"
                            expected_sig = hashlib.sha256((submitted_by + secret_salt).encode('utf-8')).hexdigest()
                            
                            if signature != expected_sig:
                                print(f"\n[!] WARNING: Fraudulent compliance feedback detected for user '{submitted_by}' (invalid signature)!")
                                write_db(f"MissingReasons/{today_str}/{unique_id}/email_sent", "fraud")
                                continue
                                
                            print(f"\n[+] Verified compliance feedback from supervisor: {submitted_by}!")
                            success = send_compliance_feedback_email(unique_id, entry)
                            if success:
                                write_db(f"MissingReasons/{today_str}/{unique_id}/email_sent", True)
                                print(f"[+] Marked feedback {unique_id} as emailed in DB.")
        except Exception as e:
            # Silently catch database errors in background thread
            pass
        pytime.sleep(15)

# Global Firebase Admin initialized flag
FIREBASE_ADMIN_INITIALIZED = False

def initialize_firebase_admin():
    """Initializes Firebase Admin SDK using service-account.json"""
    global FIREBASE_ADMIN_INITIALIZED
    if FIREBASE_ADMIN_INITIALIZED:
        return True
        
    try:
        import firebase_admin
        from firebase_admin import credentials
    except ImportError:
        print("[-] 'firebase-admin' library is not installed. Run: pip install firebase-admin")
        return False
        
    # Look in parent folder first, then local Server folder
    service_account_path = os.path.join(os.path.dirname(__file__), "..", "service-account.json")
    if not os.path.exists(service_account_path):
        service_account_path = os.path.join(os.path.dirname(__file__), "service-account.json")
        
    if not os.path.exists(service_account_path):
        print("[-] Firebase Service Account key file ('service-account.json') not found.")
        print("    Please download it from Firebase Console -> Project Settings -> Service Accounts,")
        print("    rename it to 'service-account.json', and place it in the project root folder.")
        return False
        
    try:
        cred = credentials.Certificate(service_account_path)
        firebase_admin.initialize_app(cred, {
            'databaseURL': CONFIG["db_url"]
        })
        FIREBASE_ADMIN_INITIALIZED = True
        print("[+] Firebase Admin SDK initialized successfully.")
        return True
    except Exception as e:
        print(f"[-] Failed to initialize Firebase Admin: {e}")
        return False

def trigger_fcm_notification(notification_type):
    """Triggers one of four modern and aesthetic FCM notification payloads to the client topic"""
    if not initialize_firebase_admin():
        return
        
    import firebase_admin
    from firebase_admin import messaging
    
    topic_name = CONFIG["client_id"]
    client_name = get_client_name(topic_name)
    
    title = ""
    body = ""
    image_url = None
    
    if notification_type == 1:
        title = "📋 Daily Inspection Reminder"
        body = f"Please complete today's lubrication readings for all configured assets at {client_name}. Let's maintain a 100% compliance rate!"
        
    elif notification_type == 2:
        # Alert with Image
        title = "⚠️ Alert: PM6 M704 - Oil Level Low"
        body = f"Critical violation detected at {client_name}: Oil level below 20% (Value: 15% - Take Action)."
        image_url = "https://picsum.photos/400/200" # Standard Android-renderable landscape test image
        
    elif notification_type == 3:
        # Alert without Image
        title = "⚠️ Alert: PM6 Kirma Hydraulic - High Temp"
        body = f"Warning violation detected at {client_name}: Temperature exceeded 75°C (Value: 78°C)."
        image_url = None
        
    elif notification_type == 4:
        # Group Reminder
        tree_data = normalize_to_dict(fetch_db("tank_tree"))
        nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
        folders = [n for n in nodes if n.get("type") == "folder"]
        folders.sort(key=lambda x: x.get("order", 0))
        
        if not folders:
            print("[-] No groups found in the database for this client.")
            return
            
        print("\nSelect a Group to send a reminder for:")
        for idx, folder in enumerate(folders, 1):
            print(f"  {idx}. {folder.get('name')}")
            
        try:
            sel = input(f"Select group number (1-{len(folders)}): ").strip()
            sel_idx = int(sel) - 1
            if 0 <= sel_idx < len(folders):
                selected_group = folders[sel_idx]
                group_name = selected_group.get("name")
                title = f"🔍 Group Reminder: {group_name}"
                body = f"Attention: Lubrication readings for the '{group_name}' group are due. Please complete them today."
            else:
                print("[-] Invalid group selection.")
                return
        except ValueError:
            print("[-] Invalid input. Must enter a number.")
            return
            
    try:
        notification = messaging.Notification(
            title=title,
            body=body,
            image=image_url
        )
        
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": str(notification_type)
        }
        
        message = messaging.Message(
            notification=notification,
            data=data_payload,
            topic=topic_name
        )
        
        response = messaging.send(message)
        print(f"\n[+] FCM message sent successfully! Response ID: {response}")
        print(f"    Topic: {topic_name}")
        print(f"    Title: {title}")
        print(f"    Body: {body}")
        if image_url:
            print(f"    Image: {image_url}")
    except Exception as e:
        print(f"[-] FCM publish failed: {e}")

def print_menu():
    print("\n" + "="*45)
    print(f"       VAPLI Automation Mail Server Menu")
    print("="*45)
    print(f" 1. Select Client ID     (Current: {CONFIG['client_id']})")
    print(f" 2. Select Database Mode (Current: {CONFIG['env_mode'].upper()})")
    print(f" 3. Test SMTP Connection")
    print(f" 4. Send Today's Inspection & Alert Reports")
    print(f" 5. Send Missing Tanks Compliance Alert")
    print(f" 6. Send Active Unresolved Alerts Notification")
    print(f" 7. Send Weekly Inspection Report")
    print(f" 8. Send Weekly Alerts Report")
    print(f" 9. Trigger FCM Notifications")
    print(" 10. Exit")
    print("="*45)

def main():
    import time as pytime
    load_env()
    CONFIG["server_start_time"] = pytime.time()

    # Start web server (admin panel + API) in background
    start_web_server()

    # Start compliance feedback polling thread in background
    threading.Thread(target=feedback_monitor_thread, daemon=True).start()

    # Start Multi-timezone Auto Scheduler (AST, IST, Sydney) in background
    threading.Thread(target=auto_scheduler_thread, daemon=True).start()

    # Start Real-time Alerts Push Notification check in background
    threading.Thread(target=realtime_alerts_monitor_thread, daemon=True).start()

    # Start Database Trigger Monitor in background
    threading.Thread(target=db_trigger_monitor_thread, daemon=True).start()

    
    # Check if build/web exists and copy feedback.html & admin.html there if so, to merge hosting automatically
    try:
        build_web_dir = os.path.join(os.path.dirname(__file__), "..", "build", "web")
        if os.path.exists(build_web_dir):
            import shutil
            shutil.copy(os.path.join(os.path.dirname(__file__), "feedback.html"), os.path.join(build_web_dir, "feedback.html"))
            shutil.copy(os.path.join(os.path.dirname(__file__), "..", "public", "admin.html"), os.path.join(build_web_dir, "admin.html"))
            print("[+] Merged hosting: copied feedback.html & admin.html to build/web/")
    except Exception as e:
        pass
    
    # Check dependencies warning
    if not REPORTLAB_AVAILABLE or not OPENPYXL_AVAILABLE:
        print("\n[!] WARNING: Optional rendering libraries are missing!")
        if not REPORTLAB_AVAILABLE:
            print("    - 'reportlab' is missing. PDF report generation will be skipped.")
        if not OPENPYXL_AVAILABLE:
            print("    - 'openpyxl' is missing. Excel report generation will be skipped.")
        print("    Please run: pip install -r requirements.txt\n")

    while True:
        print_menu()
        try:
            choice = input("Enter choice (1-10): ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\nExiting...")
            break

        if choice == "1":
            print("[*] Fetching clients from database...")
            url = f"{CONFIG['db_url']}/clients.json"
            try:
                req = urllib.request.Request(url)
                with urllib.request.urlopen(req, timeout=10) as res:
                    clients_data = json.loads(res.read().decode("utf-8"))
            except Exception as e:
                print(f"[-] Failed to fetch clients from database: {e}")
                continue

            clients_dict = normalize_to_dict(clients_data)
            if not clients_dict:
                print("[-] No clients found in the database.")
                continue

            client_list = []
            for cid, cinfo in clients_dict.items():
                if isinstance(cinfo, dict):
                    db_key = cinfo.get("db_key")
                    name = cinfo.get("name") or db_key or cid
                    if db_key:
                        client_list.append((db_key, name))

            if not client_list:
                print("[-] No valid client keys found in database.")
                continue

            print("\nAvailable Clients:")
            for idx, (db_key, name) in enumerate(client_list, 1):
                print(f"  {idx}. {name} ({db_key})")

            try:
                sel = input(f"Select client number (1-{len(client_list)}): ").strip()
                if sel:
                    sel_idx = int(sel) - 1
                    if 0 <= sel_idx < len(client_list):
                        selected_db_key, selected_name = client_list[sel_idx]
                        CONFIG["client_id"] = selected_db_key
                        print(f"[+] Client ID updated to: {selected_db_key} ({selected_name})")
                    else:
                        print("[-] Invalid selection number.")
            except ValueError:
                print("[-] Invalid input. Please enter a number.")

        elif choice == "2":
            print("\nAvailable Database Modes:")
            print("  1. Production")
            print("  2. Development")
            try:
                mode = input("Select mode number (1-2): ").strip()
                if mode == "1":
                    CONFIG["env_mode"] = "production"
                    print("[+] Environment Mode updated to: PRODUCTION")
                elif mode == "2":
                    CONFIG["env_mode"] = "development"
                    print("[+] Environment Mode updated to: DEVELOPMENT")
                else:
                    print("[-] Invalid selection. Mode unchanged.")
            except ValueError:
                print("[-] Invalid input. Mode unchanged.")
        elif choice == "3":
            print("[*] Testing SMTP connection...")
            if not CONFIG["smtp_email"] or not CONFIG["smtp_password"]:
                print("[-] SMTP credentials not configured in .env file.")
                continue
            try:
                server = smtplib.SMTP(CONFIG["smtp_server"], CONFIG["smtp_port"])
                server.starttls()
                server.login(CONFIG["smtp_email"], CONFIG["smtp_password"])
                server.quit()
                print("[+] SMTP Connection successful! Credentials are correct.")
            except Exception as e:
                print(f"[-] SMTP Connection test failed: {e}")
        elif choice == "4":
            execute_daily_reports_email()
        elif choice == "5":
            execute_missing_tanks_email()
        elif choice == "6":
            execute_alerts_notification_email()
        elif choice == "7":
            execute_weekly_reports_email()
        elif choice == "8":
            execute_weekly_alerts_email()
        elif choice == "9":
            print("\nFCM Trigger Submenu:")
            print(" 1. Send Reminder for Taking Tank Readings")
            print(" 2. Send Alert Notification (With Image)")
            print(" 3. Send Alert Notification (Without Image)")
            print(" 4. Send Reminder for Assessing a Group")
            print(" 5. Back")
            try:
                fcm_choice = input("Enter notification type (1-5): ").strip()
                if fcm_choice in ["1", "2", "3", "4"]:
                    trigger_fcm_notification(int(fcm_choice))
                elif fcm_choice == "5":
                    continue
                else:
                    print("[-] Invalid selection.")
            except ValueError:
                print("[-] Invalid input.")
        elif choice == "10":
            print("Exiting...")
            break
        else:
            print("[-] Invalid choice. Enter 1-10.")

if __name__ == "__main__":
    import sys
    import time as _pytime

    # Support --no-interactive flag for Cloud Run / Docker headless mode
    if "--no-interactive" in sys.argv or os.environ.get("NO_INTERACTIVE") == "1":
        load_env()
        _pytime_ref = _pytime.time()
        CONFIG["server_start_time"] = _pytime_ref

        print("[+] VAPLI Server starting in headless mode (Cloud Run / Docker)...")
        start_web_server()
        threading.Thread(target=feedback_monitor_thread, daemon=True).start()
        threading.Thread(target=auto_scheduler_thread, daemon=True).start()
        threading.Thread(target=realtime_alerts_monitor_thread, daemon=True).start()
        threading.Thread(target=db_trigger_monitor_thread, daemon=True).start()

        print(f"[+] All background workers started.")
        print(f"[+] Web server running on port {CONFIG['server_port']} — admin panel at /admin")
        print(f"[+] Health endpoint at /health")
        print(f"[+] Scheduler: AST {CONFIG['ast_reminder_hour']}:{CONFIG['ast_reminder_minute']:02d}, "
              f"IST {CONFIG['ist_reminder_hour']}:{CONFIG['ist_reminder_minute']:02d}, Sydney enabled")

        # Keep-alive main thread (Cloud Run needs the process to stay alive)
        try:
            while True:
                _pytime.sleep(60)
        except KeyboardInterrupt:
            print("\n[*] Shutting down gracefully...")
    else:
        main()

