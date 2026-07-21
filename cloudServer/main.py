import os
import re
import sys
import json
import smtplib
import hashlib
import tempfile
import urllib.request
import urllib.parse
from datetime import datetime, timedelta, timezone
from http.server import HTTPServer, BaseHTTPRequestHandler
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Attempt to load ReportLab and OpenPyXL for PDF/Excel generation
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Global configuration (defaults matched to Server/server.py config)
CONFIG = {
    "db_url": "https://dummy-firebase-project-id-default-rtdb.firebaseio.com",
    "web_api_key": "dummy-firebase-web-api-key",
    "smtp_server": "smtp.gmail.com",
    "smtp_port": 587,
    "smtp_email": "dummy@example.com",
    "smtp_password": "dummy-smtp-password",
    "client_id": "dummy_client_id",
    "env_mode": "production",
    "external_url": "https://vapli-cloud-server-43409271977.asia-east1.run.app",
    "secret_salt": "vapli_compliance_secure_salt"
}

def load_env():
    """Loads environment variables. Supports local fallback to VAPLI root .env/.env file"""
    env_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".env", ".env"))
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                m = re.match(r"^([^=]+)=(.*)$", line)
                if m:
                    key = m.group(1).strip()
                    val = m.group(2).strip()
                    if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
                        val = val[1:-1]
                    os.environ[key] = val

    # Read from environment
    if "FIREBASE_DATABASE_URL" in os.environ:
        CONFIG["db_url"] = os.environ["FIREBASE_DATABASE_URL"].rstrip("/")
    if "FIREBASE_WEB_API_KEY" in os.environ:
        CONFIG["web_api_key"] = os.environ["FIREBASE_WEB_API_KEY"]
    if "SMTP_SERVER" in os.environ:
        CONFIG["smtp_server"] = os.environ["SMTP_SERVER"]
    if "SMTP_PORT" in os.environ:
        CONFIG["smtp_port"] = int(os.environ["SMTP_PORT"])
    if "SMTP_EMAIL" in os.environ:
        CONFIG["smtp_email"] = os.environ["SMTP_EMAIL"]
    if "SMTP_PASSWORD" in os.environ:
        CONFIG["smtp_password"] = os.environ["SMTP_PASSWORD"]
    if "DEFAULT_CLIENT_ID" in os.environ:
        CONFIG["client_id"] = os.environ["DEFAULT_CLIENT_ID"]
    if "DEFAULT_ENV_MODE" in os.environ:
        CONFIG["env_mode"] = os.environ["DEFAULT_ENV_MODE"].lower()
    if "FEEDBACK_BASE_URL" in os.environ:
        CONFIG["external_url"] = os.environ["FEEDBACK_BASE_URL"].rstrip("/")
    elif "SERVER_EXTERNAL_URL" in os.environ:
        CONFIG["external_url"] = os.environ["SERVER_EXTERNAL_URL"].rstrip("/")

load_env()

# Global Firebase Admin Initializer (Lazy-Loaded only if needed)
FIREBASE_ADMIN_INITIALIZED = False
def initialize_firebase_admin():
    global FIREBASE_ADMIN_INITIALIZED
    if FIREBASE_ADMIN_INITIALIZED:
        return True
    try:
        import firebase_admin
        from firebase_admin import credentials
    except ImportError:
        print("[-] 'firebase-admin' library not installed. Skipping FCM operations.")
        return False

    service_account_path = os.path.join(os.path.dirname(__file__), "..", "service-account.json")
    if not os.path.exists(service_account_path):
        service_account_path = os.path.join(os.path.dirname(__file__), "service-account.json")

    if not os.path.exists(service_account_path):
        print("[-] Firebase Service Account key file ('service-account.json') not found.")
        return False

    try:
        cred = credentials.Certificate(service_account_path)
        firebase_admin.initialize_app(cred, {'databaseURL': CONFIG["db_url"]})
        FIREBASE_ADMIN_INITIALIZED = True
        print("[+] Firebase Admin SDK initialized successfully.")
        return True
    except Exception as e:
        print(f"[-] Failed to initialize Firebase Admin: {e}")
        return False

# ==============================================================================
# DATABASE UTILITIES
# ==============================================================================
def db_ref_path(raw_path):
    normalized = raw_path.lstrip("/")
    resolved = normalized
    global_paths = {"users", "clients"}
    scoped_paths = {"tanks", "tank_tree", "readings", "alerts", "completed_tasks", "violations", "settings", "reading_feedback"}
    if resolved:
        head = resolved.split("/")[0]
        if head not in global_paths and head in scoped_paths and CONFIG["client_id"]:
            resolved = f"{CONFIG['client_id']}/{resolved}"
    if CONFIG["env_mode"] == "development":
        resolved = f"testDB/{resolved}" if resolved else "testDB"
    return resolved

def fetch_db(path):
    full_path = db_ref_path(path)
    encoded_path = urllib.parse.quote(full_path, safe='/')
    url = f"{CONFIG['db_url']}/{encoded_path}.json"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as res:
            return json.loads(res.read().decode("utf-8"))
    except Exception as e:
        print(f"[-] Database fetch error on {path}: {e}")
        return None

def write_db(path, data, method="PUT"):
    full_path = db_ref_path(path)
    encoded_path = urllib.parse.quote(full_path, safe='/')
    url = f"{CONFIG['db_url']}/{encoded_path}.json"
    try:
        payload = json.dumps(data).encode("utf-8")
        req = urllib.request.Request(url, data=payload, method=method)
        req.add_header("Content-Type", "application/json")
        with urllib.request.urlopen(req, timeout=10) as res:
            return json.loads(res.read().decode("utf-8"))
    except Exception as e:
        print(f"[-] Database write error on {path}: {e}")
        return None

def normalize_to_dict(data):
    if not data:
        return {}
    if isinstance(data, dict):
        return data
    if isinstance(data, list):
        result = {}
        for idx, item in enumerate(data):
            if item is None:
                continue
            if isinstance(item, dict) and "id" in item:
                result[str(item["id"])] = item
            else:
                result[str(idx)] = item
        return result
    return {}

def parse_email_list(receivers_val):
    if not receivers_val:
        return []
    if isinstance(receivers_val, list):
        return [str(r).strip() for r in receivers_val if r and str(r).strip()]
    if isinstance(receivers_val, dict):
        return [str(v).strip() for v in receivers_val.values() if v and str(v).strip()]
    if isinstance(receivers_val, str):
        return [r.strip() for r in receivers_val.split(",") if r.strip()]
    return []

def get_client_name(client_id):
    try:
        clients_data = normalize_to_dict(fetch_db("clients"))
        for cid, cinfo in clients_data.items():
            if isinstance(cinfo, dict) and cinfo.get("db_key") == client_id:
                return cinfo.get("name") or client_id
    except:
        pass
    return client_id.upper().replace("_", " ")

# ==============================================================================
# TIME UTILITIES
# ==============================================================================
def get_ast_time():
    """Returns current Arabia Standard Time (AST, UTC+3)"""
    return datetime.now(timezone.utc) + timedelta(hours=3)

def get_ist_time():
    """Returns current India Standard Time (IST, UTC+5:30)"""
    return datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)

def parse_iso_datetime(iso_str):
    if not iso_str:
        return None
    try:
        clean_str = iso_str.rstrip("Z")
        if "." in clean_str:
            parts = clean_str.split(".")
            if len(parts[1]) > 6:
                clean_str = parts[0] + "." + parts[1][:6]
        return datetime.fromisoformat(clean_str)
    except:
        return None

# ==============================================================================
# EMAIL AND TEMPLATES
# ==============================================================================
def build_modern_template(title, header_title, body_content):
    client_name = get_client_name(CONFIG['client_id'])
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; background-color: #F8FAFC; color: #1E293B; margin:0; padding:20px; }}
            .container {{ max-width: 650px; margin: 0 auto; background: #FFFFFF; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.05); border: 1px solid #E2E8F0; }}
            .header {{ background: #1E3A8A; color: #FFFFFF; padding: 24px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 22px; font-weight: 700; letter-spacing: 0.5px; }}
            .header p {{ margin: 5px 0 0 0; font-size: 13px; opacity: 0.9; }}
            .content {{ padding: 30px; line-height: 1.6; font-size: 14px; }}
            .btn-cta {{ display: inline-block; background: #CB8C3E; color: #FFFFFF !important; text-decoration: none; padding: 12px 24px; font-weight: 600; border-radius: 6px; margin: 20px 0; text-align: center; font-size: 14px; }}
            .footer {{ background: #F1F5F9; color: #64748B; padding: 20px; text-align: center; font-size: 11px; border-top: 1px solid #E2E8F0; }}
            .report-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 12px; }}
            .report-table th {{ background: #475569; color: #FFFFFF; padding: 10px; text-align: left; font-weight: 600; border: 1px solid #CBD5E1; }}
            .report-table td {{ padding: 10px; border: 1px solid #CBD5E1; }}
            .badge {{ display: inline-block; padding: 2px 6px; font-size: 10px; font-weight: 700; border-radius: 4px; text-transform: uppercase; }}
            .badge-critical {{ background: #FEE2E2; color: #991B1B; }}
            .badge-warning {{ background: #FEF3C7; color: #92400E; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>{header_title}</h1>
                <p>{client_name} • VAPLI Engine</p>
            </div>
            <div class="content">
                {body_content}
            </div>
            <div class="footer">
                <p>This is an automated system notification dispatched by VAPLI Control Center.</p>
                <p>© {datetime.now().year} VAPLI Asset Integrity Management System.</p>
            </div>
        </div>
    </body>
    </html>
    """

def send_email(recipients, subject, html_body, attachments=None):
    if not CONFIG["smtp_email"] or not CONFIG["smtp_password"]:
        print("[-] SMTP is not configured. Email aborted.")
        return False
    if not recipients:
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = CONFIG["smtp_email"]
        msg["To"] = ", ".join(recipients)
        msg.attach(MIMEText(html_body, "html"))

        if attachments:
            for filepath in attachments:
                if os.path.exists(filepath):
                    with open(filepath, "rb") as f:
                        part = MIMEApplication(f.read(), Name=os.path.basename(filepath))
                    part['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
                    msg.attach(part)

        server = smtplib.SMTP(CONFIG["smtp_server"], CONFIG["smtp_port"])
        server.starttls()
        server.login(CONFIG["smtp_email"], CONFIG["smtp_password"])
        server.sendmail(CONFIG["smtp_email"], recipients, msg.as_string())
        server.quit()
        print(f"[+] Email sent successfully to: {recipients}")
        return True
    except Exception as e:
        print(f"[-] SMTP sending error: {e}")
        return False

# ==============================================================================
# FCM INTEGRATION
# ==============================================================================
def send_email_sent_fcm(email_type, recipient_count=0):
    if not initialize_firebase_admin():
        return
    try:
        from firebase_admin import messaging
        topic_name = CONFIG["client_id"]
        title = f"✉️ Email Dispatched: {email_type}"
        body = f"Report sent to {recipient_count} recipient(s) successfully."
        notification = messaging.Notification(title=title, body=body)
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": "email_sent",
            "email_type": email_type
        }
        message = messaging.Message(notification=notification, data=data_payload, topic=topic_name)
        messaging.send(message)
        print(f"[+] Email Sent FCM dispatched.")
    except Exception as e:
        print(f"[-] Email Sent FCM failed: {e}")

def trigger_fcm_tank_reminder(timezone_label="", target_date=None):
    if not initialize_firebase_admin():
        return
    try:
        from firebase_admin import messaging
        topic_name = CONFIG["client_id"]
        if not target_date:
            target_date = get_ast_time().strftime("%Y-%m-%d")

        tanks_data = normalize_to_dict(fetch_db("tanks"))
        readings_data = normalize_to_dict(fetch_db("readings"))
        tree_data = normalize_to_dict(fetch_db("tank_tree"))

        today_readings = set()
        for r in readings_data.values():
            if r and r.get("captured_at", "").split("T")[0] == target_date:
                today_readings.add(r.get("tank_id"))

        nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
        top_folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
        top_folders.sort(key=lambda x: x.get("order", 0))

        def get_folder_tanks(folder_id):
            tank_ids = []
            queue = [folder_id]
            while queue:
                curr = queue.pop(0)
                children = [n for n in nodes if n.get("parent_id") == curr]
                for c in children:
                    if c.get("type") == "leaf" and c.get("tank_id"):
                        tank_ids.append(c.get("tank_id"))
                    elif c.get("type") == "folder":
                        queue.append(c.get("id"))
            return tank_ids

        body_lines = []
        for folder in top_folders:
            t_ids = get_folder_tanks(folder.get("id"))
            if not t_ids:
                continue
            folder_tanks = [tanks_data[tid] for tid in t_ids if tid in tanks_data]
            if not folder_tanks:
                continue

            total_count = len(folder_tanks)
            inspected = [t for t in folder_tanks if t.get("id") in today_readings]
            uninspected = [t for t in folder_tanks if t.get("id") not in today_readings]
            inspected_count = len(inspected)
            percentage = (inspected_count / total_count * 100) if total_count > 0 else 0
            folder_name = folder.get("name", "Group")

            if inspected_count == 0:
                body_lines.append(f"{folder_name} has not taken Readings at all")
            elif percentage < 50:
                body_lines.append(f"{folder_name} has not been Taken reading except:")
                for t in uninspected[:3]:
                    body_lines.append(f"  • {t.get('tank_name')}")
                if len(uninspected) > 3:
                    body_lines.append(f"  • ... and {len(uninspected)-3} more")
            elif percentage < 100:
                body_lines.append(f"{folder_name} has been Taken reading except:")
                for t in uninspected[:3]:
                    body_lines.append(f"  • {t.get('tank_name')}")
                if len(uninspected) > 3:
                    body_lines.append(f"  • ... and {len(uninspected)-3} more")
            else:
                body_lines.append(f"{folder_name}: All readings complete ✓")

        if not body_lines:
            return

        tz_tag = f" [{timezone_label}]" if timezone_label else ""
        title = f"📋 3 PM Inspection Reminder{tz_tag}"
        body = "\n".join(body_lines)

        notification = messaging.Notification(title=title, body=body)
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": "1"
        }
        message = messaging.Message(notification=notification, data=data_payload, topic=topic_name)
        messaging.send(message)
        print(f"[+] Tank Reminder FCM dispatched successfully.")
    except Exception as e:
        print(f"[-] Tank Reminder FCM failed: {e}")

def trigger_batch_alerts_fcm(active_alerts):
    if not initialize_firebase_admin() or not active_alerts:
        return
    try:
        from firebase_admin import messaging
        topic_name = CONFIG["client_id"]
        critical_count = sum(1 for a in active_alerts if (a.get("constraint_severity") or a.get("severity") or "").lower() == "critical")
        warning_count = len(active_alerts) - critical_count

        title = f"🚨 VAPLI Batch Alerts Notification"
        body = f"Active violations found: {critical_count} Critical, {warning_count} Warning alerts require attention."

        notification = messaging.Notification(title=title, body=body)
        data_payload = {
            "client_id": topic_name,
            "click_action": "FLUTTER_NOTIFICATION_CLICK",
            "type": "batch_alerts",
            "critical_count": str(critical_count),
            "warning_count": str(warning_count)
        }
        message = messaging.Message(notification=notification, data=data_payload, topic=topic_name)
        messaging.send(message)
        print(f"[+] Batch Alerts FCM notification sent.")
    except Exception as e:
        print(f"[-] Batch Alerts FCM failed: {e}")

# ==============================================================================
# REPORT COMPILING LOGIC (DIRECT MATCH TO ORIGINAL SERVER)
# ==============================================================================
def format_reading_value(val):
    if val is None:
        return "-"
    if isinstance(val, dict):
        if "left" in val or "right" in val:
            return f"{val.get('left', '')} / {val.get('right', '')}"
        elif "value" in val:
            return str(val.get("value", "-"))
        else:
            return ", ".join(f"{k}: {v}" for k, v in val.items())
    return str(val)

def has_take_action_keyword(val):
    if val is None:
        return False
    val_str = str(val).lower()
    return "take action" in val_str or "not ok" in val_str

def get_selected_params_for_folder(folder_id, folder_tanks, report_format_configs):
    folder_config = report_format_configs.get(folder_id, {}) if report_format_configs else {}
    selected_params_map = folder_config.get("selected_params", {}) if folder_config else {}
    
    discovered = {}
    for tank in folder_tanks:
        props = tank.get("inspection_properties", [])
        if not isinstance(props, list):
            continue
        for prop in props:
            p_type = prop.get("type", "text")
            if p_type == "group":
                continue
            label = prop.get("label") or prop.get("name") or ""
            label = str(label).strip()
            if not label:
                continue
            
            options = prop.get("options", [])
            if isinstance(options, list):
                options = sorted([str(o).strip() for o in options if o is not None])
            else:
                options = []
            options_str = ",".join(options)
            
            db_key = f"{label}_{p_type}_{options_str}"
            fallback_key = f"{label}_{p_type}"
            
            if db_key not in discovered:
                discovered[db_key] = {
                    "db_key": db_key,
                    "fallback_key": fallback_key,
                    "label": label,
                    "type": p_type,
                    "prop_id": prop.get("id", ""),
                    "options": options,
                    "selected": True,
                    "order": 0
                }
                
    items = []
    max_order = 0
    for db_key, item in discovered.items():
        selected = True
        order = 0
        cfg = None
        if selected_params_map:
            if db_key in selected_params_map:
                cfg = selected_params_map[db_key]
            else:
                for k, v in selected_params_map.items():
                    if k.startswith(item["fallback_key"]):
                        cfg = v
                        break
        
        if cfg and isinstance(cfg, dict):
            selected = cfg.get("selected", True)
            order = int(cfg.get("order", 0))
            if selected and order > max_order:
                max_order = order
                
        items.append({
            "label": item["label"],
            "prop_id": item["prop_id"],
            "selected": selected,
            "order": order
        })
        
    selected_items = [i for i in items if i["selected"]]
    configured = sorted([i for i in selected_items if i["order"] > 0], key=lambda x: x["order"])
    unconfigured = sorted([i for i in selected_items if i["order"] == 0], key=lambda x: x["label"])
    next_order = max_order + 1
    for i in unconfigured:
        i["order"] = next_order
        next_order += 1
        
    return configured + unconfigured

def build_pdf_subtable(tanks, params, readings, active_alerts, cell_style, cell_style_bold, header_style):
    num_params = len(params)
    param_col_width = 370.0 / num_params if num_params > 0 else 370.0
    col_widths = [75, 95] + [param_col_width] * num_params
    
    headers = [
        Paragraph("Date Time", header_style),
        Paragraph("Asset Name", header_style)
    ]
    for p in params:
        headers.append(Paragraph(p["label"], header_style))
        
    table_data = [headers]
    table_commands = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F497D')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]
    
    row_idx = 1
    for tank in tanks:
        tid = tank.get("id")
        tank_code = tank.get("tank_code")
        tank_name = tank.get("tank_name", "Asset")
        display_name = f"{tank_name} ({tank_code})" if tank_code else tank_name
        
        reading = readings.get(tid)
        tank_severity = None
        for a in active_alerts:
            if a.get("tank_id") == tid:
                sev = (a.get("constraint_severity") or a.get("severity") or "warning").lower()
                if sev == "critical":
                    tank_severity = "critical"
                    break
                elif sev == "warning":
                    tank_severity = "warning"
                    
        default_row_color = '#FFFFFF' if row_idx % 2 == 1 else '#F8FAFC'
        
        if not reading:
            row = [
                Paragraph("-", cell_style),
                Paragraph(display_name, cell_style_bold),
                Paragraph("------- Readings not taken ------", cell_style_bold)
            ]
            for _ in range(num_params - 1):
                row.append(Paragraph("", cell_style))
            table_data.append(row)
            table_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#C8E6C9')))
            table_commands.append(('SPAN', (2, row_idx), (-1, row_idx)))
            table_commands.append(('ALIGN', (2, row_idx), (-1, row_idx), 'CENTER'))
        else:
            captured_at = reading.get("captured_at", "")
            try:
                dt = parse_iso_datetime(captured_at)
                dt_str = dt.strftime("%d/%m/%y %H:%M") if dt else captured_at.split(".")[0].replace("T", " ")[:14]
            except:
                dt_str = captured_at.split(".")[0].replace("T", " ")[:14]
                
            row = [
                Paragraph(dt_str, cell_style),
                Paragraph(display_name, cell_style_bold)
            ]
            
            insp_vals = reading.get("inspection_values", {})
            col_idx = 2
            for p in params:
                p_label = p["label"]
                p_id = p["prop_id"]
                val = insp_vals.get(p_label)
                if val is None:
                    val = insp_vals.get(p_id)
                val_str = format_reading_value(val)
                p_img = insp_vals.get(f"{p_id}__image_url") or insp_vals.get(f"{p_label}__image_url")
                cell_content = val_str
                if p_img:
                    cell_content += f'<br/><a href="{p_img}" color="#1565C0"><u>[Photo]</u></a>'
                    
                is_violated = False
                cell_severity = None
                for a in active_alerts:
                    if a.get("tank_id") == tid:
                        label = a.get("constraint_label") or a.get("param_label") or ""
                        if label.strip().lower() == p_label.strip().lower():
                            is_violated = True
                            cell_severity = (a.get("constraint_severity") or a.get("severity") or "critical").lower()
                            break
                            
                if not is_violated and has_take_action_keyword(val):
                    is_violated = True
                    cell_severity = "critical"
                    
                if is_violated:
                    row.append(Paragraph(cell_content, cell_style_bold))
                    color_hex = '#FFCDD2' if cell_severity == 'critical' else ('#FFF9C4' if cell_severity == 'warning' else '#BBDEFB')
                    table_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor(color_hex)))
                else:
                    row.append(Paragraph(cell_content, cell_style))
                    table_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor(default_row_color)))
                col_idx += 1
                
            table_data.append(row)
            if tank_severity:
                color_hex = '#FFCDD2' if tank_severity == 'critical' else ('#FFF9C4' if tank_severity == 'warning' else '#BBDEFB')
                table_commands.append(('BACKGROUND', (0, row_idx), (1, row_idx), colors.HexColor(color_hex)))
            else:
                table_commands.append(('BACKGROUND', (0, row_idx), (1, row_idx), colors.HexColor(default_row_color)))
        row_idx += 1
        
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle(table_commands))
    return t

def generate_inspection_pdf(filepath, date_str):
    if not REPORTLAB_AVAILABLE:
        return False
    try:
        tanks_data = normalize_to_dict(fetch_db("tanks"))
        readings_data = normalize_to_dict(fetch_db("readings"))
        tree_data = normalize_to_dict(fetch_db("tank_tree"))
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        report_format_configs = normalize_to_dict(fetch_db("settings/report_format"))
        
        today_readings = {}
        for r in readings_data.values():
            try:
                r_date = r.get("captured_at", "").split("T")[0]
                if r_date == date_str:
                    today_readings[r.get("tank_id")] = r
            except:
                continue

        active_alerts = [a for a in alerts_data.values() if not a.get("resolved", False) and a.get("status", "").lower() != "completed"]
        nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
        folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
        folders.sort(key=lambda x: x.get("order", 0))

        def get_folder_tanks(folder_id):
            tank_ids = []
            queue = [folder_id]
            while queue:
                curr = queue.pop(0)
                children = [n for n in nodes if n.get("parent_id") == curr]
                for c in children:
                    if c.get("type") == "leaf" and c.get("tank_id"):
                        tank_ids.append(c.get("tank_id"))
                    elif c.get("type") == "folder":
                        queue.append(c.get("id"))
            return tank_ids

        client_name = get_client_name(CONFIG['client_id'])
        generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")
        total_configured = len(tanks_data)
        inspected_tanks = [tid for tid in tanks_data.keys() if tid in today_readings]
        pending_tanks = [tank for tid, tank in tanks_data.items() if tid not in today_readings]
        pending_tanks.sort(key=lambda x: x.get("tank_name", ""))
        compliance_rate = (len(inspected_tanks) / total_configured * 100) if total_configured > 0 else 0.0

        story = []
        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#CB8C3E'), alignment=1, spaceAfter=8)
        meta_style = ParagraphStyle('DocMeta', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#555555'), alignment=1, spaceAfter=15)
        section_style = ParagraphStyle('SecTitle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#1A365D'), spaceBefore=10, spaceAfter=6)
        subsection_style = ParagraphStyle('SubSecTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9.5, textColor=colors.HexColor('#2B6CB0'), spaceBefore=6, spaceAfter=4)
        cell_style = ParagraphStyle('CellText', parent=styles['Normal'], fontName='Helvetica', fontSize=7, leading=9, textColor=colors.HexColor('#2D3748'))
        cell_style_bold = ParagraphStyle('CellTextBold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7, leading=9, textColor=colors.HexColor('#2D3748'))
        header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#FFFFFF'), alignment=1)

        story.append(Paragraph("VAPLI Inspection Telemetry Report", title_style))
        story.append(Paragraph(f"Client: {client_name} | Date: {date_str} | Generated: {generated_at}", meta_style))
        story.append(Spacer(1, 10))
        
        story.append(Paragraph("Professional Inspection Summary", section_style))
        story.append(Spacer(1, 5))
        
        stats_table_data = [
            [Paragraph("<b>Metric</b>", cell_style_bold), Paragraph("<b>Value / Count</b>", cell_style_bold)],
            [Paragraph("Total Configured Assets", cell_style), Paragraph(str(total_configured), cell_style_bold)],
            [Paragraph("Assets with Readings", cell_style), Paragraph(str(len(inspected_tanks)), cell_style_bold)],
            [Paragraph("Assets Pending Readings", cell_style), Paragraph(str(len(pending_tanks)), cell_style_bold)],
            [Paragraph("Compliance Rate", cell_style), Paragraph(f"{compliance_rate:.1f}%", cell_style_bold)],
            [Paragraph("Active Unresolved Alerts", cell_style), Paragraph(str(len(active_alerts)), cell_style_bold)]
        ]
        stats_table = Table(stats_table_data, colWidths=[200, 100])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#EDF2F7')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 15))

        story.append(Paragraph("Assets Pending Inspections (No Readings Today)", ParagraphStyle('PendingTitle', parent=section_style, textColor=colors.HexColor('#9B2C2C'))))
        story.append(Spacer(1, 5))
        
        if not pending_tanks:
            story.append(Paragraph("<b>No assets pending readings today. All recorded.</b>", ParagraphStyle('NoPending', parent=cell_style, textColor=colors.HexColor('#2F855A'))))
        else:
            pending_rows = []
            temp_row = []
            for t in pending_tanks:
                t_display = f"&bull; {t.get('tank_name')} ({t.get('tank_code', '')})"
                temp_row.append(Paragraph(t_display, cell_style))
                if len(temp_row) == 3:
                    pending_rows.append(temp_row)
                    temp_row = []
            if temp_row:
                while len(temp_row) < 3:
                    temp_row.append(Paragraph("", cell_style))
                pending_rows.append(temp_row)
                
            pending_table = Table(pending_rows, colWidths=[180, 180, 180])
            pending_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(pending_table)
            
        story.append(PageBreak())

        has_detailed_data = False
        for folder in folders:
            tank_ids = get_folder_tanks(folder.get("id"))
            if not tank_ids:
                continue
            folder_tanks = [tanks_data[tid] for tid in tank_ids if tid in tanks_data]
            if not folder_tanks:
                continue
            selected_params = get_selected_params_for_folder(folder.get("id"), folder_tanks, report_format_configs)
            if not selected_params:
                continue
                
            has_detailed_data = True
            normal_tanks_list = []
            violated_tanks_list = []
            
            for tank in folder_tanks:
                tid = tank.get("id")
                has_active_alert = any(a.get("tank_id") == tid for a in active_alerts)
                if has_active_alert:
                    violated_tanks_list.append(tank)
                else:
                    normal_tanks_list.append(tank)
                    
            folder_name = folder.get("name")
            if normal_tanks_list:
                story.append(Paragraph(f"{folder_name} — Normal Assets", subsection_style))
                story.append(Spacer(1, 3))
                story.append(build_pdf_subtable(normal_tanks_list, selected_params, today_readings, active_alerts, cell_style, cell_style_bold, header_style))
                story.append(Spacer(1, 15))
                
            if violated_tanks_list:
                story.append(Paragraph(f"{folder_name} — Violated Assets", ParagraphStyle('ViolatedSubSec', parent=subsection_style, textColor=colors.HexColor('#9B2C2C'))))
                story.append(Spacer(1, 3))
                story.append(build_pdf_subtable(violated_tanks_list, selected_params, today_readings, active_alerts, cell_style, cell_style_bold, header_style))
                story.append(Spacer(1, 15))
                
        if not has_detailed_data:
            story.append(Paragraph("No detailed asset records found matching report config filters.", cell_style))
            
        doc.build(story)
        return True
    except Exception as e:
        print(f"[-] Error generating PDF: {e}")
        return False

def generate_inspection_excel(filepath, date_str):
    if not OPENPYXL_AVAILABLE:
        return False
    try:
        tanks_data = normalize_to_dict(fetch_db("tanks"))
        readings_data = normalize_to_dict(fetch_db("readings"))
        tree_data = normalize_to_dict(fetch_db("tank_tree"))
        
        today_readings = {}
        for r in readings_data.values():
            if r and r.get("captured_at", "").split("T")[0] == date_str:
                today_readings[r.get("tank_id")] = r

        nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
        folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
        folders.sort(key=lambda x: x.get("order", 0))

        wb = Workbook()
        ws = wb.active
        ws.title = "Inspection_Report"

        title_font = Font(name="Calibri", size=16, bold=True, color="CB8C3E")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=12, bold=True, color="252830")
        header_fill = PatternFill(start_color="CB8C3E", end_color="CB8C3E", fill_type="solid")
        section_fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
        border_side = Side(border_style="thin", color="E0E0E0")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws["A1"] = "VAPLI Telemetry Summary"
        ws["A1"].font = title_font
        ws["A2"] = f"Client: {CONFIG['client_id'].upper()} | Date: {date_str} | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)

        curr_row = 4
        def get_folder_tanks(folder_id):
            tank_ids = []
            queue = [folder_id]
            while queue:
                curr = queue.pop(0)
                children = [n for n in nodes if n.get("parent_id") == curr]
                for c in children:
                    if c.get("type") == "leaf" and c.get("tank_id"):
                        tank_ids.append(c.get("tank_id"))
                    elif c.get("type") == "folder":
                        queue.append(c.get("id"))
            return tank_ids

        for folder in folders:
            tank_ids = get_folder_tanks(folder.get("id"))
            if not tank_ids:
                continue
            folder_tanks = [tanks_data[tid] for tid in tank_ids if tid in tanks_data]
            if not folder_tanks:
                continue

            ws.cell(row=curr_row, column=1, value=f"Group: {folder.get('name')}").font = section_font
            ws.cell(row=curr_row, column=1).fill = section_fill
            curr_row += 1

            group_params = []
            for tank in folder_tanks:
                props = tank.get("inspection_properties", [])
                for p in props:
                    label = p.get("label", "")
                    if label and label not in group_params:
                        group_params.append(label)

            headers = ["Asset Name"] + group_params + ["Capture Time"]
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=curr_row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
            curr_row += 1

            for tank in folder_tanks:
                tid = tank.get("id")
                reading = today_readings.get(tid)
                ws.cell(row=curr_row, column=1, value=tank.get("tank_name", "Asset")).border = thin_border
                
                if not reading:
                    ws.cell(row=curr_row, column=2, value="Readings not taken").border = thin_border
                    ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=len(group_params)+2)
                    for c in range(1, len(group_params) + 3):
                        ws.cell(row=curr_row, column=c).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        ws.cell(row=curr_row, column=c).border = thin_border
                else:
                    insp_vals = reading.get("inspection_values", {})
                    for col_idx, gp in enumerate(group_params, start=2):
                        val = insp_vals.get(gp, "-")
                        ws.cell(row=curr_row, column=col_idx, value=str(format_reading_value(val))).border = thin_border
                    
                    dt = parse_iso_datetime(reading.get("captured_at"))
                    time_str = dt.strftime("%I:%M %p") if dt else (reading.get("captured_at", "") or "").split("T")[-1][:5]
                    ws.cell(row=curr_row, column=len(group_params)+2, value=time_str).border = thin_border
                curr_row += 1
            curr_row += 2

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"[-] Error generating Excel: {e}")
        return False

def generate_alerts_pdf(filepath, date_str):
    if not REPORTLAB_AVAILABLE:
        return False
    try:
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        today_alerts = []
        for a in alerts_data.values():
            raw_time = a.get("captured_at") or a.get("timestamp") or ""
            if raw_time.split("T")[0] == date_str:
                today_alerts.append(a)

        client_name = get_client_name(CONFIG['client_id'])
        generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")

        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#D32F2F'), alignment=1, spaceAfter=8)
        meta_style = ParagraphStyle('DocMeta', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#555555'), alignment=1, spaceAfter=15)
        cell_style = ParagraphStyle('CellText', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#2D3748'))
        cell_style_bold = ParagraphStyle('CellTextBold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#2D3748'))
        header_cell_style = ParagraphStyle('HeaderCellText', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor('#FFFFFF'), alignment=1)

        story = []
        story.append(Paragraph(f"VAPLI Daily Alert Report", title_style))
        story.append(Paragraph(f"Client: {client_name} | Date: {date_str} | Generated: {generated_at}", meta_style))

        headers = [
            Paragraph("Time", header_cell_style),
            Paragraph("Asset Name", header_cell_style),
            Paragraph("Severity", header_cell_style),
            Paragraph("Parameter", header_cell_style),
            Paragraph("Value", header_cell_style),
            Paragraph("Alert Message", header_cell_style),
            Paragraph("IF-THEN Detail", header_cell_style)
        ]
        table_data = [headers]
        table_commands = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#9B2C2C')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]

        row_idx = 1
        for a in today_alerts:
            raw_time = a.get("captured_at") or a.get("timestamp") or ""
            dt = parse_iso_datetime(raw_time)
            time_str = dt.strftime("%I:%M %p") if dt else raw_time.split("T")[-1][:5]
            
            sev = (a.get("constraint_severity") or a.get("severity") or "WARNING").upper()
            label = a.get("constraint_label") or a.get("param_label") or "-"
            val = a.get("violated_value") or a.get("param_value") or "-"
            msg = a.get("message") or a.get("alert_title") or "-"
            if_then_val = a.get("if_then", "-") or "-"
            
            row = [
                Paragraph(time_str, cell_style),
                Paragraph(f"{a.get('tank_name')} ({a.get('tank_code', '')})", cell_style_bold),
                Paragraph(sev, cell_style_bold),
                Paragraph(label, cell_style),
                Paragraph(str(val), cell_style_bold),
                Paragraph(msg, cell_style),
                Paragraph(if_then_val, cell_style)
            ]
            table_data.append(row)
            
            sev_lower = sev.lower()
            color_hex = '#FFCDD2' if sev_lower == 'critical' else ('#FFF9C4' if sev_lower == 'warning' else '#EDF2F7')
            table_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(color_hex)))
            row_idx += 1

        col_widths = [55, 95, 55, 75, 45, 125, 90]
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle(table_commands))
        story.append(t)
        doc.build(story)
        return True
    except Exception as e:
        print(f"[-] Error generating Alerts PDF: {e}")
        return False

def generate_alerts_excel(filepath, date_str):
    if not OPENPYXL_AVAILABLE:
        return False
    try:
        alerts_data = normalize_to_dict(fetch_db("alerts"))
        today_alerts = []
        for a in alerts_data.values():
            raw_time = a.get("captured_at") or a.get("timestamp") or ""
            if raw_time.split("T")[0] == date_str:
                today_alerts.append(a)

        wb = Workbook()
        ws = wb.active
        ws.title = "Alerts_Summary"

        title_font = Font(name="Calibri", size=16, bold=True, color="D32F2F")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        
        red_header_fill = PatternFill(start_color="9B2C2C", end_color="9B2C2C", fill_type="solid")
        red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        gray_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
        )
        
        left_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')

        ws.append(["VAPLI Daily Alert Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.row_dimensions[1].height = 30
        
        client_name = get_client_name(CONFIG['client_id'])
        ws.append([f"Client: {client_name} | Date: {date_str} | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.append([])

        headers = ["Time", "Asset Name", "Asset Code", "Severity", "Parameter Checked", "Violated Value", "Alert Title", "Alert Message", "IF-THEN Detail"]
        ws.append(headers)
        ws.row_dimensions[4].height = 24
        
        for col_idx in range(1, 10):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = header_font
            cell.fill = red_header_fill
            cell.alignment = center_align
            cell.border = thin_border

        curr_row = 5
        for a in today_alerts:
            raw_time = a.get("captured_at") or a.get("timestamp") or ""
            dt = parse_iso_datetime(raw_time)
            time_str = dt.strftime("%I:%M %p") if dt else raw_time.split("T")[-1][:5]
            sev = (a.get("constraint_severity") or a.get("severity") or "WARNING").upper()
            label = a.get("constraint_label") or a.get("param_label") or "-"
            val = a.get("violated_value") or a.get("param_value") or "-"
            title = a.get("alert_title") or "-"
            msg = a.get("message") or "-"
            if_then_val = a.get("if_then", "-") or "-"

            ws.append([time_str, a.get("tank_name"), a.get("tank_code"), sev, label, val, title, msg, if_then_val])
            ws.row_dimensions[curr_row].height = 20
            sev_lower = sev.lower()
            fill_to_use = red_fill if sev_lower == "critical" else (yellow_fill if sev_lower == "warning" else gray_fill)
            
            for col_idx in range(1, 10):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = bold_font if col_idx in [2, 3, 4, 6] else regular_font
                cell.fill = fill_to_use
                cell.border = thin_border
                cell.alignment = center_align if col_idx in [1, 3, 4, 6] else left_align
            curr_row += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"[-] Error writing Alerts Excel: {e}")
        return False

# ==============================================================================
# BATCH WORKFLOW EXECUTORS
# ==============================================================================
def execute_daily_reports_email(target_date):
    print(f"[*] Starting Daily Reports compilation for {target_date}...")
    temp_dir = tempfile.gettempdir()
    
    insp_pdf_path = os.path.join(temp_dir, f"InspectionReport_{target_date}.pdf")
    insp_xls_path = os.path.join(temp_dir, f"InspectionReport_{target_date}.xlsx")
    alerts_pdf_path = os.path.join(temp_dir, f"AlertsReport_{target_date}.pdf")
    alerts_xls_path = os.path.join(temp_dir, f"AlertsReport_{target_date}.xlsx")

    generate_inspection_pdf(insp_pdf_path, target_date)
    generate_inspection_excel(insp_xls_path, target_date)
    generate_alerts_pdf(alerts_pdf_path, target_date)
    generate_alerts_excel(alerts_xls_path, target_date)

    settings = fetch_db("settings/Report_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))
    if not receivers:
        print("[-] No report receivers configured.")
        return False

    client_name = get_client_name(CONFIG['client_id'])
    subject = f"VAPLI Telemetry & Alert Reports - {client_name}"
    
    body = f"""
    <p>Dear Team,</p>
    <p>Please find attached the official VAPLI telemetry and alert inspection reports compiled for today.</p>
    <p><b>Summary Details:</b></p>
    <ul>
        <li><b>Client:</b> {client_name}</li>
        <li><b>Date:</b> {datetime.now().strftime('%A, %d %B %Y')}</li>
    </ul>
    <p>The attachments contain:
        <ol>
            <li><b>Today's Inspection Summary (PDF & Excel)</b></li>
            <li><b>Alert Violation Log (PDF & Excel)</b></li>
        </ol>
    </p>
    <p>Please inspect critical alerts immediately to protect asset lifecycles.</p>
    <br>
    <p>Warm regards,<br><b>VAPLI Automation Engine</b></p>
    """
    
    html_email = build_modern_template("Telemetry & Alerts Summary", "Inspection Summary Reports", body)
    attachments = [insp_pdf_path, insp_xls_path, alerts_pdf_path, alerts_xls_path]
    ok = send_email(receivers, subject, html_email, attachments)
    if ok:
        send_email_sent_fcm("Daily Inspection Report", len(receivers))

    for path in attachments:
        try:
            if os.path.exists(path):
                os.remove(path)
        except:
            pass
    return ok

def render_3_column_list(tanks):
    if not tanks:
        return "<p style='color:#777;'>None</p>"
    html = "<table style='width:100%; border:none; margin: 10px 0;'><tr>"
    col = 0
    for t in tanks:
        if col > 0 and col % 3 == 0:
            html += "</tr><tr>"
        html += f"<td style='width:33.3%; padding: 4px 8px; border:none; font-size:12px; color:#555;'>&bull; {t.get('tank_name')} ({t.get('tank_code', '')})</td>"
        col += 1
    while col % 3 != 0:
        html += "<td style='width:33.3%; border:none;'></td>"
        col += 1
    html += "</tr></table>"
    return html

def execute_missing_tanks_email(target_date):
    print(f"[*] Starting Missing Tanks scans for {target_date}...")
    tanks_data = normalize_to_dict(fetch_db("tanks"))
    readings_data = normalize_to_dict(fetch_db("readings"))
    tree_data = normalize_to_dict(fetch_db("tank_tree"))
    
    settings = fetch_db("settings/Missing Tanks_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))
    if not receivers:
        print("[-] No receivers for missing tanks.")
        return False

    today_readings = set()
    for r in readings_data.values():
        if r and r.get("captured_at", "").split("T")[0] == target_date:
            today_readings.add(r.get("tank_id"))

    nodes = list(tree_data.values()) if isinstance(tree_data, dict) else []
    top_folders = [n for n in nodes if n.get("type") == "folder" and not n.get("parent_id")]
    top_folders.sort(key=lambda x: x.get("order", 0))

    def get_folder_tanks(folder_id):
        tank_ids = []
        queue = [folder_id]
        while queue:
            curr = queue.pop(0)
            children = [n for n in nodes if n.get("parent_id") == curr]
            for c in children:
                if c.get("type") == "leaf" and c.get("tank_id"):
                    tank_ids.append(c.get("tank_id"))
                elif c.get("type") == "folder":
                    queue.append(c.get("id"))
        return tank_ids

    compliance_content = ""
    missing_tank_ids = []

    for folder in top_folders:
        t_ids = get_folder_tanks(folder.get("id"))
        if not t_ids:
            continue
        folder_tanks = [tanks_data[tid] for tid in t_ids if tid in tanks_data]
        if not folder_tanks:
            continue

        total_count = len(folder_tanks)
        inspected = [t for t in folder_tanks if t.get("id") in today_readings]
        uninspected = [t for t in folder_tanks if t.get("id") not in today_readings]
        inspected_count = len(inspected)
        missing_tank_ids.extend([t.get("id") for t in uninspected])
        percentage = (inspected_count / total_count) * 100 if total_count > 0 else 0

        compliance_content += f"<h3 style='margin-top:20px; color:#252830;'>Group: {folder.get('name')}</h3>"
        if inspected_count == 0:
            compliance_content += f"<p style='color:#D97706; font-weight:600;'>Group <b>{folder.get('name')}</b> has no tank recording Checkout for Readings</p>"
        elif percentage == 100.0:
            compliance_content += f"<p style='color:#2ECC71; font-weight:600;'>Group <b>{folder.get('name')}</b> is fully recorded.</p>"
        elif percentage < 50.0:
            compliance_content += f"<p>Group <b>{folder.get('name')}</b> has not been Recorded except these tanks</p>"
            compliance_content += render_3_column_list(inspected)
        else:
            compliance_content += f"<p>Group <b>{folder.get('name')}</b> has been Recorded except these tanks</p>"
            compliance_content += render_3_column_list(uninspected)

    if not missing_tank_ids:
        print("[*] 100% compliance achieved. No missing tanks emails required.")
        return True

    client_name = get_client_name(CONFIG['client_id'])
    encoded_tanks = ",".join(missing_tank_ids)
    feedback_base = CONFIG.get("external_url") or "https://helloworld-email-43409271977.asia-east1.run.app"

    for receiver in receivers:
        sig = hashlib.sha256((receiver + CONFIG["secret_salt"]).encode('utf-8')).hexdigest()
        # Points directly to our Cloud Run questioning server endpoint
        feedback_url = f"{feedback_base}/feedback?client_id={CONFIG['client_id']}&date={target_date}&tanks={encoded_tanks}&env={CONFIG['env_mode']}&submitted_by={urllib.parse.quote(receiver)}&sig={sig}"
        
        feedback_btn = f"""
        <div style='text-align: center;'>
            <a href='{feedback_url}' class='btn-cta' target='_blank'>Submit Reason for Missed Inspections</a>
        </div>
        """

        body = f"""
        <p>Dear Supervisor,</p>
        <p>Here is the compliance report highlighting lubrication inspection gaps for today.</p>
        <hr style='border:0; border-top:1px solid #E2E8F0;'>
        {compliance_content}
        <hr style='border:0; border-top:1px solid #E2E8F0; margin-top:20px;'>
        <p style='margin-top:20px;'><b>Inspection Action Required:</b></p>
        <p>To keep the automation metrics clean, please submit the reasons for any missed inspections today by clicking the button below:</p>
        {feedback_btn}
        <br>
        <p>Thank you,<br><b>VAPLI Compliance Monitor</b></p>
        """

        subject = f"VAPLI Inspection Compliance Gaps - {datetime.now().strftime('%d/%m/%Y')} ({client_name})"
        html_email = build_modern_template("Inspection Gaps & Form Link", "Daily Compliance Alert", body)
        ok = send_email([receiver], subject, html_email)
        if ok:
            send_email_sent_fcm("Missing Tanks Compliance Alert", 1)
    return True

def send_compliance_feedback_email(entry):
    settings = fetch_db("settings/Report_Recievers") or {}
    receivers = parse_email_list(settings.get("Emailids"))
    if not receivers:
        return False
        
    client_name = get_client_name(CONFIG['client_id'])
    
    tanks_content = ""
    tanks_by_group = entry.get("tanks_by_group", {})
    for group_name, tank_list in tanks_by_group.items():
        tanks_content += f"<h4 style='color:#CB8C3E; margin-bottom:5px;'>{group_name}</h4><ul style='margin-top:0; padding-left:20px;'>"
        for t in tank_list:
            tanks_content += f"<li>{t.get('name')} ({t.get('code', '')})</li>"
        tanks_content += "</ul>"

    body = f"""
    <p>Dear Administrator,</p>
    <p>Supervisor <b>{entry.get('submitted_by')}</b> has submitted the compliance feedback form explaining missed inspections.</p>
    <div style='background: #F1F5F9; border-left: 4px solid #CB8C3E; padding: 15px; margin: 20px 0; border-radius: 4px;'>
        <p style='margin:0;'><b>Date of Missed Readings:</b> {entry.get('date')}</p>
        <p style='margin:10px 0 0 0;'><b>Reason Category:</b> {entry.get('reason')}</p>
        <p style='margin:10px 0 0 0;'><b>Detailed Comments:</b> {entry.get('comments')}</p>
        <p style='margin:10px 0 0 0;'><b>Submitted At:</b> {entry.get('submitted_at')}</p>
    </div>
    <h3 style='color:#1E3A8A; margin-top:20px;'>Unattended Assets Covered:</h3>
    {tanks_content}
    <br>
    <p>Regards,<br><b>VAPLI Compliance Monitor</b></p>
    """
    
    subject = f"Missed Inspection Reason Submitted - {client_name}"
    html_email = build_modern_template("Supervisor Compliance Reason", "Missed Inspection Report", body)
    return send_email(receivers, subject, html_email)

# ==============================================================================
# WEB CONTROLLER (HTTP ROUTER)
# ==============================================================================
class CloudRunServer(BaseHTTPRequestHandler):
    
    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed_url.query)
        
        # 1. TRIGGER SCHEDULER DAILY WORKFLOW
        if parsed_url.path in ["/trigger", "/"]:
            # Evaluate time context in AST
            ast_now = get_ast_time()
            ast_date = ast_now.strftime("%Y-%m-%d")
            
            is_force = params.get("force", ["false"])[0].lower() == "true"
            is_weekend = ast_now.weekday() >= 5
            
            # Check if a specific client was requested, otherwise fetch all active clients
            requested_client = params.get("client_id", [""])[0]
            clients_to_process = []
            if requested_client:
                clients_to_process.append(requested_client)
            else:
                try:
                    # Temporary clear client ID to access the root clients path
                    CONFIG["client_id"] = ""
                    clients_data = normalize_to_dict(fetch_db("clients"))
                    for cid, cinfo in clients_data.items():
                        if isinstance(cinfo, dict):
                            db_key = cinfo.get("db_key")
                            if db_key:
                                clients_to_process.append(db_key)
                except Exception as e:
                    print(f"[-] Failed to fetch clients list from database: {e}")
            
            # Fallback if no clients resolved
            if not clients_to_process:
                clients_to_process.append("dummy_client_id")
                
            results = []
            for cid in clients_to_process:
                CONFIG["client_id"] = cid
                print(f"[*] Processing scheduler workflow for client: {cid}")
                
                status_msg = ""
                email_status = False
                
                # Execute daily compilation / missing checks
                if not is_weekend:
                    status_msg = f"Triggered Weekday workflow (AST Date: {ast_date})."
                    ok_reports = execute_daily_reports_email(ast_date)
                    ok_missing = execute_missing_tanks_email(ast_date)
                    email_status = ok_reports and ok_missing
                else:
                    status_msg = f"Triggered Weekend workflow (AST Date: {ast_date}). Weekly compilation is supported (see logs)."
                    ok_missing = execute_missing_tanks_email(ast_date)
                    email_status = ok_missing
                    
                # Process Batch Alerts & FCM (Consolidated Alerts)
                alerts_data = normalize_to_dict(fetch_db("alerts"))
                active_alerts = []
                for a_id, a in alerts_data.items():
                    if a and not a.get("resolved", False) and a.get("captured_at", "").split("T")[0] == ast_date:
                        active_alerts.append(a)
                
                if active_alerts:
                    trigger_batch_alerts_fcm(active_alerts)
                    
                trigger_fcm_tank_reminder("AST", ast_date)
                
                results.append({
                    "client_id": cid,
                    "email_sent_successfully": email_status,
                    "alerts_processed_count": len(active_alerts),
                    "message": status_msg
                })
            
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            
            response = {
                "status": "success",
                "mode": "weekend" if is_weekend else "weekday",
                "triggered_date": ast_date,
                "processed_clients": results
            }
            self.wfile.write(json.dumps(response).encode("utf-8"))
            return

        # 2. FEEDBACK UI PAGE GET
        elif parsed_url.path == "/feedback":
            client_id = params.get("client_id", [""])[0]
            date_str = params.get("date", [""])[0]
            tanks_param = params.get("tanks", [""])[0]
            submitted_by = params.get("submitted_by", [""])[0]
            sig = params.get("sig", [""])[0]
            
            # Validate link parameters
            if not client_id or not date_str or not tanks_param or not submitted_by or not sig:
                self.send_error_page("Unauthorized Access", "The compliance link parameters are invalid or missing.")
                return
                
            # Render aesthetic Dark Glassmorphic questionnaire web page
            html_page = self.get_feedback_form_html(client_id, date_str, tanks_param, submitted_by, sig)
            
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(html_page.encode("utf-8"))
            return
            
        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")
            
    def do_POST(self):
        parsed_url = urllib.parse.urlparse(self.path)
        if parsed_url.path == "/submit-feedback":
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length).decode('utf-8')
            fields = urllib.parse.parse_qs(post_data)
            
            client_id = fields.get("client_id", [""])[0]
            date_str = fields.get("date", [""])[0]
            tanks_param = fields.get("tanks", [""])[0]
            submitted_by = fields.get("submitted_by", [""])[0]
            sig = fields.get("sig", [""])[0]
            
            reason = fields.get("reason", [""])[0]
            comments = fields.get("comments", [""])[0]
            
            # Secure signature verification
            expected_sig = hashlib.sha256((submitted_by + CONFIG["secret_salt"]).encode('utf-8')).hexdigest()
            if sig != expected_sig:
                self.send_error_page("Fraudulent Submission Detected", "Access link signature is invalid. Form editing is restricted for security.")
                return
                
            # Process missing tanks list
            tank_ids = tanks_param.split(",")
            tanks_data = normalize_to_dict(fetch_db("tanks"))
            
            grouped_tanks = {}
            for tid in tank_ids:
                tank = tanks_data.get(tid)
                if not tank: continue
                g_name = Object.values(tank.get("Groups", {}))[0] if tank.get("Groups") else "Unassigned"
                if g_name not in grouped_tanks:
                    grouped_tanks[g_name] = []
                grouped_tanks[g_name].append({
                    "tank_id": tid,
                    "name": tank.get("tank_name", "Asset"),
                    "code": tank.get("tank_code", "")
                })
                
            feedback_entry = {
                "date": date_str,
                "reason": reason,
                "comments": comments,
                "submitted_by": submitted_by,
                "signature": sig,
                "submitted_at": datetime.now(timezone.utc).isoformat(),
                "tanks_by_group": grouped_tanks,
                "email_sent": False
            }
            
            # Save response to DB
            unique_id = int(datetime.now(timezone.utc).timestamp() * 1000)
            write_db(f"MissingReasons/{date_str}/{unique_id}", feedback_entry)
            
            # Send immediate notification email to administrators
            email_sent = send_compliance_feedback_email(feedback_entry)
            if email_sent:
                write_db(f"MissingReasons/{date_str}/{unique_id}/email_sent", True)

            # Render aesthetic success screen
            success_html = self.get_success_html()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(success_html.encode("utf-8"))
            return
            
        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")
            
    def send_error_page(self, title, message):
        self.send_response(403)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.end_headers()
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>{title}</title>
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background-color: #0E1012; color: #F0EEE9; padding: 20px; text-align: center; }}
                .card {{ max-width: 450px; margin: 100px auto; background: #141618; border: 1px solid #252830; border-radius: 12px; padding: 30px; box-shadow: 0 4px 15px rgba(0,0,0,0.25); }}
                h2 {{ color: #EF4444; margin: 0 0 15px 0; }}
                p {{ color: #8A8F9C; font-size: 14px; line-height: 1.5; }}
            </style>
        </head>
        <body>
            <div class="card">
                <h2>{title}</h2>
                <p>{message}</p>
            </div>
        </body>
        </html>
        """
        self.self.wfile.write(html.encode("utf-8"))

    def get_feedback_form_html(self, client_id, date_str, tanks_param, submitted_by, sig):
        tanks_data = normalize_to_dict(fetch_db("tanks"))
        tank_ids = tanks_param.split(",")
        
        grouped_tanks = {}
        for tid in tank_ids:
            tank = tanks_data.get(tid)
            if not tank: continue
            # Find group
            groups_dict = tank.get("Groups", {})
            g_name = list(groups_dict.values())[0] if groups_dict else "Unassigned"
            if g_name not in grouped_tanks:
                grouped_tanks[g_name] = []
            grouped_tanks[g_name].append(tank)
            
        tanks_html = ""
        for g_name, t_list in grouped_tanks.items():
            tanks_html += f"""
            <div class='group-container'>
                <h4>{g_name}</h4>
                <ul>
            """
            for t in t_list:
                tanks_html += f"<li>• {t.get('tank_name')} <span class='code'>({t.get('tank_code', '')})</span></li>"
            tanks_html += "</ul></div>"

        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>Missed Inspection Feedback</title>
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background-color: #0E1012; color: #F0EEE9; margin: 0; padding: 20px; }}
                .wrapper {{ max-width: 550px; margin: 40px auto; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .header h1 {{ color: #CB8C3E; font-size: 24px; margin: 0; font-weight: 700; }}
                .header p {{ color: #8A8F9C; font-size: 13px; margin: 5px 0 0 0; }}
                .card {{ background: #141618; border: 1px solid #252830; border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(0,0,0,0.25); }}
                .card h3 {{ margin: 0 0 15px 0; font-size: 16px; color: #F0EEE9; border-bottom: 1px solid #252830; padding-bottom: 8px; }}
                .group-container {{ margin-bottom: 12px; }}
                .group-container h4 {{ color: #CB8C3E; margin: 0 0 6px 0; font-size: 13px; }}
                .group-container ul {{ list-style: none; padding-left: 10px; margin: 0; }}
                .group-container li {{ font-size: 13px; margin-bottom: 4px; }}
                .code {{ color: #8A8F9C; font-size: 11px; }}
                .form-group {{ margin-bottom: 15px; }}
                .form-group label {{ display: block; font-size: 12px; color: #8A8F9C; margin-bottom: 6px; font-weight: 600; }}
                .form-group select, .form-group textarea {{ width: 100%; background: #1C1E22; border: 1px solid #252830; border-radius: 6px; color: #F0EEE9; padding: 10px; font-size: 13px; box-sizing: border-box; }}
                .form-group select:focus, .form-group textarea:focus {{ border-color: #CB8C3E; outline: none; }}
                .form-group textarea {{ height: 80px; resize: vertical; }}
                .readonly-box {{ padding: 10px; background: #1C1E22; border: 1px solid #252830; border-radius: 6px; color: #8A8F9C; font-size: 13px; }}
                .submit-btn {{ width: 100%; background-color: #CB8C3E; color: #FFFFFF; border: none; border-radius: 8px; padding: 14px; font-size: 14px; font-weight: 600; cursor: pointer; transition: background 0.2s; }}
                .submit-btn:hover {{ background-color: #B57B32; }}
            </style>
        </head>
        <body>
            <div class="wrapper">
                <div class="header">
                    <h1>Missed Inspection Questionnaire</h1>
                    <p>Reporting missed readings for client <b>{client_id.upper()}</b> on {date_str}</p>
                </div>
                
                <form action="/submit-feedback" method="POST">
                    <input type="hidden" name="client_id" value="{client_id}">
                    <input type="hidden" name="date" value="{date_str}">
                    <input type="hidden" name="tanks" value="{tanks_param}">
                    <input type="hidden" name="submitted_by" value="{submitted_by}">
                    <input type="hidden" name="sig" value="{sig}">
                    
                    <div class="card">
                        <h3>Unattended Assets</h3>
                        {tanks_html}
                    </div>

                    <div class="card">
                        <h3>Compliance Question</h3>
                        
                        <div class="form-group">
                            <label>Supervisor Email (Secured)</label>
                            <div class="readonly-box">{submitted_by}</div>
                        </div>

                        <div class="form-group">
                            <label>Reason for not taking readings today? *</label>
                            <select name="reason" required>
                                <option value="">-- Select Reason --</option>
                                <option value="Breakdown work">Breakdown work</option>
                                <option value="Leave">Leave</option>
                                <option value="Lack of Manpower">Lack of Manpower</option>
                                <option value="Stock Out / No Lube">Stock Out / No Lube</option>
                                <option value="Machine Offline / Under Maintenance">Machine Offline / Under Maintenance</option>
                                <option value="Other">Other</option>
                            </select>
                        </div>
                        
                        <div class="form-group">
                            <label>Explanation / Comments *</label>
                            <textarea name="comments" placeholder="Describe the reason in detail..." required></textarea>
                        </div>
                    </div>

                    <button type="submit" class="submit-btn">Submit Explanation</button>
                </form>
            </div>
        </body>
        </html>
        """

    def get_success_html(self):
        return """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>Feedback Submitted</title>
            <style>
                body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background-color: #0E1012; color: #F0EEE9; padding: 20px; text-align: center; }}
                .card {{ max-width: 450px; margin: 100px auto; background: #141618; border: 1px solid #252830; border-radius: 12px; padding: 40px 20px; box-shadow: 0 4px 15px rgba(0,0,0,0.25); }}
                .success-icon {{ width: 60px; height: 60px; background-color: #D5F5E3; color: #2ECC71; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 32px; font-weight: bold; margin: 0 auto 20px auto; }}
                h2 {{ color: #2ECC71; margin: 0 0 10px 0; }}
                p {{ color: #8A8F9C; font-size: 13px; line-height: 1.5; }}
            </style>
        </head>
        <body>
            <div class="card">
                <div class="success-icon">✓</div>
                <h2>Submission Successful</h2>
                <p>Explanation for missed readings has been recorded in the database, and the report has been dispatched to the administrators.</p>
                <p style="margin-top:20px; font-size:11px;">You may now safely close this tab.</p>
            </div>
        </body>
        </html>
        """

# ==============================================================================
# ENTRY POINT
# ==============================================================================
def run():
    port = int(os.environ.get("PORT", 8080))
    server_address = ("", port)
    httpd = HTTPServer(server_address, CloudRunServer)
    print(f"[*] Cloud Run Email & Questioning server running on port {port}...")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
    httpd.server_close()

if __name__ == "__main__":
    run()
