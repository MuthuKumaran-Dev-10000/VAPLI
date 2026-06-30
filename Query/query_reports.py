import os
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Database path configuration
BACKUP_PATH = r"C:\Users\muthu\Freelance\vapli\Backup-26-06-2026.json"
OUTPUT_DIR = r"C:\Users\muthu\Freelance\vapli\Query"

# Style configurations
FONT_NAME = "Segoe UI"

# Colors for headers
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Classic corporate dark blue
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")

# Colors for alternating rows
ALT_ROW_FILL = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")

# Colors for alerts (Classic Excel colors)
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
CRITICAL_FONT = Font(name=FONT_NAME, size=10, bold=True, color="9C0006")

WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
WARNING_FONT = Font(name=FONT_NAME, size=10, bold=True, color="9C6500")

NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)

# Grid borders
THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)

def format_datetime(dt_str):
    """Safely extracts date and time in a human-readable format YYYY-MM-DD HH:MM:SS"""
    if not dt_str:
        return ""
    try:
        clean = dt_str.split(".")[0].replace("T", " ")
        return clean
    except Exception:
        return dt_str

def format_reading_value(val):
    """Formats reading values, handling dictionaries (e.g. cooler temperatures)"""
    if val is None:
        return ""
    if isinstance(val, dict):
        return ", ".join(f"{k.capitalize()}: {v}" for k, v in val.items() if v != "")
    return str(val)

def check_value_violation(val, constraint):
    """Evaluates if a value violates a parameter's constraint"""
    op = constraint.get("op")
    c_val = constraint.get("value")
    
    if val is None or val == "":
        return False
        
    if isinstance(val, dict):
        # Dual inputs: if any of the values violates, return True
        return any(check_value_violation(v, constraint) for v in val.values())
        
    val_str = str(val).strip()
    c_val_str = str(c_val).strip()
    
    # Handle numerical comparison operators
    if op in [">", "<", ">=", "<="]:
        try:
            val_num = float(val_str)
            c_val_num = float(c_val_str)
            if op == ">": return val_num > c_val_num
            if op == "<": return val_num < c_val_num
            if op == ">=": return val_num >= c_val_num
            if op == "<=": return val_num <= c_val_num
        except ValueError:
            pass
    elif op == "==":
        return val_str.lower() == c_val_str.lower()
    elif op == "!=":
        return val_str.lower() != c_val_str.lower()
    elif op == "contains":
        return c_val_str.lower() in val_str.lower()
        
    return False

def has_take_action_keyword(val):
    """Checks if the value contains 'take action' or 'low level take action' keywords"""
    if val is None or val == "":
        return False
    if isinstance(val, dict):
        return any(has_take_action_keyword(v) for v in val.values())
    
    val_str = str(val).lower()
    return "take action" in val_str or "low level take action" in val_str

def sanitize_sheet_name(name, existing_names):
    """Sanitizes sheet names to conform to Excel rules (max 31 chars, no forbidden chars)"""
    for char in ['\\', '/', '?', '*', ':', '[', ']']:
        name = name.replace(char, '')
    name = name[:30] # Excel limit is 31
    if not name:
        name = "Sheet"
    
    base_name = name
    counter = 1
    while name.lower() in existing_names:
        suffix = f"_{counter}"
        name = base_name[:31 - len(suffix)] + suffix
        counter += 1
    existing_names.add(name.lower())
    return name

def apply_sheet_formatting(ws, include_header=True):
    """Applies column widths, alignment, and basic styles to a worksheet"""
    # Freeze pane below header
    if include_header:
        ws.freeze_panes = "A2"
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Add thin border to all cells
            if cell.row > 1 or not include_header:
                cell.border = THIN_BORDER
                if not cell.font or cell.font.name != FONT_NAME:
                    cell.font = NORMAL_FONT
            
            if cell.value:
                # Ignore lines starting with http or [Image when calculating widths
                lines = str(cell.value).split('\n')
                for line in lines:
                    line_str = line.strip()
                    if line_str.startswith("http") or line_str.startswith("[Image:"):
                        continue
                    max_len = max(max_len, len(line_str))
        
        # Set final column widths
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def process_client_reports(data, client_db_key, folder_name):
    print(f"\n[*] Processing client: {folder_name} ({client_db_key})")
    
    client_node = data.get(client_db_key, {})
    if not client_node:
        print(f"[-] Client node not found for {client_db_key}. Skipping.")
        return
        
    tanks = client_node.get("tanks", {})
    readings = client_node.get("readings", {})
    
    if not tanks:
        print(f"[-] No tanks found for {client_db_key}.")
        return

    # Create client directory
    client_dir = os.path.join(OUTPUT_DIR, folder_name)
    os.makedirs(client_dir, exist_ok=True)
    
    # --------------------------------------------------------------------------
    # Initialize Workbooks
    # --------------------------------------------------------------------------
    wb_normal = Workbook()
    wb_normal.remove(wb_normal.active) # Remove default sheet
    
    wb_alerts = Workbook()
    ws_all_alerts = wb_alerts.active
    ws_all_alerts.title = "All Violations"
    
    # Write Headers for Combined Alerts Sheet
    alert_headers = [
        "Date Time", "Asset Name", "Captured By", "Parameter", "Value", 
        "Violation Details", "Severity", "Main Image"
    ]
    for col_idx, h in enumerate(alert_headers, 1):
        cell = ws_all_alerts.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_all_alerts.row_dimensions[1].height = 26
    
    existing_normal_sheets = set()
    existing_alerts_sheets = set()
    
    all_violations_rows = []
    
    # Filter readings and index them by tank_id
    readings_by_tank = {}
    for rid, r in readings.items():
        tid = r.get("tank_id")
        if tid:
            readings_by_tank.setdefault(tid, []).append(r)
            
    # Sort readings inside each tank chronologically
    for tid in readings_by_tank:
        readings_by_tank[tid].sort(key=lambda x: x.get("captured_at", ""))
        
    # Iterate through each tank
    for tid, tank in tanks.items():
        tank_name = tank.get("tank_name") or "Asset"
        tank_code = tank.get("tank_code") or tid
        display_name = f"{tank_name} ({tank_code})" if tank_code else tank_name
        
        # Get parameter configurations
        params = tank.get("inspection_properties", [])
        if not isinstance(params, list):
            params = []
            
        # Get readings for this tank
        tank_readings = readings_by_tank.get(tid, [])
        if not tank_readings:
            continue
            
        print(f"  - Tank: {tank_name} ({len(tank_readings)} readings)")
        
        # Determine Sheet Name
        sheet_name_normal = sanitize_sheet_name(tank_name, existing_normal_sheets)
        sheet_name_alerts = sanitize_sheet_name(tank_name, existing_alerts_sheets)
        
        # Create sheet in Normal readings
        ws_normal = wb_normal.create_sheet(title=sheet_name_normal)
        normal_headers = ["Date Time", "Asset Name", "Main Image", "Captured By"] + [p.get("label") or p.get("id") for p in params]
        
        # Set up headers for Normal
        for col_idx, h in enumerate(normal_headers, 1):
            cell = ws_normal.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_normal.row_dimensions[1].height = 26
        
        # Create sheet in Alerts (will only add if there are violations)
        ws_alerts = None
        alerts_header_written = False
        
        # Fill rows for normal readings
        for row_idx, r in enumerate(tank_readings, 2):
            dt_str = format_datetime(r.get("captured_at"))
            cap_by = r.get("captured_by_name") or ""
            main_img = r.get("image_url") or ""
            
            row_data = [dt_str, display_name, main_img, cap_by]
            row_violations = [] # Keeps track of violations in this reading row
            
            # Map parameter values
            param_cells_info = [] # Store tuple of (col_idx, val, is_violated, severity_level, img_url)
            for idx, p in enumerate(params):
                pid = p.get("id")
                label = p.get("label")
                
                # Retrieve parameter value
                val = r.get("inspection_values", {}).get(label)
                if val is None:
                    val = r.get("inspection_values", {}).get(pid)
                    
                # Retrieve parameter-specific image URL
                p_img = r.get("inspection_values", {}).get(f"{pid}__image_url")
                if not p_img:
                    p_img = r.get("inspection_values", {}).get(f"{label}__image_url")
                    
                val_str = format_reading_value(val)
                cell_display_val = val_str
                if p_img:
                    cell_display_val += f"\n[Image: {p_img}]"
                
                # Check for alert triggers (Rules: constraints or Take Action keywords)
                is_violated = False
                severity_level = None
                violation_desc = ""
                
                # Check constraints
                constraints = p.get("constraints", [])
                for c in constraints:
                    if check_value_violation(val, c):
                        is_violated = True
                        severity_level = c.get("severity", "critical").lower()
                        violation_desc = c.get("message") or c.get("alert_title") or "Constraint Violated"
                        break
                        
                # Check keyword alert
                if not is_violated and has_take_action_keyword(val):
                    is_violated = True
                    severity_level = "critical" # Treat manual "Take Action" as critical severity
                    violation_desc = "Keyword 'Take Action' detected"
                    
                if is_violated:
                    row_violations.append({
                        "param_label": label or pid,
                        "value": val_str,
                        "p_img": p_img or "",
                        "violation_desc": violation_desc,
                        "severity": severity_level
                    })
                    
                param_cells_info.append((col_idx_param := (5 + idx), cell_display_val, is_violated, severity_level, p_img))
                row_data.append(cell_display_val)
                
            # Write row to Normal readings sheet
            ws_normal.row_dimensions[row_idx].height = 20
            # Apply zebra striping
            row_fill = ALT_ROW_FILL if row_idx % 2 == 1 else None
            
            for col_idx, cell_val in enumerate(row_data, 1):
                cell = ws_normal.cell(row=row_idx, column=col_idx, value=cell_val)
                if row_fill:
                    cell.fill = row_fill
                
                # Enable text wrapping on parameter cells
                if col_idx >= 5:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="center")
                    
            # Color-code violated cells in Normal readings sheet
            for col_idx, cell_val, is_violated, severity, p_img in param_cells_info:
                if is_violated:
                    cell = ws_normal.cell(row=row_idx, column=col_idx)
                    cell.font = CRITICAL_FONT if severity == "critical" else WARNING_FONT
                    cell.fill = CRITICAL_FILL if severity == "critical" else WARNING_FILL
            
            # If there are violations in this reading row, record it for Alerts
            if row_violations:
                # Add to the global violations registry
                for viol in row_violations:
                    all_violations_rows.append([
                        dt_str, display_name, cap_by, viol["param_label"], 
                        viol["value"] + (f"\n[Image: {viol['p_img']}]" if viol["p_img"] else ""),
                        viol["violation_desc"], viol["severity"].upper(), main_img
                    ])
                
                # Ensure sheet inside Alerts is initialized
                if ws_alerts is None:
                    ws_alerts = wb_alerts.create_sheet(title=sheet_name_alerts)
                    
                # Write header to Alerts tank sheet if it hasn't been written
                if not alerts_header_written:
                    for col_idx, h in enumerate(normal_headers, 1):
                        cell = ws_alerts.cell(row=1, column=col_idx, value=h)
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    ws_alerts.row_dimensions[1].height = 26
                    alerts_header_written = True
                    
                # Write row data to Alerts tank sheet
                alerts_row_idx = ws_alerts.max_row + 1
                ws_alerts.row_dimensions[alerts_row_idx].height = 20
                
                for col_idx, cell_val in enumerate(row_data, 1):
                    cell = ws_alerts.cell(row=alerts_row_idx, column=col_idx, value=cell_val)
                    if col_idx >= 5:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    else:
                        cell.alignment = Alignment(vertical="center")
                        
                # Color code specific violated parameter cells in Alerts tank sheet
                for col_idx, cell_val, is_violated, severity, p_img in param_cells_info:
                    if is_violated:
                        cell = ws_alerts.cell(row=alerts_row_idx, column=col_idx)
                        cell.font = CRITICAL_FONT if severity == "critical" else WARNING_FONT
                        cell.fill = CRITICAL_FILL if severity == "critical" else WARNING_FILL
        
        # Apply standard layout spacing and formatting to Normal readings sheets
        apply_sheet_formatting(ws_normal)
        
        # Apply formatting to the Alerts tank sheet (if it was created)
        if ws_alerts:
            apply_sheet_formatting(ws_alerts)

    # --------------------------------------------------------------------------
    # Save Normal readings.xlsx
    # --------------------------------------------------------------------------
    if len(wb_normal.sheetnames) > 0:
        normal_path = os.path.join(client_dir, "Normal readings.xlsx")
        wb_normal.save(normal_path)
        print(f"[+] Saved Normal readings to: {normal_path}")
    else:
        print("[-] No readings data found. 'Normal readings.xlsx' was not generated.")

    # --------------------------------------------------------------------------
    # Populating and Formatting Combined Alerts (Sheet 1)
    # --------------------------------------------------------------------------
    if all_violations_rows:
        # Sort violations chronologically
        all_violations_rows.sort(key=lambda x: x[0])
        
        for idx, row_data in enumerate(all_violations_rows, 2):
            ws_all_alerts.row_dimensions[idx].height = 22
            row_fill = ALT_ROW_FILL if idx % 2 == 1 else None
            
            for col_idx, cell_val in enumerate(row_data, 1):
                cell = ws_all_alerts.cell(row=idx, column=col_idx, value=cell_val)
                if row_fill:
                    cell.fill = row_fill
                
                # Check alignment and text wraps
                if col_idx in [5, 6]: # Value, Violation details
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="center")
                    
                # Format severity cell
                if col_idx == 7: # Severity column
                    sev = str(cell_val).lower()
                    cell.font = CRITICAL_FONT if "critical" in sev else WARNING_FONT
                    cell.fill = CRITICAL_FILL if "critical" in sev else WARNING_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Format the combined sheet
        apply_sheet_formatting(ws_all_alerts, include_header=True)
        
        alerts_path = os.path.join(client_dir, "Alerts.xlsx")
        wb_alerts.save(alerts_path)
        print(f"[+] Saved Alerts report to: {alerts_path}")
    else:
        # Write empty state indicator to Alerts sheet 1
        ws_all_alerts.cell(row=2, column=1, value="No active constraint violations or action items found.")
        ws_all_alerts.merge_cells("A2:H2")
        ws_all_alerts.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        ws_all_alerts.cell(row=2, column=1).font = Font(italic=True, name=FONT_NAME)
        
        alerts_path = os.path.join(client_dir, "Alerts.xlsx")
        wb_alerts.save(alerts_path)
        print(f"[+] Saved empty Alerts report to: {alerts_path}")

def main():
    print("="*60)
    print("      VAPLI Custom Telemetry & Alarms Spreadsheet Generator")
    print("="*60)
    
    if not os.path.exists(BACKUP_PATH):
        print(f"[-] ERROR: Backup JSON file not found at: {BACKUP_PATH}")
        return

    # Load Backup data
    print("[*] Loading database backup file...")
    start_time = datetime.now()
    with open(BACKUP_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print(f"[+] Loaded backup file in {(datetime.now() - start_time).total_seconds():.2f} seconds.")
    
    # Process Visy Paper (db_key: vsy_papers)
    process_client_reports(data, "vsy_papers", "Visy_Paper")
    
    # Process Muthu Stores (db_key: dummy_client_id)
    process_client_reports(data, "dummy_client_id", "muthu_Stores")
    
    print("\n" + "="*60)
    print("[+] Spreadsheet generation complete. Check 'Query/' subfolders.")
    print("="*60)

if __name__ == "__main__":
    main()
